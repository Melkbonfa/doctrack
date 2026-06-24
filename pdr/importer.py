"""
pdr/importer.py — Importação da planilha tratada (Lista Mestra PDR) para o banco.

Popula pdr_produtos / pdr_apresentacoes / pdr_documentos a partir de
pdr/data/Lista_Mestra_PDR_tratada.xlsx. Idempotente quando chamado pela rota
de reimport (que limpa as tabelas antes).
"""
import os
import warnings

import openpyxl

from models import db
from .models import Produto, Apresentacao, PdrDocumento, TIPOS_DOC

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
EXCEL_PATH = os.environ.get(
    "PDR_EXCEL_PATH", os.path.join(DATA_DIR, "Lista_Mestra_PDR_tratada.xlsx")
)

# Colunas da planilha por tipo de documento: (fase, status, codificação, versão)
DOC_MAP = {
    "especificacao":      ("ESPEC_FASE", "ESPEC_STATUS", None, "ESPEC_VERSAO"),
    "descritivo":         ("DESCRITIVO_FASE", "DESCRITIVO_STATUS", None, "DESCRITIVO_VERSAO"),
    "instrucao_trabalho": ("IT_FASE", "IT_STATUS", "IT_CODIFICACAO", "IT_VERSAO"),
    "manual":             ("MANUAL_FASE", "MANUAL_STATUS", None, "MANUAL_VERSAO"),
}


def _s(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none", "-", "—") else s


def importar_planilha():
    """Lê a planilha tratada e popula Produto/Apresentacao/PdrDocumento."""
    if not os.path.exists(EXCEL_PATH):
        print(f"[AVISO] Planilha tratada do PDR não encontrada: {EXCEL_PATH}")
        return 0
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Apresentações"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    def g(row, key):
        i = idx.get(key)
        return _s(row[i]) if i is not None and i < len(row) else ""

    produtos = {}
    n_apres = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome = g(row, "PRODUTO")
        if not nome:
            continue
        if nome not in produtos:
            p = Produto(nome=nome, sigla=g(row, "SIGLA"), linha=g(row, "LINHA") or "Extracta KITs")
            db.session.add(p)
            db.session.flush()
            produtos[nome] = p
        prod = produtos[nome]
        if not prod.sigla and g(row, "SIGLA"):
            prod.sigla = g(row, "SIGLA")

        apres = Apresentacao(
            produto_id=prod.id,
            apresentacao=g(row, "APRESENTACAO"),
            descricao=g(row, "DESCRICAO"),
            modelo=g(row, "MODELO"),
            sku=g(row, "SKU"),
            cadastro_protheus=g(row, "CADASTRO_PROTHEUS"),
            anvisa=g(row, "ANVISA"),
            numero_anvisa=g(row, "NUMERO_ANVISA"),
            fornecedor=g(row, "FORNECEDOR"),
            etiqueta=g(row, "ETIQUETA"),
            rotulagem=g(row, "ROTULAGEM"),
            planilha_rastreabilidade=g(row, "PLANILHA_RASTREABILIDADE"),
            observacoes=g(row, "OBSERVACOES"),
        )
        db.session.add(apres)
        db.session.flush()
        for tipo, (fase_c, stat_c, cod_c, ver_c) in DOC_MAP.items():
            db.session.add(PdrDocumento(
                apresentacao_id=apres.id, tipo=tipo,
                fase=g(row, fase_c), status=g(row, stat_c),
                codificacao=g(row, cod_c) if cod_c else "",
                versao=g(row, ver_c),
            ))
        n_apres += 1
    db.session.commit()
    print(f"[OK] PDR importado: {len(produtos)} produtos, {n_apres} apresentações, "
          f"{n_apres * 4} documentos.")
    return n_apres

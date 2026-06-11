"""
importar_entregaveis.py — Importação única da aba "Controle Projetos 2026"
da planilha files/Entregáveis - Engenharia (rev fev).xlsm para o doctrack.db.

Uso:
  ./venv/Scripts/python.exe importar_entregaveis.py            # importa (aborta se já houver dados)
  ./venv/Scripts/python.exe importar_entregaveis.py --substituir  # apaga projetos do ano e reimporta
  ./venv/Scripts/python.exe importar_entregaveis.py --dry-run     # só mostra o resumo, não grava
"""
import os
import sys
import argparse

XLSM = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                    "files", "Entregáveis - Engenharia (rev fev).xlsm")
ABA = "Controle Projetos 2026"
ANO = 2026
CATEGORIAS_VALIDAS = ["Produto", "Sistema", "Documentação", "Capacitação", "Marketing"]
# Cabeçalhos (linha 2) das colunas de metadados, em lower
META = {"ordem", "moscow", "prioridade", "entregáveis", "descrição", "consumível?",
        "cronograma mapeado", "sku", "lançamentos"}


def extrair_colunas(linha1, linha2, linha3):
    """Retorna [(idx0, tipo, categoria, responsaveis)] das colunas de entregável.

    Categoria vem da linha 1 com forward-fill; corta quando a categoria
    deixa de ser uma das válidas (ex.: '% janeiro') ou o tipo começa com '%'.
    """
    cols, categoria = [], None
    for i, tipo in enumerate(linha2):
        cab1 = linha1[i] if i < len(linha1) else None
        if isinstance(cab1, str) and cab1.strip():
            categoria = cab1.strip()
        nome = (tipo or "").strip() if isinstance(tipo, str) else ""
        if not nome or nome.lower() in META:
            continue
        nl = nome.lower()
        # Colunas de resumo/indicadores encerram a região de entregáveis
        if (nome.startswith("%") or nl.startswith("idp")
                or nl in ("idc", "andamento", "esperado", "executado",
                          "lançamento previsto", "lançamento real")):
            break
        if categoria not in CATEGORIAS_VALIDAS:
            # primeira coluna de % mensal encerra a região de entregáveis
            if categoria is not None:
                break
            continue
        resp = linha3[i] if i < len(linha3) else None
        resp = (resp or "").strip() if isinstance(resp, str) else ""
        # normaliza quebras de linha em nomes tipo "Software\nNeutro"
        nome = " ".join(nome.split())
        cols.append((i, nome, categoria, resp))
    return cols


def carregar_planilha():
    from openpyxl import load_workbook
    wb = load_workbook(XLSM, read_only=True, data_only=True)
    ws = wb[ABA]
    linhas = list(ws.iter_rows(values_only=True))
    return linhas


def indices_metadados(linha2):
    """Mapeia cabeçalho de metadado → índice de coluna (0-based)."""
    idx = {}
    for i, v in enumerate(linha2):
        if isinstance(v, str) and v.strip().lower() in META:
            # setdefault: a planilha repete "Entregáveis" como coluna de
            # contagem no fim — só a primeira ocorrência vale
            idx.setdefault(v.strip().lower(), i)
    return idx


def importar(substituir=False, dry_run=False):
    os.environ.setdefault("JWT_SECRET", "import-local-secret-32-chars-xxxxxxxx")
    from servidor import app
    from models import db, Projeto, Entregavel, converter_celula

    linhas = carregar_planilha()
    l1, l2, l3 = linhas[0], linhas[1], linhas[2]
    cols = extrair_colunas(l1, l2, l3)
    meta = indices_metadados(l2)
    ignoradas = 0
    projetos = []

    for row in linhas[3:]:
        nome = row[meta["entregáveis"]] if meta.get("entregáveis") is not None else None
        if not (isinstance(nome, str) and nome.strip()):
            if projetos:
                # bloco de projetos é contíguo: primeira linha em branco após
                # os dados encerra (abaixo há rodapé de "Equipe"/totais)
                break
            continue
        def mv(chave, default=""):
            i = meta.get(chave)
            v = row[i] if i is not None and i < len(row) else None
            return v if v is not None else default
        lanc = mv("lançamentos")
        if hasattr(lanc, "strftime"):
            lanc = lanc.strftime("%d/%m/%Y")
        p = dict(
            nome=" ".join(str(nome).split()),
            descricao=str(mv("descrição") or "").strip(),
            sku=str(mv("sku") or "").strip(),
            moscow=str(mv("moscow") or "").strip(),
            prioridade=int(mv("prioridade") or 0) if str(mv("prioridade") or "").strip().isdigit() else 0,
            consumivel=str(mv("consumível?") or "").strip().lower() == "sim",
            lancamento=str(lanc or "").strip(),
            entregaveis=[],
        )
        for (i, tipo, categoria, resp) in cols:
            valor = row[i] if i < len(row) else None
            status, pct = converter_celula(valor)
            if isinstance(valor, str) and valor.strip().startswith("#"):
                ignoradas += 1
            p["entregaveis"].append(dict(tipo=tipo, categoria=categoria,
                                         responsaveis=resp, status=status,
                                         percentual=pct))
        projetos.append(p)

    total_e = sum(len(p["entregaveis"]) for p in projetos)
    print(f"Planilha lida: {len(projetos)} projetos, {total_e} entregáveis, "
          f"{len(cols)} tipos de entregável, {ignoradas} células com lixo de fórmula.")
    for p in projetos:
        aplic = sum(1 for e in p["entregaveis"] if e["status"] != "na")
        print(f"  - {p['nome']}  [{p['moscow'] or '—'}]  {aplic} entregáveis aplicáveis")

    if dry_run:
        print("\n--dry-run: nada gravado.")
        return

    with app.app_context():
        db.create_all()
        existentes = Projeto.query.filter_by(ano=ANO).count()
        if existentes and not substituir:
            print(f"\nABORTADO: já existem {existentes} projetos de {ANO} no banco. "
                  f"Use --substituir para apagar e reimportar.")
            sys.exit(1)
        if existentes and substituir:
            Projeto.query.filter_by(ano=ANO).delete()
            db.session.commit()
            print(f"Projetos de {ANO} anteriores removidos.")
        for p in projetos:
            proj = Projeto(nome=p["nome"], descricao=p["descricao"], sku=p["sku"],
                           moscow=p["moscow"], prioridade=p["prioridade"],
                           consumivel=p["consumivel"], lancamento=p["lancamento"],
                           ano=ANO)
            db.session.add(proj)
            db.session.flush()
            for e in p["entregaveis"]:
                db.session.add(Entregavel(projeto_id=proj.id, **e,
                                          atualizado_por="importacao"))
        db.session.commit()
        print(f"\nOK: {len(projetos)} projetos e {total_e} entregáveis gravados no banco.")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--substituir", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    importar(substituir=args.substituir, dry_run=args.dry_run)

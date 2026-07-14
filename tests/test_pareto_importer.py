"""Testes do importador da análise Pareto (qtd_saidas + classe ABC por SKU)."""
import io

import openpyxl
import pytest


def _pareto_xlsx(linhas):
    """Monta um .xlsx no layout real: cabeçalho na 11ª linha, dados a partir da 12ª.

    `linhas` = lista de (sku_venda, qtd_saidas, classe).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pareto 80-20"
    # ruído nas 10 primeiras linhas (como na planilha original)
    ws.cell(row=1, column=1, value="ANÁLISE DE PARETO")
    header = ["Rank", "SKU Venda", "SKU Importação", "Equipamento", "Bloqueio",
              "Qtd Saídas", "% do Total", "% Acumulado", "Faz 80%?", "Classe ABC"]
    for j, h in enumerate(header, start=1):
        ws.cell(row=11, column=j, value=h)
    for i, (sku, qtd, classe) in enumerate(linhas):
        r = 12 + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=sku)        # SKU Venda (string, com zeros)
        ws.cell(row=r, column=4, value=f"EQUIP {sku}")
        ws.cell(row=r, column=6, value=qtd)
        ws.cell(row=r, column=10, value=classe)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_importa_casa_por_sku_normalizado(app):
    """SKU com zero à esquerda ('01.000404') casa com a planilha ('1.000404')."""
    from models import db, Equipamento
    from pareto_importer import importar_pareto

    with app.app_context():
        db.session.add(Equipamento(nome="Alfa", sku="01.000404"))
        db.session.add(Equipamento(nome="Beta", sku="01.000200"))  # zero à direita
        db.session.commit()

        xlsx = _pareto_xlsx([("1.000404", 107, "A"), ("01.000200", 3, "C")])
        rel = importar_pareto(file_bytes=xlsx, dryrun=False)

        assert rel["a_atualizar"] == 2
        assert rel["sem_match_n"] == 0
        alfa = Equipamento.query.filter_by(sku="01.000404").first()
        beta = Equipamento.query.filter_by(sku="01.000200").first()
        assert alfa.pareto_classe == "A" and alfa.qtd_saidas == 107
        assert beta.pareto_classe == "C" and beta.qtd_saidas == 3


def test_dryrun_nao_grava(app):
    from models import db, Equipamento
    from pareto_importer import importar_pareto

    with app.app_context():
        db.session.add(Equipamento(nome="Alfa", sku="01.000404"))
        db.session.commit()

        xlsx = _pareto_xlsx([("1.000404", 107, "A")])
        rel = importar_pareto(file_bytes=xlsx, dryrun=True)

        assert rel["aplicado"] is False
        assert rel["a_atualizar"] == 1
        alfa = Equipamento.query.filter_by(sku="01.000404").first()
        assert alfa.pareto_classe == "" and (alfa.qtd_saidas or 0) == 0


def test_sem_match_reportado(app):
    from models import db, Equipamento
    from pareto_importer import importar_pareto

    with app.app_context():
        db.session.add(Equipamento(nome="Alfa", sku="01.000404"))
        db.session.commit()

        xlsx = _pareto_xlsx([("1.000404", 10, "A"), ("09.999999", 5, "B")])
        rel = importar_pareto(file_bytes=xlsx, dryrun=False)

        assert rel["a_atualizar"] == 1
        assert rel["sem_match_n"] == 1
        assert rel["sem_match"][0]["sku"] == "09.999999"


def test_snapshot_zera_quem_saiu_do_ranking(app):
    """Equipamento que tinha classe/qtd mas não vem na planilha é zerado ao aplicar."""
    from models import db, Equipamento
    from pareto_importer import importar_pareto

    with app.app_context():
        db.session.add(Equipamento(nome="Alfa", sku="01.000404",
                                   pareto_classe="A", qtd_saidas=50))
        db.session.commit()

        xlsx = _pareto_xlsx([("1.000200", 8, "C")])  # Alfa não está mais
        rel = importar_pareto(file_bytes=xlsx, dryrun=False)

        assert rel["limpos_n"] == 1
        alfa = Equipamento.query.filter_by(sku="01.000404").first()
        assert alfa.pareto_classe == "" and alfa.qtd_saidas == 0

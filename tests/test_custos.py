"""Testes do módulo Custos.

Cobrem o que a revisão precisa poder confiar:

  * o corte de acesso (custo é dinheiro — gestão para cima);
  * o motor de cálculo, especialmente a separação NRE × COGS, o `aplicavel` que
    de fato exclui e a reserva incidindo só sobre a exposição cambial;
  * o versionamento automático da baseline;
  * a convenção de exportação (`;`, BOM, nome datado, mesmos filtros da tela);
  * o comportamento offline-first do câmbio.
"""
import io
import re

import pytest

BASE = "/custos/api"


# ── helpers ───────────────────────────────────────────────────────────────────

def _criar(client, headers, token, **extra):
    corpo = {
        "produto": "Analisador XPTO-1000",
        "sku": "01.000999",
        "fornecedor": "Allsheng",
        "tipo": "OEM",
        "incoterm": "FOB",
        "moeda_base": "USD",
        "status": "vigente",
        "valor_fob": 10000,
        "qtd_invoice": 3,
        "volume_projetado": 10,
        "preco_venda": 300000,
        "taxa_planejamento": 5,
        "reserva_cambial_pct": 10,
        "taxa_planejamento_justificativa": "Teto do ciclo.",
    }
    corpo.update(extra)
    res = client.post(f"{BASE}/composicoes", json=corpo, headers=headers(token))
    assert res.status_code == 201, res.get_json()
    return res.get_json()


def _lanc(comp, subcat):
    return next(l for l in comp["lancamentos"] if l["subcategoria"].startswith(subcat))


# ── acesso ────────────────────────────────────────────────────────────────────

@pytest.mark.parametrize("rota", [
    "/composicoes", "/meta", "/portfolio", "/saude", "/cotacoes",
    "/export/composicoes.csv", "/export/custos.xlsx",
])
def test_leitura_e_tecnico_nao_acessam(client, auth_headers, leitura_token,
                                       tecnico_token, rota):
    """Custo é dinheiro: nem leitura nem técnico enxergam qualquer rota."""
    for token in (leitura_token, tecnico_token):
        res = client.get(f"{BASE}{rota}", headers=auth_headers(token))
        assert res.status_code == 403, f"{rota} vazou para um perfil sem direito"


def test_sem_token_e_401(client):
    assert client.get(f"{BASE}/composicoes").status_code == 401


def test_gestor_acessa(client, auth_headers, gestor_token):
    res = client.get(f"{BASE}/composicoes", headers=auth_headers(gestor_token))
    assert res.status_code == 200
    assert res.get_json()["composicoes"] == []


def test_tecnico_nao_escreve(client, auth_headers, tecnico_token):
    res = client.post(f"{BASE}/composicoes", json={"produto": "X"},
                      headers=auth_headers(tecnico_token))
    assert res.status_code == 403


# ── criação e estrutura padrão ────────────────────────────────────────────────

def test_criar_gera_codigo_estrutura_e_baseline(client, auth_headers, gestor_token):
    c = _criar(client, auth_headers, gestor_token)
    assert re.match(r"^CC-\d{4}-\d{3}$", c["codigo"])
    subs = [l["subcategoria"] for l in c["lancamentos"]]
    # A estrutura de importação não pode nascer em branco: sem ela cada
    # composição seria redigitada diferente e o comparativo perderia sentido.
    assert "Mercadoria (FOB unitário)" in subs
    assert any(s.startswith("II —") for s in subs)
    assert "Reserva cambial" in subs
    # OEM ganha as linhas de hora; revenda não.
    assert any(l["tipo_calculo"] == "horas" for l in c["lancamentos"])
    assert c["versao"] == 1


def test_revenda_nao_recebe_linhas_de_hora(client, auth_headers, gestor_token):
    c = _criar(client, auth_headers, gestor_token, tipo="Revenda")
    assert not any(l["tipo_calculo"] == "horas" for l in c["lancamentos"])


def test_codigo_sequencial(client, auth_headers, gestor_token):
    a = _criar(client, auth_headers, gestor_token)
    b = _criar(client, auth_headers, gestor_token, produto="Outro")
    assert int(b["codigo"][-3:]) == int(a["codigo"][-3:]) + 1


@pytest.mark.parametrize("campo,valor", [
    ("tipo", "Consignado"), ("status", "publicada"),
    ("incoterm", "XPTO"), ("moeda_base", "JPY"),
])
def test_vocabulario_controlado(client, auth_headers, gestor_token, campo, valor):
    res = client.post(f"{BASE}/composicoes", json={"produto": "X", campo: valor},
                      headers=auth_headers(gestor_token))
    assert res.status_code == 400
    assert campo in res.get_json()["erro"]


def test_taxa_de_planejamento_precisa_ser_positiva(client, auth_headers, gestor_token):
    res = client.post(f"{BASE}/composicoes",
                      json={"produto": "X", "taxa_planejamento": 0},
                      headers=auth_headers(gestor_token))
    assert res.status_code == 400


# ── motor de cálculo ──────────────────────────────────────────────────────────

def test_nre_e_cogs_nao_se_misturam(client, auth_headers, gestor_token):
    """O erro estrutural que este módulo existe para evitar.

    NRE amortiza sobre o volume; COGS não. Somados na mesma coluna, o "custo do
    projeto" vira o custo de uma unidade e todo indicador derivado sai errado.
    """
    c = _criar(client, auth_headers, gestor_token,
               custo_hora_engenharia=100, volume_projetado=10)
    lh = next(l for l in c["lancamentos"] if l["perfil_hora"] == "eng")
    client.put(f"{BASE}/lancamentos/{lh['id']}", json={"horas": 200},
               headers=auth_headers(gestor_token))

    d = client.get(f"{BASE}/composicoes/{c['id']}",
                   headers=auth_headers(gestor_token)).get_json()
    k = d["calculo"]
    assert k["nre_realizado"] == 20000.0           # 200h × R$ 100
    assert k["nre_unitario"] == 2000.0             # amortizado em 10 unidades
    assert k["nre_realizado"] != k["cogs_orcado"]
    # O custo unitário leva o NRE amortizado, nunca o NRE cheio.
    assert k["custo_unitario"] == round(k["cogs_efetivo"] + 2000.0, 2)


def test_fob_e_percentuais(client, auth_headers, gestor_token):
    c = _criar(client, auth_headers, gestor_token)
    k = c["calculo"]
    linhas = {x["subcategoria"]: x for x in k["linhas"]}
    # FOB 10.000 × taxa 5 = 50.000
    assert linhas["Mercadoria (FOB unitário)"]["orcado"] == 50000.0
    # II 16% sobre o FOB em BRL
    assert linhas["II — Imposto de Importação"]["orcado"] == 8000.0
    # PIS 2,1%
    assert linhas["PIS/Pasep — Importação"]["orcado"] == 1050.0


def test_reserva_incide_so_sobre_exposicao_cambial(client, auth_headers, gestor_token):
    """Tributos são apurados em BRL sobre valor já convertido — não têm exposição.

    Incluí-los na base infla a reserva sem cobrir risco nenhum.
    """
    c = _criar(client, auth_headers, gestor_token)
    frete = _lanc(c, "Frete internacional")
    client.put(f"{BASE}/lancamentos/{frete['id']}",
               json={"valor_moeda": 500, "moeda": "USD"},
               headers=auth_headers(gestor_token))

    d = client.get(f"{BASE}/composicoes/{c['id']}",
                   headers=auth_headers(gestor_token)).get_json()
    k = d["calculo"]
    assert k["exposicao_cambial"] == 10500.0        # FOB 10.000 + frete 500
    reserva = next(x for x in k["linhas"] if x["subcategoria"] == "Reserva cambial")
    assert reserva["orcado"] == 5250.0              # 10% × 10.500 × taxa 5


def test_aplicavel_falso_exclui_de_fato(client, auth_headers, gestor_token):
    """Na planilha de origem esta coluna era decorativa: marcava "Não" e o valor
    entrava no total assim mesmo."""
    c = _criar(client, auth_headers, gestor_token)
    antes = c["calculo"]["cogs_orcado"]
    ii = _lanc(c, "II —")
    res = client.put(f"{BASE}/lancamentos/{ii['id']}", json={"aplicavel": False},
                     headers=auth_headers(gestor_token))
    depois = res.get_json()["calculo"]["cogs_orcado"]
    # Desconta exatamente o II. A reserva não se mexe: tributo não é exposição
    # cambial, então não entra na base dela.
    assert depois == round(antes - 8000.0, 2)


def test_margem_e_payback(client, auth_headers, gestor_token):
    c = _criar(client, auth_headers, gestor_token,
               custo_hora_engenharia=100, preco_venda=300000)
    k = c["calculo"]
    assert k["custo_unitario"] > 0
    assert k["margem_valor"] == round(300000 - k["custo_unitario"], 2)
    assert 0 < k["margem_pct"] < 1
    # Sem NRE lançado, o payback é zero — e não uma divisão que estoura.
    assert k["payback_unidades"] == 0.0


def test_sem_realizado_o_custo_usa_o_orcado(client, auth_headers, gestor_token):
    """Antes desta regra, uma composição recém-criada mostrava custo zero e
    margem de 100% — o tipo de número enganoso que o módulo existe para evitar."""
    c = _criar(client, auth_headers, gestor_token)
    k = c["calculo"]
    assert k["cogs_realizado"] is None      # nada lançado ainda
    assert k["desvio"] is None              # e nada a comparar
    assert k["cogs_efetivo"] == k["cogs_orcado"]
    assert k["custo_unitario"] == k["cogs_orcado"]
    assert k["margem_pct"] < 1.0


def test_desvio_compara_so_o_comparavel(client, auth_headers, gestor_token):
    """Subtrair o orçado inteiro de um realizado parcial daria um desvio
    negativo enorme enquanto a DI não chega."""
    c = _criar(client, auth_headers, gestor_token)
    ii = _lanc(c, "II —")
    res = client.put(f"{BASE}/lancamentos/{ii['id']}",
                     json={"realizado_valor_brl": 8600},
                     headers=auth_headers(gestor_token))
    k = res.get_json()["calculo"]
    # Só o II tem os dois lados: 8.600 − 8.000. O resto do COGS fica de fora.
    assert k["desvio"] == 600.0
    assert k["cogs_realizado"] == 8600.0
    assert k["desvio_pct"] == 0.075


def test_sem_preco_nao_calcula_margem(client, auth_headers, gestor_token):
    c = _criar(client, auth_headers, gestor_token, preco_venda=None)
    assert c["calculo"]["margem_pct"] is None
    assert c["calculo"]["payback_unidades"] is None


def test_realizado_gera_desvio(client, auth_headers, gestor_token):
    c = _criar(client, auth_headers, gestor_token)
    ii = _lanc(c, "II —")
    res = client.put(f"{BASE}/lancamentos/{ii['id']}",
                     json={"realizado_valor_brl": 9500},
                     headers=auth_headers(gestor_token))
    linha = next(x for x in res.get_json()["calculo"]["linhas"] if x["id"] == ii["id"])
    assert linha["realizado"] == 9500.0
    assert linha["desvio"] == 1500.0     # 9.500 realizado − 8.000 orçado


def test_mercadoria_realizada_usa_a_taxa_da_di(client, auth_headers, gestor_token):
    c = _criar(client, auth_headers, gestor_token)
    client.put(f"{BASE}/composicoes/{c['id']}", json={"taxa_realizada": 4.5},
               headers=auth_headers(gestor_token))
    d = client.get(f"{BASE}/composicoes/{c['id']}",
                   headers=auth_headers(gestor_token)).get_json()
    fob = next(x for x in d["calculo"]["linhas"]
               if x["subcategoria"] == "Mercadoria (FOB unitário)")
    assert fob["orcado"] == 50000.0       # 10.000 × 5,00 (planejamento)
    assert fob["realizado"] == 45000.0    # 10.000 × 4,50 (DI)
    assert fob["desvio"] == -5000.0


# ── versionamento ─────────────────────────────────────────────────────────────

def test_mexer_no_baseline_versiona(client, auth_headers, gestor_token):
    c = _criar(client, auth_headers, gestor_token)
    assert c["versao"] == 1
    res = client.put(f"{BASE}/composicoes/{c['id']}", json={"taxa_planejamento": 6},
                     headers=auth_headers(gestor_token))
    assert res.get_json()["versao"] == 2

    d = client.get(f"{BASE}/composicoes/{c['id']}",
                   headers=auth_headers(gestor_token)).get_json()
    assert [v["numero"] for v in d["versoes"]] == [2, 1]
    # A v1 guarda o cálculo de antes — é contra ela que o realizado é medido.
    assert d["versoes"][-1]["snapshot"]["taxa_planejamento"] == 5.0


def test_composicao_sem_baseline_nao_grava_alteracao_como_v1(app):
    """Uma composição criada fora da API (seed, importação) nasce com versao=1 e
    nenhuma linha em custo_versoes. A primeira alteração dela é v2, não v1."""
    from models import db
    from custos.models import Composicao
    from custos.routes import _congelar

    with app.app_context():
        c = Composicao(produto="Importada", taxa_planejamento=5, versao=1)
        db.session.add(c)
        db.session.flush()
        assert c.versoes == []
        v = _congelar(c, "Alteração de taxa_planejamento", "teste")
        assert v.numero == 2
        assert c.versao == 2


def test_campo_fora_do_baseline_nao_versiona(client, auth_headers, gestor_token):
    c = _criar(client, auth_headers, gestor_token)
    res = client.put(f"{BASE}/composicoes/{c['id']}", json={"fornecedor": "Outro"},
                     headers=auth_headers(gestor_token))
    assert res.get_json()["versao"] == 1


# ── lançamentos ───────────────────────────────────────────────────────────────

def test_crud_de_lancamento(client, auth_headers, gestor_token):
    c = _criar(client, auth_headers, gestor_token)
    n = len(c["lancamentos"])

    res = client.post(f"{BASE}/composicoes/{c['id']}/lancamentos", json={
        "natureza": "nre", "categoria": "Regulatório",
        "subcategoria": "Ensaio de compatibilidade eletromagnética",
        "tipo_calculo": "montante", "moeda": "BRL", "valor_moeda": 28500,
        "procedencia": "cotacao", "confianca": "media",
    }, headers=auth_headers(gestor_token))
    assert res.status_code == 201
    novo = res.get_json()["lancamento"]
    assert res.get_json()["calculo"]["nre_orcado"] == 28500.0

    client.put(f"{BASE}/lancamentos/{novo['id']}", json={"valor_moeda": 30000},
               headers=auth_headers(gestor_token))
    res = client.delete(f"{BASE}/lancamentos/{novo['id']}",
                        headers=auth_headers(gestor_token))
    assert res.status_code == 200
    assert res.get_json()["calculo"]["nre_orcado"] == 0.0

    d = client.get(f"{BASE}/composicoes/{c['id']}",
                   headers=auth_headers(gestor_token)).get_json()
    assert len(d["lancamentos"]) == n


def test_lancamento_sem_subcategoria_e_400(client, auth_headers, gestor_token):
    c = _criar(client, auth_headers, gestor_token)
    res = client.post(f"{BASE}/composicoes/{c['id']}/lancamentos", json={},
                      headers=auth_headers(gestor_token))
    assert res.status_code == 400


def test_lancamento_em_moeda_congela_a_taxa(client, auth_headers, gestor_token):
    c = _criar(client, auth_headers, gestor_token)
    frete = _lanc(c, "Frete internacional")
    res = client.put(f"{BASE}/lancamentos/{frete['id']}",
                     json={"valor_moeda": 100, "moeda": "USD"},
                     headers=auth_headers(gestor_token))
    l = res.get_json()["lancamento"]
    # Saber "quanto custou" exige saber "convertido a quanto".
    assert l["taxa_aplicada"] == 5.0
    assert l["valor_brl"] == 500.0


# ── saúde ─────────────────────────────────────────────────────────────────────

def test_saude_sem_composicoes(client, auth_headers, gestor_token):
    d = client.get(f"{BASE}/saude", headers=auth_headers(gestor_token)).get_json()
    assert d["total"] > 0
    assert 0 <= d["indice"] <= 100


def test_saude_acusa_hora_zerada_e_melhora_ao_corrigir(client, auth_headers, gestor_token):
    c = _criar(client, auth_headers, gestor_token, custo_hora_engenharia=0)
    d = client.get(f"{BASE}/saude", headers=auth_headers(gestor_token)).get_json()
    hora = next(x for x in d["verificacoes"] if x["id"] == "hora")
    assert not hora["ok"] and hora["severidade"] == "falha"
    antes = d["indice"]

    client.put(f"{BASE}/composicoes/{c['id']}", json={"custo_hora_engenharia": 150},
               headers=auth_headers(gestor_token))
    d2 = client.get(f"{BASE}/saude", headers=auth_headers(gestor_token)).get_json()
    assert next(x for x in d2["verificacoes"] if x["id"] == "hora")["ok"]
    assert d2["indice"] > antes


def test_saude_da_credito_proporcional(client, auth_headers, gestor_token):
    """Tudo-ou-nada faria "1 de 4 sem preço" pesar igual a "tudo quebrado"."""
    for i in range(4):
        _criar(client, auth_headers, gestor_token, produto=f"P{i}",
               preco_venda=None if i == 0 else 1000)
    d = client.get(f"{BASE}/saude", headers=auth_headers(gestor_token)).get_json()
    preco = next(x for x in d["verificacoes"] if x["id"] == "preco")
    assert preco["quantidade"] == 1
    assert preco["frac"] == 0.75


# ── portfólio ─────────────────────────────────────────────────────────────────

def test_portfolio_ordena_e_agrega_por_categoria(client, auth_headers, gestor_token):
    _criar(client, auth_headers, gestor_token, produto="Barato", valor_fob=1000)
    _criar(client, auth_headers, gestor_token, produto="Caro", valor_fob=50000)
    d = client.get(f"{BASE}/portfolio", headers=auth_headers(gestor_token)).get_json()
    assert [i["produto"] for i in d["itens"]] == ["Caro", "Barato"]
    assert {c["categoria"] for c in d["categorias"]} <= {
        "Parceiro OEM", "Integração Local", "Regulatório",
        "Logística e Tributos", "Comercial"}
    assert abs(sum(c["pct"] for c in d["categorias"]) - 1.0) < 0.01


# ── cotações e câmbio ─────────────────────────────────────────────────────────

def test_cotacao_manual_e_upsert(client, auth_headers, gestor_token):
    for valor in (5.10, 5.25):
        res = client.post(f"{BASE}/cotacoes",
                          json={"moeda": "USD", "data": "2026-08-03", "valor": valor},
                          headers=auth_headers(gestor_token))
        assert res.status_code == 201
    d = client.get(f"{BASE}/cotacoes", headers=auth_headers(gestor_token)).get_json()
    # Unicidade por (moeda, data, tipo): a segunda gravação atualiza, não duplica.
    assert len(d["cotacoes"]) == 1
    assert d["cotacoes"][0]["valor"] == 5.25
    assert d["referencia"] == 5.25


@pytest.mark.parametrize("corpo", [
    {"moeda": "JPY", "valor": 1}, {"moeda": "USD", "valor": 0},
    {"moeda": "USD", "valor": 5, "data": "03/08/2026"},
    {"moeda": "USD", "valor": 5, "tipo": "chute"},
])
def test_cotacao_invalida_e_400(client, auth_headers, gestor_token, corpo):
    res = client.post(f"{BASE}/cotacoes", json=corpo, headers=auth_headers(gestor_token))
    assert res.status_code == 400


def test_cambio_desligado_nao_chama_rede(app, monkeypatch):
    """Offline-first: sem rede o módulo continua utilizável com taxa manual."""
    from custos import cambio
    monkeypatch.setenv("DOCTRACK_CAMBIO", "0")
    assert cambio.habilitado() is False
    assert cambio.buscar_ptax("USD") == []
    assert cambio.sincronizar() == "desativado (DOCTRACK_CAMBIO=0)"


def test_cambio_falha_de_rede_nao_propaga(app, monkeypatch):
    """Falhar é normal e não pode derrubar a tarefa diária nem a tela."""
    from custos import cambio

    class _Boom:
        @staticmethod
        def get(*a, **k):
            raise OSError("rede fabril sem saída")

    monkeypatch.setenv("DOCTRACK_CAMBIO", "1")
    monkeypatch.setitem(__import__("sys").modules, "requests", _Boom)
    assert cambio.buscar_ptax("USD") == []


def test_cambio_consolida_boletins_do_dia(app, monkeypatch):
    """A PTAX publica várias vezes por dia; o que vale é o fechamento."""
    from custos import cambio

    class _Resp:
        status_code = 200
        @staticmethod
        def raise_for_status(): pass
        @staticmethod
        def json():
            return {"value": [
                {"cotacaoVenda": 5.10, "dataHoraCotacao": "2026-08-03 10:02:00.000"},
                {"cotacaoVenda": 5.19, "dataHoraCotacao": "2026-08-03 13:05:00.000"},
                {"cotacaoVenda": 5.22, "dataHoraCotacao": "2026-08-04 13:05:00.000"},
            ]}

    class _Req:
        @staticmethod
        def get(*a, **k):
            return _Resp()

    monkeypatch.setenv("DOCTRACK_CAMBIO", "1")
    monkeypatch.setitem(__import__("sys").modules, "requests", _Req)
    serie = cambio.buscar_ptax("USD")
    assert len(serie) == 2
    assert serie[0][1] == 5.19     # fechamento do dia 03, não o boletim das 10h
    assert serie[1][1] == 5.22


# ── exportações ───────────────────────────────────────────────────────────────

def test_export_csv_segue_a_convencao(client, auth_headers, gestor_token):
    _criar(client, auth_headers, gestor_token)
    res = client.get(f"{BASE}/export/composicoes.csv", headers=auth_headers(gestor_token))
    assert res.status_code == 200

    cd = res.headers["Content-Disposition"]
    assert re.search(r"\d{8}", cd), "nome do arquivo precisa ser datado"
    assert res.data.startswith(b"\xef\xbb\xbf"), "sem BOM o Excel pt-BR lê como Latin-1"
    texto = res.data.decode("utf-8-sig")
    assert ";" in texto.splitlines()[0]
    assert "," not in texto.splitlines()[0]
    assert "Analisador XPTO-1000" in texto


def test_export_respeita_os_filtros_da_tela(client, auth_headers, gestor_token):
    _criar(client, auth_headers, gestor_token, produto="Vigente", status="vigente")
    _criar(client, auth_headers, gestor_token, produto="Rascunho", status="rascunho")
    res = client.get(f"{BASE}/export/composicoes.csv?status=vigente",
                     headers=auth_headers(gestor_token))
    texto = res.data.decode("utf-8-sig")
    assert "Vigente" in texto and "Rascunho" not in texto


def test_export_com_filtro_invalido_e_400(client, auth_headers, gestor_token):
    res = client.get(f"{BASE}/export/composicoes.csv?status=inexistente",
                     headers=auth_headers(gestor_token))
    assert res.status_code == 400


def test_export_xlsx_tem_tres_abas(client, auth_headers, gestor_token):
    import openpyxl
    _criar(client, auth_headers, gestor_token)
    res = client.get(f"{BASE}/export/custos.xlsx", headers=auth_headers(gestor_token))
    assert res.status_code == 200
    assert re.search(r"\d{8}", res.headers["Content-Disposition"])

    wb = openpyxl.load_workbook(io.BytesIO(res.data))
    assert wb.sheetnames == ["Composicoes", "Lancamentos", "Cotacoes"]
    assert wb["Composicoes"].max_row == 2
    # Uma linha por lançamento da composição criada.
    assert wb["Lancamentos"].max_row == 13


# ── arquivamento ──────────────────────────────────────────────────────────────

def test_arquivar_some_da_lista(client, auth_headers, gestor_token):
    c = _criar(client, auth_headers, gestor_token)
    client.delete(f"{BASE}/composicoes/{c['id']}", headers=auth_headers(gestor_token))
    d = client.get(f"{BASE}/composicoes", headers=auth_headers(gestor_token)).get_json()
    assert d["composicoes"] == []


# ── página ────────────────────────────────────────────────────────────────────

def test_pagina_responde(client):
    """A página valida o token no front, como o PDR e o hub; os dados é que são
    barrados em cada rota /custos/api/*."""
    res = client.get("/custos/")
    assert res.status_code == 200
    assert b"custos.js" in res.data

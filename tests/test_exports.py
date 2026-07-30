"""Testes das exportações — formato, filtros e permissão.

Cobre as três regressões que os exports carregavam em comum:

1. O CSV de consumíveis saía com vírgula, enquanto os outros usam ponto-e-vírgula
   (o Excel pt-BR abre o arquivo com vírgula numa coluna só).
2. Consumíveis, projetos e PDR ignoravam os filtros da tela e devolviam sempre a
   base inteira.
3. Documentos, equipamentos e consumíveis exportavam a base completa para
   qualquer login, inclusive o papel `leitura`.
"""
import csv
import io


def _texto(res):
    """Corpo do CSV sem o BOM que o Excel exige para ler UTF-8."""
    return res.data.decode("utf-8-sig")


def _linhas(res):
    return list(csv.reader(io.StringIO(_texto(res)), delimiter=";"))


def _cabecalho(client, token):
    return {"Authorization": f"Bearer {token}"}


# ── Formato: ponto-e-vírgula em todos os CSVs ────────────────────────────────

def test_todos_os_csv_usam_ponto_e_virgula(client, gestor_token):
    """O separador é o mesmo em todos os módulos.

    Um export por módulo com convenção diferente é o tipo de coisa que só
    aparece na máquina de quem abre o arquivo.
    """
    h = _cabecalho(client, gestor_token)
    for rota in ("/api/documentos/export",
                 "/api/equipamentos/export",
                 "/api/consumiveis/export"):
        res = client.get(rota, headers=h)
        assert res.status_code == 200, rota
        primeira = _texto(res).splitlines()[0]
        assert ";" in primeira, f"{rota} não usa ';' como separador"
        # Uma coluna só significa que o separador não foi reconhecido.
        assert len(_linhas(res)[0]) > 1, rota


def test_csv_sai_com_bom_para_o_excel(client, gestor_token):
    """Sem o BOM o Excel lê UTF-8 como Latin-1 e os acentos viram lixo."""
    res = client.get("/api/consumiveis/export", headers=_cabecalho(client, gestor_token))
    assert res.status_code == 200
    assert res.data.startswith(b"\xef\xbb\xbf")


def test_nome_do_arquivo_tem_data(client, gestor_token):
    """Sem data no nome, cada export sobrescreve o anterior em Downloads."""
    import re
    h = _cabecalho(client, gestor_token)
    for rota in ("/api/documentos/export",
                 "/api/equipamentos/export",
                 "/api/consumiveis/export"):
        res = client.get(rota, headers=h)
        cd = res.headers.get("Content-Disposition", "")
        assert re.search(r"\d{8}", cd), f"{rota} exporta sem data no nome: {cd}"


# ── Filtros: consumíveis ─────────────────────────────────────────────────────

def _semear_consumiveis(app):
    from models import db, Consumivel, TipoConsumivel
    with app.app_context():
        t_agulha = TipoConsumivel(nome="Agulha")
        t_tubo = TipoConsumivel(nome="Tubo")
        db.session.add_all([t_agulha, t_tubo])
        db.session.flush()
        db.session.add_all([
            Consumivel(nome="Agulha 21G", sku="AG-21", tipo_id=t_agulha.id, ativo=True),
            Consumivel(nome="Agulha 23G", sku="", tipo_id=t_agulha.id,
                       pendente_sku=True, ativo=True),
            Consumivel(nome="Tubo EDTA", sku="TB-01", tipo_id=t_tubo.id, ativo=True),
        ])
        db.session.commit()


def test_export_consumiveis_filtra_por_tipo(client, app, gestor_token):
    _semear_consumiveis(app)
    res = client.get("/api/consumiveis/export?tipo=Agulha",
                     headers=_cabecalho(client, gestor_token))
    assert res.status_code == 200
    nomes = [l[0] for l in _linhas(res)[1:] if l]
    assert sorted(nomes) == ["Agulha 21G", "Agulha 23G"]


def test_export_consumiveis_filtra_por_busca(client, app, gestor_token):
    _semear_consumiveis(app)
    res = client.get("/api/consumiveis/export?q=tubo",
                     headers=_cabecalho(client, gestor_token))
    nomes = [l[0] for l in _linhas(res)[1:] if l]
    assert nomes == ["Tubo EDTA"]


def test_export_consumiveis_filtra_pendentes(client, app, gestor_token):
    _semear_consumiveis(app)
    res = client.get("/api/consumiveis/export?pendente=1",
                     headers=_cabecalho(client, gestor_token))
    nomes = [l[0] for l in _linhas(res)[1:] if l]
    assert nomes == ["Agulha 23G"]


def test_export_consumiveis_sem_filtro_traz_tudo(client, app, gestor_token):
    _semear_consumiveis(app)
    res = client.get("/api/consumiveis/export",
                     headers=_cabecalho(client, gestor_token))
    nomes = [l[0] for l in _linhas(res)[1:] if l]
    assert len(nomes) == 3


# ── Filtros: projetos (mesmo helper da listagem) ─────────────────────────────

def _semear_projetos(app):
    from models import db, Projeto
    with app.app_context():
        db.session.add_all([
            Projeto(nome="Alfa", ano=2026, status="execucao", ativo=True),
            Projeto(nome="Beta", ano=2026, status="concluido", ativo=True),
        ])
        db.session.commit()


def test_export_projetos_respeita_status(client, app, gestor_token):
    """O export usa `_filtrar_projetos`, o mesmo da listagem."""
    from openpyxl import load_workbook
    _semear_projetos(app)
    res = client.get("/api/entregaveis/export?status=execucao",
                     headers=_cabecalho(client, gestor_token))
    assert res.status_code == 200
    wb = load_workbook(io.BytesIO(res.data))
    nomes = [c.value for c in wb["Entregáveis 2026"]["A"][1:] if c.value]
    assert nomes == ["Alfa"]


def test_export_projetos_respeita_busca(client, app, gestor_token):
    from openpyxl import load_workbook
    _semear_projetos(app)
    res = client.get("/api/entregaveis/export?busca=Beta",
                     headers=_cabecalho(client, gestor_token))
    wb = load_workbook(io.BytesIO(res.data))
    nomes = [c.value for c in wb["Entregáveis 2026"]["A"][1:] if c.value]
    assert nomes == ["Beta"]


def test_export_projetos_status_invalido_da_400(client, app, gestor_token):
    """Mesma validação da listagem — o helper devolve o 400 pronto."""
    _semear_projetos(app)
    res = client.get("/api/entregaveis/export?status=inexistente",
                     headers=_cabecalho(client, gestor_token))
    assert res.status_code == 400


def test_export_projetos_sem_filtro_traz_ativos(client, app, gestor_token):
    from openpyxl import load_workbook
    _semear_projetos(app)
    res = client.get("/api/entregaveis/export",
                     headers=_cabecalho(client, gestor_token))
    wb = load_workbook(io.BytesIO(res.data))
    nomes = [c.value for c in wb["Entregáveis 2026"]["A"][1:] if c.value]
    assert sorted(nomes) == ["Alfa", "Beta"]


# ── Filtros: PDR ─────────────────────────────────────────────────────────────

def _semear_pdr(app):
    from models import db
    from pdr.models import Produto, Apresentacao
    with app.app_context():
        extracta = Produto(nome="Extracta A", linha="Extracta KITs", sigla="EXA", ativo=True)
        outra = Produto(nome="Outro B", linha="Linha B", sigla="OTB", ativo=True)
        db.session.add_all([extracta, outra])
        db.session.flush()
        db.session.add_all([
            Apresentacao(produto_id=extracta.id, apresentacao="96 testes", sku="EXA-96",
                         fornecedor="Loccus", anvisa="Sim", ativo=True),
            Apresentacao(produto_id=outra.id, apresentacao="48 testes", sku="OTB-48",
                         fornecedor="Terceiro", anvisa="Não", ativo=True),
        ])
        db.session.commit()


def test_export_pdr_filtra_por_linha(client, app, admin_token):
    _semear_pdr(app)
    res = client.get("/pdr/api/export/apresentacoes.csv?linha=Linha B",
                     headers=_cabecalho(client, admin_token))
    assert res.status_code == 200
    skus = [l[4] for l in _linhas(res)[1:] if l]
    assert skus == ["OTB-48"]


def test_export_pdr_filtra_por_fornecedor_e_busca(client, app, admin_token):
    _semear_pdr(app)
    h = _cabecalho(client, admin_token)
    res = client.get("/pdr/api/export/apresentacoes.csv?fornecedor=Loccus", headers=h)
    assert [l[4] for l in _linhas(res)[1:] if l] == ["EXA-96"]
    # a busca varre produto, sku, apresentação, modelo, fornecedor e descrição
    res = client.get("/pdr/api/export/apresentacoes.csv?busca=terceiro", headers=h)
    assert [l[4] for l in _linhas(res)[1:] if l] == ["OTB-48"]


def test_export_pdr_sem_filtro_traz_tudo(client, app, admin_token):
    _semear_pdr(app)
    res = client.get("/pdr/api/export/apresentacoes.csv",
                     headers=_cabecalho(client, admin_token))
    assert len([l for l in _linhas(res)[1:] if l]) == 2


# ── Permissão: exportar a base é técnico pra cima ────────────────────────────

def test_leitura_nao_exporta(client, leitura_token):
    """`leitura` via a tela, mas levava a base inteira para fora do sistema."""
    h = _cabecalho(client, leitura_token)
    for rota in ("/api/documentos/export",
                 "/api/equipamentos/export",
                 "/api/consumiveis/export"):
        res = client.get(rota, headers=h)
        assert res.status_code == 403, f"{rota} liberou export para leitura"


def test_tecnico_exporta(client, tecnico_token):
    h = _cabecalho(client, tecnico_token)
    for rota in ("/api/documentos/export",
                 "/api/equipamentos/export",
                 "/api/consumiveis/export"):
        assert client.get(rota, headers=h).status_code == 200, rota


def test_export_projetos_continua_restrito_a_gestao(client, tecnico_token):
    """A aba PMO traz orçado/gasto/EAC, e dinheiro é gestão pra cima
    (`pode_ver_financeiro`). Por isso este export não desce para técnico."""
    res = client.get("/api/entregaveis/export",
                     headers=_cabecalho(client, tecnico_token))
    assert res.status_code == 403

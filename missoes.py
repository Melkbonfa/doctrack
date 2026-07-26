"""
missoes.py — Módulo Missões (kanban nativo tipo Planner)

Rotas (todas técnico+; `leitura` recebe 403):
  GET    /api/missoes                        — lista missões (leve; ?arquivadas=1)
  POST   /api/missoes                        — criar (3 colunas padrão ou ?modelo_id)
  GET    /api/missoes/<id>                   — board completo (colunas + cartões, sem descrição)
  PATCH  /api/missoes/<id>                   — renomear/cor/arquivar
  DELETE /api/missoes/<id>                   — ARQUIVA (?definitivo=1 + admin → exclui)
  POST   /api/missoes/<id>/colunas           — criar coluna
  PATCH  /api/missoes/colunas/<id>           — renomear/cor/categoria/limite WIP
  DELETE /api/missoes/colunas/<id>           — excluir coluna (cartões migram, não somem)
  POST   /api/missoes/colunas/<id>/cartoes   — criar cartão
  GET    /api/missoes/cartoes/<id>           — cartão completo (descrição + checklist + comentários)
  PATCH  /api/missoes/cartoes/<id>           — editar (guardado por `versao` → 409)
  DELETE /api/missoes/cartoes/<id>           — excluir cartão
  POST   /api/missoes/reordenar              — reordenar/mover (transação; `versao` → 409)
  GET    /api/missoes/refs                   — busca Equipamento/Projeto/Documento p/ vínculo
  GET    /api/missoes/usuarios               — nomes p/ o seletor de responsáveis
  GET    /api/missoes/etiquetas              — vocabulário de etiquetas já usadas
  GET    /api/missoes/meus-cartoes           — visão cross-missão do usuário logado
  GET    /api/missoes/cartoes-vinculados     — cartões de uma entidade (ficha do documento)
  POST   /api/missoes/cartao-rapido          — criar cartão a partir de uma entidade

  Checklist / comentários (paridade com o Planner):
  POST   /api/missoes/cartoes/<id>/itens          PATCH|DELETE /api/missoes/itens/<id>
  POST   /api/missoes/cartoes/<id>/comentarios    DELETE /api/missoes/comentarios/<id>

  Leitura dos dados (o módulo não produzia informação nenhuma):
  GET    /api/missoes/<id>/metricas          — WIP, aging, cycle time, throughput, carga
  GET    /api/missoes/alertas                — fatos acionáveis (mesmo formato de projetos)
  GET    /api/missoes/<id>/historico         — trilha da missão   (?cartao_id=)
  GET    /api/missoes/<id>/snapshots         — série diária dos indicadores
  POST   /api/missoes/snapshot               — grava a foto do dia (boot/cron)
  GET    /api/missoes/<id>/export            — Excel (cartões + métricas + histórico)

  Modelos de missão (processos que se repetem):
  GET|POST /api/missoes/modelos              DELETE /api/missoes/modelos/<id>

Decisões (pesquisa em planner-missoes-pesquisa/report.md): ordem int reindexada
em transação (suficiente no MVP; fractional/LexoRank só em escala), `versao` como
lock otimista (equivalente do @odata.etag/412 do Graph), board sem `descricao`
(split leve/pesado do plannerTask↔plannerTaskDetails).

Coleta: `MissaoCartaoHistorico` é a série temporal do fluxo — o AuditLog registra
"alguém mexeu", mas não responde quanto tempo um cartão ficou em cada coluna nem
quantos foram concluídos na semana (mesmo papel do DocumentoHistorico).
"""
import io
import json
import re
import unicodedata
from datetime import date, datetime, timedelta

from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import get_jwt_identity

from models import (db, Missao, MissaoColuna, MissaoCartao, MissaoCartaoHistorico,
                    MissaoSnapshot, MissaoCartaoItem, MissaoCartaoComentario,
                    MissaoModelo, missao_cartao_responsaveis,
                    PRIORIDADES_CARTAO, REF_TIPOS_CARTAO, CATEGORIAS_COLUNA,
                    RECORRENCIAS_CARTAO,
                    Equipamento, Projeto, Documento, User, STATUS_MAP)
from auth import require_role, log_action, get_client_ip

# SQLite não aplica String(n) — os limites precisam ser validados aqui.
_RE_COR_HEX = re.compile(r"^#[0-9a-fA-F]{3,8}$")
MAX_TITULO, MAX_NOME_MISSAO, MAX_NOME_COLUNA = 200, 160, 80
MAX_RESPONSAVEIS, MAX_ETIQUETAS = 200, 300
MAX_ITEM_CHECKLIST, MAX_COMENTARIO = 300, 4000
MAX_WIP = 999


def _erro_tamanho(valor, maximo, campo):
    if valor and len(valor) > maximo:
        return f"{campo} muito longo (máx. {maximo} caracteres)"
    return None


def _parse_data(valor):
    """'AAAA-MM-DD' → date, ou None se vazio/inválido. Um regex de formato deixa
    passar 2026-02-31 e 2026-13-45; o calendário é quem sabe se a data existe."""
    if not valor:
        return None
    try:
        return date.fromisoformat(str(valor).strip())
    except ValueError:
        return None


missoes_bp = Blueprint("missoes", __name__)

COLUNAS_PADRAO = [("A fazer", "todo"), ("Fazendo", "doing"), ("Concluído", "done")]

# preenchido por servidor.py para emitir tempo real sem import circular
_rt = {"socketio": None, "publish_event": None, "AuditLog": None, "EventType": None}


def init_realtime(socketio, publish_event, AuditLog, EventType):
    _rt.update(socketio=socketio, publish_event=publish_event,
               AuditLog=AuditLog, EventType=EventType)


def _emit(event_type, payload, email, *, campo="", ip=""):
    if _rt["socketio"] and _rt["publish_event"]:
        try:
            _rt["publish_event"](event_type, payload, user_email=email,
                                 db=db, AuditLog=_rt["AuditLog"],
                                 socketio=_rt["socketio"], campo=campo, ip=ip)
        except Exception:
            pass  # tempo real é best-effort; a gravação já foi feita


def _role():
    u = User.query.filter_by(email=get_jwt_identity(), ativo=True).first()
    return u.role if u else ""


# ── TRILHA TEMPORAL ──────────────────────────────────────────────────────────

def _hist(cartao, evento, *, campo="", antigo="", novo="", origem="manual",
          coluna_origem_id=None, coluna_destino_id=None, por="", em=None):
    """Registra um ponto da série temporal do cartão. Não comita — quem chama
    controla a transação (o sync de documentos roda dentro da transação deles)."""
    db.session.add(MissaoCartaoHistorico(
        cartao_id=cartao.id, missao_id=cartao.missao_id, evento=evento,
        coluna_origem_id=coluna_origem_id, coluna_destino_id=coluna_destino_id,
        campo=campo or "", valor_antigo=_txt(antigo), valor_novo=_txt(novo),
        origem=origem, por=por or "", em=em or datetime.now()))


def _txt(v):
    if v is None:
        return ""
    if isinstance(v, bool):
        return "sim" if v else "não"
    return str(v)


# Campos cuja mudança vira linha no histórico. `descricao` fica de fora: o texto
# inteiro em cada edição inchava a tabela sem responder nenhuma pergunta.
_CAMPOS_RASTREADOS = ("titulo", "responsaveis", "prazo", "data_inicio", "prioridade",
                      "etiquetas", "peso", "recorrencia", "ref_tipo", "ref_id")


def _snapshot_campos(cartao):
    return {c: getattr(cartao, c) for c in _CAMPOS_RASTREADOS}


def _registrar_diff(cartao, antes, email, origem="manual"):
    """Uma linha por campo que realmente mudou."""
    for campo in _CAMPOS_RASTREADOS:
        novo = getattr(cartao, campo)
        if (antes.get(campo) or "") != (novo or ""):
            _hist(cartao, "campo", campo=campo, antigo=antes.get(campo),
                  novo=novo, origem=origem, por=email,
                  coluna_destino_id=cartao.coluna_id)


# ── RESPONSÁVEIS (CSV exibido ↔ N:N consultável) ─────────────────────────────

def _sync_responsaveis(cartao):
    """Reconcilia o N:N a partir do CSV de nomes.

    O CSV continua sendo o texto que aparece no cartão (aceita gente de fora da
    plataforma), mas quem responde "esse cartão é meu?" é a tabela: o
    `responsaveis ILIKE '%Ana%'` casava com "Mariana" e se perdia ao renomear o
    usuário. Nome sem usuário ativo correspondente fica só como texto."""
    nomes = [n.strip() for n in (cartao.responsaveis or "").split(",") if n.strip()]
    if not nomes:
        cartao.responsaveis_users = []
        return
    alvo = {n.lower() for n in nomes}
    achados = [u for u in User.query.filter_by(ativo=True).all()
               if (u.nome or "").strip().lower() in alvo]
    cartao.responsaveis_users = achados


def _cond_responsavel(user):
    """Filtro "cartões de <user>": N:N (fonte oficial) OU nome exato no CSV para
    os cartões que ainda não passaram por um save depois da migração."""
    nome = (user.nome or "").strip()
    sub = (db.session.query(missao_cartao_responsaveis.c.cartao_id)
           .filter(missao_cartao_responsaveis.c.user_id == user.id))
    cond = MissaoCartao.id.in_(sub)
    if nome:
        # ",Ana Paula," dentro de ",Ana Paula,Bruno," — delimitado, sem casar prefixo
        csv = (db.literal(",") +
               db.func.replace(db.func.coalesce(MissaoCartao.responsaveis, ""), ", ", ",") +
               db.literal(","))
        cond = db.or_(cond, csv.like(f"%,{nome},%"))
    return cond


# ── SINCRONIZAÇÃO DOCUMENTO → CARTÃO ─────────────────────────────────────────
# O documento é a fonte da verdade: quando o status dele muda, os cartões
# vinculados (ref_tipo='documento') movem-se para a coluna da categoria alvo.

def _categoria_alvo(setor, status):
    """Mapeia o status do documento para a categoria de coluna do kanban:
    primeira etapa → 'todo'; etapa final → 'done'; intermediárias → 'doing'.
    Status legado (fora do fluxo do setor) → None (no-op)."""
    fluxo = STATUS_MAP.get(setor, [])
    if status not in fluxo:
        return None
    if status == fluxo[0]:
        return "todo"
    if status == fluxo[-1]:
        return "done"
    return "doing"


def sincronizar_cartoes_documento(doc, email):
    """Move os cartões vinculados ao documento para a coluna da categoria alvo
    da respectiva missão (menor `ordem` se houver mais de uma; missão sem coluna
    da categoria → no-op). `concluido` acompanha o 'done' nos dois sentidos
    (regressão do documento reabre o cartão). NÃO comita — roda na transação do
    endpoint de documentos; devolve [(event_type, payload)] para o caller emitir
    após o commit (payload["cartao"] é o objeto, serializar na emissão).

    Grava histórico com origem='doc-sync': antes essa marca só existia no payload
    do socket, que não fica em lugar nenhum consultável."""
    alvo = _categoria_alvo(doc.setor, doc.status)
    if alvo is None:
        return []
    eventos = []
    agora = datetime.now()
    cartoes = MissaoCartao.query.filter_by(ref_tipo="documento", ref_id=doc.id).all()
    for cartao in cartoes:
        destino = (MissaoColuna.query
                   .filter_by(missao_id=cartao.missao_id, categoria=alvo)
                   .order_by(MissaoColuna.ordem).first())
        if destino is None:
            continue
        mudou = False
        origem_id = cartao.coluna_id
        if destino.id != cartao.coluna_id:
            ult = (MissaoCartao.query.filter_by(coluna_id=destino.id)
                   .order_by(MissaoCartao.ordem.desc()).first())
            cartao.coluna_id = destino.id
            cartao.ordem = (ult.ordem + 1) if ult else 0   # append no fim
            cartao.entrou_coluna_em = agora
            mudou = True
            _hist(cartao, "movido", origem="doc-sync", por=email, em=agora,
                  coluna_origem_id=origem_id, coluna_destino_id=destino.id,
                  campo="documento", novo=doc.status or "")
        novo_concluido = (alvo == "done")
        if bool(cartao.concluido) != novo_concluido:
            _marcar_conclusao(cartao, novo_concluido, email, origem="doc-sync",
                              em=agora, reagendar=False)
            mudou = True
        if mudou:
            cartao.versao = (cartao.versao or 0) + 1   # invalida drags concorrentes
            cartao.atualizado_por = email
            cartao.atualizado_em = agora
            eventos.append(("MISSAO_CARTAO_MOVIDO",
                            {"missao_id": cartao.missao_id, "cartao": cartao,
                             "coluna_origem_id": origem_id, "origem": "doc-sync",
                             "documento_id": doc.id}))
    return eventos


def emitir_eventos_sync(eventos, email):
    """Serializa e emite (pós-commit) os eventos devolvidos pelo sync."""
    for ev, payload in eventos:
        payload = dict(payload, cartao=payload["cartao"].to_dict())
        _emit(ev, payload, email)


# ── CONCLUSÃO E RECORRÊNCIA ──────────────────────────────────────────────────

def _marcar_conclusao(cartao, concluido, email, *, origem="manual", em=None,
                      reagendar=True):
    """Aplica a conclusão registrando QUANDO e POR QUEM — sem `concluido_em` não
    havia throughput nem cycle time, só um booleano. Devolve o cartão gerado
    quando a conclusão dispara uma recorrência (ou None)."""
    agora = em or datetime.now()
    if concluido:
        cartao.concluido = True
        cartao.concluido_em = agora
        cartao.concluido_por = email
        _hist(cartao, "concluido", origem=origem, por=email, em=agora,
              coluna_destino_id=cartao.coluna_id)
        # Reagendamento só na conclusão manual: o doc-sync alterna concluído/
        # reaberto conforme o documento avança e regride, e clonaria em cascata.
        if reagendar and (cartao.recorrencia or ""):
            return _reagendar_recorrente(cartao, email)
    else:
        cartao.concluido = False
        cartao.concluido_em = None
        cartao.concluido_por = ""
        _hist(cartao, "reaberto", origem=origem, por=email, em=agora,
              coluna_destino_id=cartao.coluna_id)
    return None


def _somar_periodo(base, recorrencia):
    """Próxima ocorrência. Meses são somados no calendário (uma calibração
    'anual' não pode andar 365 dias e escorregar um dia por ano bissexto)."""
    passo = RECORRENCIAS_CARTAO.get(recorrencia or "")
    if not passo:
        return None
    unidade, n = passo
    if unidade == "dias":
        return base + timedelta(days=n)
    mes = base.month - 1 + n
    ano = base.year + mes // 12
    mes = mes % 12 + 1
    dia = min(base.day, [31, 29 if (ano % 4 == 0 and (ano % 100 != 0 or ano % 400 == 0))
                         else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][mes - 1])
    return date(ano, mes, dia)


def _reagendar_recorrente(cartao, email):
    """Clona o cartão concluído para a próxima ocorrência, na coluna 'todo'.

    O domínio é cheio de obrigação periódica (calibração, requalificação, revisão
    anual) e o kanban não tinha como representá-la: fechava-se o cartão e o
    próximo dependia de alguém lembrar."""
    base = _parse_data(cartao.prazo) or date.today()
    proximo = _somar_periodo(base, cartao.recorrencia)
    if not proximo:
        return None
    destino = (MissaoColuna.query
               .filter_by(missao_id=cartao.missao_id, categoria="todo")
               .order_by(MissaoColuna.ordem).first()) or \
              (MissaoColuna.query.filter_by(missao_id=cartao.missao_id)
               .order_by(MissaoColuna.ordem).first())
    if destino is None:
        return None
    ult = (MissaoCartao.query.filter_by(coluna_id=destino.id)
           .order_by(MissaoCartao.ordem.desc()).first())
    novo = MissaoCartao(
        missao_id=cartao.missao_id, coluna_id=destino.id, titulo=cartao.titulo,
        descricao=cartao.descricao or "", responsaveis=cartao.responsaveis or "",
        prazo=proximo.isoformat(), data_inicio="", prioridade=cartao.prioridade or "media",
        etiquetas=cartao.etiquetas or "", peso=cartao.peso if cartao.peso is not None else 1.0,
        recorrencia=cartao.recorrencia, ref_tipo=cartao.ref_tipo or "", ref_id=cartao.ref_id,
        criado_por=email, atualizado_por=email,
        ordem=(ult.ordem + 1) if ult else 0)
    db.session.add(novo)
    db.session.flush()          # precisa do id para o histórico e o N:N
    novo.responsaveis_users = list(cartao.responsaveis_users)
    for item in cartao.itens:   # checklist volta zerado — é a mesma rotina de novo
        db.session.add(MissaoCartaoItem(cartao_id=novo.id, texto=item.texto,
                                        feito=False, ordem=item.ordem))
    _hist(novo, "criado", origem="recorrencia", por=email,
          coluna_destino_id=destino.id, campo="recorrencia",
          novo=f"{cartao.recorrencia} · a partir do cartão {cartao.id}")
    return novo


# ── METADADOS EM LOTE (evita N+1 no board) ───────────────────────────────────

def _mapa_refs(cartoes):
    """Metadados dos vínculos em lote (no máx. 1 query IN por tipo).

    Inclui os inativos de propósito: filtrar por `ativo` fazia o chip sumir sem
    aviso quando o documento era desativado, e o cartão perdia o contexto."""
    ids = {t: {c.ref_id for c in cartoes if c.ref_tipo == t and c.ref_id}
           for t in REF_TIPOS_CARTAO}
    mapa = {}
    if ids["equipamento"]:
        for e in Equipamento.query.filter(Equipamento.id.in_(ids["equipamento"])):
            mapa[("equipamento", e.id)] = {"label": e.nome, "status": "",
                                           "status_global": "", "ativo": bool(e.ativo)}
    if ids["projeto"]:
        for p in Projeto.query.filter(Projeto.id.in_(ids["projeto"])):
            mapa[("projeto", p.id)] = {"label": p.nome, "status": "",
                                       "status_global": "", "ativo": bool(p.ativo)}
    if ids["documento"]:
        for d in Documento.query.filter(Documento.id.in_(ids["documento"])):
            mapa[("documento", d.id)] = {"label": d.documento,
                                         "status": d.status or "",
                                         "status_global": d.status_global,
                                         "ativo": bool(d.ativo)}
    return mapa


def _mapa_extras(cartoes):
    """Contagem de checklist e comentários por cartão em 3 queries agregadas."""
    ids = [c.id for c in cartoes]
    if not ids:
        return {}
    out = {i: {"itens": 0, "itens_feitos": 0, "comentarios": 0} for i in ids}
    q_item = db.session.query(MissaoCartaoItem.cartao_id,
                              db.func.count(MissaoCartaoItem.id))
    for cid, n in q_item.filter(MissaoCartaoItem.cartao_id.in_(ids)) \
                        .group_by(MissaoCartaoItem.cartao_id).all():
        out[cid]["itens"] = n or 0
    for cid, n in q_item.filter(MissaoCartaoItem.cartao_id.in_(ids),
                                MissaoCartaoItem.feito == True) \
                        .group_by(MissaoCartaoItem.cartao_id).all():
        out[cid]["itens_feitos"] = n or 0
    for cid, n in db.session.query(MissaoCartaoComentario.cartao_id,
                                   db.func.count(MissaoCartaoComentario.id)) \
                            .filter(MissaoCartaoComentario.cartao_id.in_(ids)) \
                            .group_by(MissaoCartaoComentario.cartao_id).all():
        out[cid]["comentarios"] = n or 0
    return out


def _validar_ref(ref_tipo, ref_id):
    """Vínculo opcional: confere existência ao gravar (sem FK rígida).
    Devolve (ref_tipo, ref_id, erro|None)."""
    if not ref_tipo:
        return "", None, None
    if ref_tipo not in REF_TIPOS_CARTAO:
        return None, None, f"ref_tipo inválido. Use: {', '.join(REF_TIPOS_CARTAO)}"
    try:
        rid = int(ref_id)
    except (TypeError, ValueError):
        return None, None, "ref_id deve ser número"
    modelo = {"equipamento": Equipamento, "projeto": Projeto, "documento": Documento}[ref_tipo]
    alvo = modelo.query.get(rid)
    if not alvo or not alvo.ativo:
        return None, None, f"{ref_tipo} {rid} não encontrado"
    return ref_tipo, rid, None


# ── MISSÕES ──────────────────────────────────────────────────────────────────

def _pode_administrar(m, email, role):
    """Excluir/arquivar missão: gestão sempre; técnico só a que ele criou.
    `criado_por` era gravado e nunca consultado — qualquer técnico apagava a
    missão de qualquer área."""
    return role in ("admin", "gestor") or (m.criado_por or "") == email


@missoes_bp.route("/api/missoes", methods=["GET"])
@require_role("admin", "gestor", "tecnico")
def listar_missoes():
    arquivadas = request.args.get("arquivadas", "").strip() == "1"
    missoes = (Missao.query.filter_by(arquivado=arquivadas)
               .order_by(Missao.ordem, Missao.id).all())
    # contagens agregadas (evita carregar os cartões de cada missão só para contar)
    base = db.session.query(MissaoCartao.missao_id, db.func.count(MissaoCartao.id))
    counts = dict(base.group_by(MissaoCartao.missao_id).all())
    abertos = dict(base.filter(MissaoCartao.concluido == False)
                       .group_by(MissaoCartao.missao_id).all())
    return jsonify({"missoes": [m.to_dict(n_cartoes=counts.get(m.id, 0),
                                          n_abertos=abertos.get(m.id, 0))
                                for m in missoes]})


@missoes_bp.route("/api/missoes", methods=["POST"])
@require_role("admin", "gestor", "tecnico")
def criar_missao():
    data = request.get_json(silent=True) or {}
    nome = (data.get("nome") or "").strip()
    if not nome:
        return jsonify({"erro": "nome é obrigatório"}), 400
    erro = _erro_tamanho(nome, MAX_NOME_MISSAO, "nome")
    if erro:
        return jsonify({"erro": erro}), 400
    accent = (data.get("accent") or "").strip()
    if accent and not _RE_COR_HEX.match(accent):
        return jsonify({"erro": "accent deve ser cor hex (#rrggbb)"}), 400
    email = get_jwt_identity()
    ult = Missao.query.order_by(Missao.ordem.desc()).first()
    m = Missao(nome=nome,
               descricao=(data.get("descricao") or "").strip(),
               accent=accent,
               ordem=(ult.ordem + 1) if ult else 0,
               criado_por=email)
    db.session.add(m)
    db.session.flush()   # garante m.id para as colunas

    modelo = MissaoModelo.query.get(int(data["modelo_id"])) if data.get("modelo_id") else None
    if modelo is not None:
        _aplicar_modelo(m, modelo, email)
    else:
        for i, (nome_col, cat) in enumerate(COLUNAS_PADRAO):
            db.session.add(MissaoColuna(missao_id=m.id, nome=nome_col,
                                        categoria=cat, ordem=i))
    db.session.commit()
    log_action(email, "CREATE", entidade=f"Missão: {m.nome}", ip=get_client_ip())
    _emit("MISSAO_CREATED", {"missao": m.to_dict()}, email, ip=get_client_ip())
    return jsonify({"missao": m.to_dict(com_colunas=True)}), 201


@missoes_bp.route("/api/missoes/<int:mid>", methods=["GET"])
@require_role("admin", "gestor", "tecnico")
def detalhe_missao(mid):
    # eager-load colunas+cartões em 3 queries (sem isso o board faz 1 query por coluna)
    m = (Missao.query
         .options(db.selectinload(Missao.colunas).selectinload(MissaoColuna.cartoes))
         .get_or_404(mid))
    cartoes = [c for col in m.colunas for c in col.cartoes]
    return jsonify({"missao": m.to_dict(com_colunas=True,
                                        refs_map=_mapa_refs(cartoes),
                                        extras_map=_mapa_extras(cartoes))})


@missoes_bp.route("/api/missoes/<int:mid>", methods=["PATCH"])
@require_role("admin", "gestor", "tecnico")
def editar_missao(mid):
    m = Missao.query.get_or_404(mid)
    data = request.get_json(silent=True) or {}
    email = get_jwt_identity()
    if "nome" in data:
        nome = (data.get("nome") or "").strip()
        if not nome:
            return jsonify({"erro": "nome não pode ficar vazio"}), 400
        erro = _erro_tamanho(nome, MAX_NOME_MISSAO, "nome")
        if erro:
            return jsonify({"erro": erro}), 400
        m.nome = nome
    if "descricao" in data:
        m.descricao = (data.get("descricao") or "").strip()
    if "accent" in data:
        accent = (data.get("accent") or "").strip()
        if accent and not _RE_COR_HEX.match(accent):
            return jsonify({"erro": "accent deve ser cor hex (#rrggbb)"}), 400
        m.accent = accent
    if "arquivado" in data:
        if not _pode_administrar(m, email, _role()):
            return jsonify({"erro": "só quem criou a missão (ou a gestão) pode arquivá-la"}), 403
        m.arquivado = bool(data.get("arquivado"))
    db.session.commit()
    log_action(email, "UPDATE", entidade=f"Missão: {m.nome}", ip=get_client_ip())
    _emit("MISSAO_UPDATED", {"missao": m.to_dict()}, email, ip=get_client_ip())
    return jsonify({"missao": m.to_dict()})


@missoes_bp.route("/api/missoes/<int:mid>", methods=["DELETE"])
@require_role("admin", "gestor", "tecnico")
def excluir_missao(mid):
    """Arquiva por padrão. O DELETE era um `db.session.delete()` em cascata que
    levava colunas, cartões e — agora — a série histórica inteira junto; o resto
    do sistema usa desativação (`ativo`/`arquivado`). A exclusão definitiva
    continua existindo, mas é explícita e só para admin."""
    m = Missao.query.get_or_404(mid)
    email = get_jwt_identity()
    role = _role()
    if not _pode_administrar(m, email, role):
        return jsonify({"erro": "só quem criou a missão (ou a gestão) pode excluí-la"}), 403
    definitivo = request.args.get("definitivo", "").strip() == "1"
    nome = m.nome
    if definitivo:
        if role != "admin":
            return jsonify({"erro": "exclusão definitiva é restrita ao administrador"}), 403
        db.session.delete(m)   # cascade: colunas + cartões + histórico + snapshots
        db.session.commit()
        log_action(email, "DELETE", entidade=f"Missão: {nome}", ip=get_client_ip())
        _emit("MISSAO_DELETED", {"missao_id": mid}, email, ip=get_client_ip())
        return jsonify({"ok": True, "definitivo": True})
    m.arquivado = True
    db.session.commit()
    log_action(email, "UPDATE", entidade=f"Missão: {nome}", campo="arquivado",
               antigo="não", novo="sim", ip=get_client_ip())
    _emit("MISSAO_ARQUIVADA", {"missao_id": mid, "missao": m.to_dict()}, email,
          campo="arquivado", ip=get_client_ip())
    return jsonify({"ok": True, "arquivado": True})


# ── COLUNAS ──────────────────────────────────────────────────────────────────

def _validar_wip(valor):
    """Devolve (limite, erro|None). 0/vazio = sem limite."""
    if valor in (None, ""):
        return 0, None
    try:
        n = int(valor)
    except (TypeError, ValueError):
        return None, "limite_wip deve ser número"
    if not (0 <= n <= MAX_WIP):
        return None, f"limite_wip deve estar entre 0 e {MAX_WIP}"
    return n, None


@missoes_bp.route("/api/missoes/<int:mid>/colunas", methods=["POST"])
@require_role("admin", "gestor", "tecnico")
def criar_coluna(mid):
    m = Missao.query.get_or_404(mid)
    data = request.get_json(silent=True) or {}
    nome = (data.get("nome") or "").strip()
    if not nome:
        return jsonify({"erro": "nome é obrigatório"}), 400
    erro = _erro_tamanho(nome, MAX_NOME_COLUNA, "nome")
    if erro:
        return jsonify({"erro": erro}), 400
    cat = (data.get("categoria") or "").strip()
    if cat and cat not in CATEGORIAS_COLUNA:
        return jsonify({"erro": f"categoria inválida. Use: {', '.join(CATEGORIAS_COLUNA)}"}), 400
    # A cor cai num atributo style=""; sem validar, aceitava CSS arbitrário
    # (url(...) externo). Mesma regra do accent da missão.
    cor = (data.get("cor") or "").strip()
    if cor and not _RE_COR_HEX.match(cor):
        return jsonify({"erro": "cor deve ser hex (#rrggbb)"}), 400
    limite, erro = _validar_wip(data.get("limite_wip"))
    if erro:
        return jsonify({"erro": erro}), 400
    ult = (MissaoColuna.query.filter_by(missao_id=m.id)
           .order_by(MissaoColuna.ordem.desc()).first())
    c = MissaoColuna(missao_id=m.id, nome=nome, categoria=cat, cor=cor,
                     limite_wip=limite, ordem=(ult.ordem + 1) if ult else 0)
    db.session.add(c)
    db.session.commit()
    email = get_jwt_identity()
    log_action(email, "CREATE", entidade=f"Missão: {m.nome} · Coluna: {nome}", ip=get_client_ip())
    _emit("MISSAO_COLUNA_CREATED", {"missao_id": m.id, "coluna": c.to_dict(com_cartoes=True)},
          email, ip=get_client_ip())
    return jsonify({"coluna": c.to_dict(com_cartoes=True)}), 201


@missoes_bp.route("/api/missoes/colunas/<int:cid>", methods=["PATCH"])
@require_role("admin", "gestor", "tecnico")
def editar_coluna(cid):
    c = MissaoColuna.query.get_or_404(cid)
    data = request.get_json(silent=True) or {}
    if "nome" in data:
        nome = (data.get("nome") or "").strip()
        if not nome:
            return jsonify({"erro": "nome não pode ficar vazio"}), 400
        erro = _erro_tamanho(nome, MAX_NOME_COLUNA, "nome")
        if erro:
            return jsonify({"erro": erro}), 400
        c.nome = nome
    if "cor" in data:
        cor = (data.get("cor") or "").strip()
        if cor and not _RE_COR_HEX.match(cor):
            return jsonify({"erro": "cor deve ser hex (#rrggbb)"}), 400
        c.cor = cor
    if "categoria" in data:
        cat = (data.get("categoria") or "").strip()
        if cat and cat not in CATEGORIAS_COLUNA:
            return jsonify({"erro": f"categoria inválida. Use: {', '.join(CATEGORIAS_COLUNA)}"}), 400
        c.categoria = cat
    if "limite_wip" in data:
        limite, erro = _validar_wip(data.get("limite_wip"))
        if erro:
            return jsonify({"erro": erro}), 400
        c.limite_wip = limite
    db.session.commit()
    email = get_jwt_identity()
    log_action(email, "UPDATE", entidade=f"Coluna de missão: {c.nome}", ip=get_client_ip())
    _emit("MISSAO_COLUNA_UPDATED", {"missao_id": c.missao_id, "coluna": c.to_dict()},
          email, ip=get_client_ip())
    return jsonify({"coluna": c.to_dict()})


@missoes_bp.route("/api/missoes/colunas/<int:cid>", methods=["DELETE"])
@require_role("admin", "gestor", "tecnico")
def excluir_coluna(cid):
    """Os cartões migram para outra coluna em vez de serem destruídos junto.
    `?destino_id=` escolhe para onde; sem ele, vai para a coluna anterior (ou a
    primeira). Excluir a última coluna da missão é recusado."""
    c = MissaoColuna.query.get_or_404(cid)
    mid, nome = c.missao_id, c.nome
    email = get_jwt_identity()
    cartoes = list(c.cartoes)
    if cartoes:
        destino_id = request.args.get("destino_id", type=int)
        outras = (MissaoColuna.query.filter(MissaoColuna.missao_id == mid,
                                            MissaoColuna.id != cid)
                  .order_by(MissaoColuna.ordem).all())
        if not outras:
            return jsonify({"erro": "esta é a única coluna da missão — mova ou "
                                    "exclua os cartões antes"}), 400
        destino = next((o for o in outras if o.id == destino_id), None) or \
            next((o for o in reversed(outras) if (o.ordem or 0) < (c.ordem or 0)), outras[0])
        ult = (MissaoCartao.query.filter_by(coluna_id=destino.id)
               .order_by(MissaoCartao.ordem.desc()).first())
        pos = (ult.ordem + 1) if ult else 0
        agora = datetime.now()
        for cartao in sorted(cartoes, key=lambda x: x.ordem or 0):
            _hist(cartao, "movido", por=email, em=agora, campo="coluna_excluida",
                  antigo=nome, novo=destino.nome,
                  coluna_origem_id=cid, coluna_destino_id=destino.id)
            # Reatribui pela relação, não pela FK: `cartoes` tem delete-orphan e
            # a coleção em memória é o que a cascata consulta — mexer só no
            # coluna_id deixava o cartão na lista da coluna velha e ele morria
            # junto com ela.
            cartao.coluna = destino
            cartao.ordem = pos
            cartao.entrou_coluna_em = agora
            cartao.versao = (cartao.versao or 0) + 1
            pos += 1
        db.session.flush()
    db.session.delete(c)
    db.session.commit()
    log_action(email, "DELETE", entidade=f"Coluna de missão: {nome}", ip=get_client_ip())
    _emit("MISSAO_COLUNA_DELETED", {"missao_id": mid, "coluna_id": cid}, email,
          ip=get_client_ip())
    return jsonify({"ok": True, "cartoes_movidos": len(cartoes)})


# ── CARTÕES ──────────────────────────────────────────────────────────────────

_CARTAO_STR = ("responsaveis", "etiquetas")


def _aplicar_campos_cartao(cartao, data):
    """Aplica campos simples do payload ao cartão. Devolve erro|None.
    A conclusão NÃO é tratada aqui — passa por _marcar_conclusao (marcos + histórico)."""
    if "titulo" in data:
        titulo = (data.get("titulo") or "").strip()
        if not titulo:
            return "titulo não pode ficar vazio"
        erro = _erro_tamanho(titulo, MAX_TITULO, "titulo")
        if erro:
            return erro
        cartao.titulo = titulo
    if "descricao" in data:
        cartao.descricao = (data.get("descricao") or "").strip()
    _maximos = {"responsaveis": MAX_RESPONSAVEIS, "etiquetas": MAX_ETIQUETAS}
    for campo in _CARTAO_STR:
        if campo in data:
            valor = (data.get(campo) or "").strip()
            erro = _erro_tamanho(valor, _maximos[campo], campo)
            if erro:
                return erro
            setattr(cartao, campo, valor)
    for campo, rotulo in (("prazo", "prazo"), ("data_inicio", "data de início")):
        if campo in data:
            valor = (data.get(campo) or "").strip()
            if valor and _parse_data(valor) is None:
                return f"{rotulo} inválido — use uma data real no formato AAAA-MM-DD"
            setattr(cartao, campo, valor)
    ini, fim = _parse_data(cartao.data_inicio), _parse_data(cartao.prazo)
    if ini and fim and ini > fim:
        return "a data de início não pode ser depois do prazo"
    if "prioridade" in data:
        pri = (data.get("prioridade") or "media").strip()
        if pri not in PRIORIDADES_CARTAO:
            return f"prioridade inválida. Use: {', '.join(PRIORIDADES_CARTAO)}"
        cartao.prioridade = pri
    if "recorrencia" in data:
        rec = (data.get("recorrencia") or "").strip()
        if rec not in RECORRENCIAS_CARTAO:
            return ("recorrência inválida. Use: " +
                    ", ".join(k for k in RECORRENCIAS_CARTAO if k))
        cartao.recorrencia = rec
    if "peso" in data:
        try:
            peso = float(data.get("peso") if data.get("peso") not in (None, "") else 1)
        except (TypeError, ValueError):
            return "peso deve ser número"
        if not (0 < peso <= 1000):
            return "peso deve estar entre 0 (exclusivo) e 1000"
        cartao.peso = peso
    if "ref_tipo" in data or "ref_id" in data:
        ref_tipo, ref_id, erro = _validar_ref(
            (data.get("ref_tipo") or "").strip(), data.get("ref_id"))
        if erro:
            return erro
        cartao.ref_tipo, cartao.ref_id = ref_tipo, ref_id
    return None


@missoes_bp.route("/api/missoes/colunas/<int:cid>/cartoes", methods=["POST"])
@require_role("admin", "gestor", "tecnico")
def criar_cartao(cid):
    col = MissaoColuna.query.get_or_404(cid)
    data = request.get_json(silent=True) or {}
    titulo = (data.get("titulo") or "").strip()
    if not titulo:
        return jsonify({"erro": "titulo é obrigatório"}), 400
    email = get_jwt_identity()
    agora = datetime.now()
    ult = (MissaoCartao.query.filter_by(coluna_id=col.id)
           .order_by(MissaoCartao.ordem.desc()).first())
    cartao = MissaoCartao(missao_id=col.missao_id, coluna_id=col.id,
                          titulo=titulo, criado_por=email, atualizado_por=email,
                          criado_em=agora, entrou_coluna_em=agora,
                          ordem=(ult.ordem + 1) if ult else 0)
    erro = _aplicar_campos_cartao(cartao, data)
    if erro:
        return jsonify({"erro": erro}), 400
    db.session.add(cartao)
    db.session.flush()               # id para o N:N e para o histórico
    _sync_responsaveis(cartao)
    if data.get("concluido"):
        _marcar_conclusao(cartao, True, email, em=agora, reagendar=False)
    _hist(cartao, "criado", por=email, em=agora, coluna_destino_id=col.id,
          campo="titulo", novo=titulo)
    db.session.commit()
    log_action(email, "CREATE", entidade=f"Cartão: {titulo}", ip=get_client_ip())
    _emit("MISSAO_CARTAO_CREATED",
          {"missao_id": col.missao_id, "cartao": cartao.to_dict()}, email,
          ip=get_client_ip())
    return jsonify({"cartao": cartao.to_dict()}), 201


@missoes_bp.route("/api/missoes/cartoes/<int:cid>", methods=["GET"])
@require_role("admin", "gestor", "tecnico")
def detalhe_cartao(cid):
    cartao = MissaoCartao.query.get_or_404(cid)
    return jsonify({"cartao": cartao.to_dict(com_descricao=True)})


@missoes_bp.route("/api/missoes/cartoes/<int:cid>", methods=["PATCH"])
@require_role("admin", "gestor", "tecnico")
def editar_cartao(cid):
    cartao = MissaoCartao.query.get_or_404(cid)
    data = request.get_json(silent=True) or {}
    # Lock otimista: se o cliente mandou a versão que leu e ela não bate mais,
    # outro usuário editou no meio — 409 e o cliente recarrega o cartão.
    if "versao" in data and int(data.get("versao") or 0) != (cartao.versao or 0):
        return jsonify({"erro": "conflito de edição — recarregue o cartão",
                        "cartao": cartao.to_dict(com_descricao=True)}), 409
    email = get_jwt_identity()
    antes = _snapshot_campos(cartao)
    concluido_antes = bool(cartao.concluido)
    erro = _aplicar_campos_cartao(cartao, data)
    if erro:
        return jsonify({"erro": erro}), 400
    if "responsaveis" in data:
        _sync_responsaveis(cartao)
    _registrar_diff(cartao, antes, email)
    gerado = None
    if "concluido" in data and bool(data.get("concluido")) != concluido_antes:
        gerado = _marcar_conclusao(cartao, bool(data.get("concluido")), email)
    cartao.versao = (cartao.versao or 0) + 1
    cartao.atualizado_por = email
    cartao.atualizado_em = datetime.now()
    db.session.commit()
    _emit("MISSAO_CARTAO_UPDATED",
          {"missao_id": cartao.missao_id, "cartao": cartao.to_dict()}, email,
          ip=get_client_ip())
    resposta = {"cartao": cartao.to_dict(com_descricao=True)}
    if gerado is not None:
        _emit("MISSAO_CARTAO_CREATED",
              {"missao_id": gerado.missao_id, "cartao": gerado.to_dict(),
               "origem": "recorrencia"}, email)
        resposta["recorrencia"] = {"cartao_id": gerado.id, "prazo": gerado.prazo}
    return jsonify(resposta)


@missoes_bp.route("/api/missoes/cartoes/<int:cid>", methods=["DELETE"])
@require_role("admin", "gestor", "tecnico")
def excluir_cartao(cid):
    cartao = MissaoCartao.query.get_or_404(cid)
    mid, titulo = cartao.missao_id, cartao.titulo
    db.session.delete(cartao)   # cascade: histórico, checklist e comentários do cartão
    db.session.commit()
    email = get_jwt_identity()
    log_action(email, "DELETE", entidade=f"Cartão: {titulo}", ip=get_client_ip())
    _emit("MISSAO_CARTAO_DELETED", {"missao_id": mid, "cartao_id": cid}, email,
          ip=get_client_ip())
    return jsonify({"ok": True})


# ── CHECKLIST DO CARTÃO ──────────────────────────────────────────────────────

@missoes_bp.route("/api/missoes/cartoes/<int:cid>/itens", methods=["POST"])
@require_role("admin", "gestor", "tecnico")
def criar_item(cid):
    cartao = MissaoCartao.query.get_or_404(cid)
    texto = ((request.get_json(silent=True) or {}).get("texto") or "").strip()
    if not texto:
        return jsonify({"erro": "texto é obrigatório"}), 400
    erro = _erro_tamanho(texto, MAX_ITEM_CHECKLIST, "texto")
    if erro:
        return jsonify({"erro": erro}), 400
    ult = (MissaoCartaoItem.query.filter_by(cartao_id=cid)
           .order_by(MissaoCartaoItem.ordem.desc()).first())
    item = MissaoCartaoItem(cartao_id=cid, texto=texto,
                            ordem=(ult.ordem + 1) if ult else 0)
    db.session.add(item)
    db.session.commit()
    _emit("MISSAO_CARTAO_UPDATED",
          {"missao_id": cartao.missao_id, "cartao": cartao.to_dict()},
          get_jwt_identity())
    return jsonify({"item": item.to_dict()}), 201


@missoes_bp.route("/api/missoes/itens/<int:iid>", methods=["PATCH", "DELETE"])
@require_role("admin", "gestor", "tecnico")
def editar_item(iid):
    item = MissaoCartaoItem.query.get_or_404(iid)
    cartao = item.cartao
    if request.method == "DELETE":
        db.session.delete(item)
    else:
        data = request.get_json(silent=True) or {}
        if "texto" in data:
            texto = (data.get("texto") or "").strip()
            if not texto:
                return jsonify({"erro": "texto não pode ficar vazio"}), 400
            erro = _erro_tamanho(texto, MAX_ITEM_CHECKLIST, "texto")
            if erro:
                return jsonify({"erro": erro}), 400
            item.texto = texto
        if "feito" in data:
            item.feito = bool(data.get("feito"))
    db.session.commit()
    _emit("MISSAO_CARTAO_UPDATED",
          {"missao_id": cartao.missao_id, "cartao": cartao.to_dict()},
          get_jwt_identity())
    return jsonify({"ok": True} if request.method == "DELETE" else {"item": item.to_dict()})


# ── COMENTÁRIOS DO CARTÃO ────────────────────────────────────────────────────

@missoes_bp.route("/api/missoes/cartoes/<int:cid>/comentarios", methods=["POST"])
@require_role("admin", "gestor", "tecnico")
def comentar(cid):
    cartao = MissaoCartao.query.get_or_404(cid)
    texto = ((request.get_json(silent=True) or {}).get("texto") or "").strip()
    if not texto:
        return jsonify({"erro": "texto é obrigatório"}), 400
    erro = _erro_tamanho(texto, MAX_COMENTARIO, "texto")
    if erro:
        return jsonify({"erro": erro}), 400
    email = get_jwt_identity()
    user = User.query.filter_by(email=email).first()
    c = MissaoCartaoComentario(cartao_id=cid, texto=texto,
                               por=(user.nome if user and user.nome else email))
    db.session.add(c)
    db.session.commit()
    _emit("MISSAO_CARTAO_COMENTADO",
          {"missao_id": cartao.missao_id, "cartao_id": cid,
           "comentario": c.to_dict()}, email, ip=get_client_ip())
    return jsonify({"comentario": c.to_dict()}), 201


@missoes_bp.route("/api/missoes/comentarios/<int:coid>", methods=["DELETE"])
@require_role("admin", "gestor", "tecnico")
def excluir_comentario(coid):
    c = MissaoCartaoComentario.query.get_or_404(coid)
    email = get_jwt_identity()
    user = User.query.filter_by(email=email).first()
    autor = (user.nome if user and user.nome else email)
    if (c.por or "") != autor and _role() not in ("admin", "gestor"):
        return jsonify({"erro": "só o autor (ou a gestão) pode apagar o comentário"}), 403
    missao_id, cartao_id = c.cartao.missao_id, c.cartao_id
    db.session.delete(c)
    db.session.commit()
    _emit("MISSAO_CARTAO_COMENTADO",
          {"missao_id": missao_id, "cartao_id": cartao_id}, email)
    return jsonify({"ok": True})


# ── REORDENAR / MOVER ────────────────────────────────────────────────────────

def _reindexar(coluna_id, ids):
    """Reatribui ordem = índice para os cartões da coluna, na ordem enviada.
    Cartões da coluna fora da lista vão para o fim (ordem preservada)."""
    ids = [int(i) for i in ids]
    cartoes = {c.id: c for c in MissaoCartao.query.filter_by(coluna_id=coluna_id).all()}
    pos = 0
    for i in ids:
        c = cartoes.pop(i, None)
        if c is not None:
            c.ordem = pos
            pos += 1
    for c in sorted(cartoes.values(), key=lambda x: x.ordem or 0):
        c.ordem = pos
        pos += 1


@missoes_bp.route("/api/missoes/reordenar", methods=["POST"])
@require_role("admin", "gestor", "tecnico")
def reordenar():
    """Duas formas (tudo numa única transação, ordem = índice da lista):
      {coluna_id, ids:[...]}                                — reordenar na coluna
      {cartao_id, versao, coluna_destino_id, ids:[...],
       ids_origem:[...]}                                    — mover entre colunas
    O move é guardado pela `versao` do cartão (conflito ⇒ 409)."""
    data = request.get_json(silent=True) or {}
    email = get_jwt_identity()

    cartao_id = data.get("cartao_id")
    if cartao_id:
        cartao = MissaoCartao.query.get_or_404(int(cartao_id))
        destino = MissaoColuna.query.get_or_404(int(data.get("coluna_destino_id") or 0))
        if destino.missao_id != cartao.missao_id:
            return jsonify({"erro": "coluna de outra missão"}), 400
        # Lock otimista do move: dois usuários arrastando o mesmo cartão não
        # podem se sobrescrever em silêncio (pesquisa: @odata.etag → 412/409).
        if "versao" in data and int(data.get("versao") or 0) != (cartao.versao or 0):
            return jsonify({"erro": "conflito — o cartão foi movido por outro usuário",
                            "cartao": cartao.to_dict()}), 409
        origem_id = cartao.coluna_id
        agora = datetime.now()
        gerado = None
        if origem_id != destino.id:
            origem = MissaoColuna.query.get(origem_id)
            _hist(cartao, "movido", por=email, em=agora,
                  antigo=(origem.nome if origem else ""), novo=destino.nome,
                  coluna_origem_id=origem_id, coluna_destino_id=destino.id)
            cartao.entrou_coluna_em = agora
            # A coluna É o estado: arrastar para uma coluna 'done' conclui o
            # cartão (e o inverso reabre). Sem isso o board acumulava cartão na
            # coluna de concluído com o `concluido` ainda em falso.
            alvo = (destino.categoria or "")
            if alvo == "done" and not cartao.concluido:
                gerado = _marcar_conclusao(cartao, True, email, em=agora)
            elif alvo and alvo != "done" and cartao.concluido:
                _marcar_conclusao(cartao, False, email, em=agora)
        cartao.coluna_id = destino.id
        cartao.versao = (cartao.versao or 0) + 1
        cartao.atualizado_por = email
        cartao.atualizado_em = agora
        db.session.flush()
        _reindexar(destino.id, data.get("ids") or [])
        if origem_id != destino.id:
            _reindexar(origem_id, data.get("ids_origem") or [])
        db.session.commit()
        _emit("MISSAO_CARTAO_MOVIDO",
              {"missao_id": cartao.missao_id, "cartao": cartao.to_dict(),
               "coluna_origem_id": origem_id}, email, ip=get_client_ip())
        if gerado is not None:
            _emit("MISSAO_CARTAO_CREATED",
                  {"missao_id": gerado.missao_id, "cartao": gerado.to_dict(),
                   "origem": "recorrencia"}, email)
        return jsonify({"cartao": cartao.to_dict()})

    coluna_id = data.get("coluna_id")
    if coluna_id:
        col = MissaoColuna.query.get_or_404(int(coluna_id))
        _reindexar(col.id, data.get("ids") or [])
        db.session.commit()
        _emit("MISSAO_COLUNA_REORDENADA",
              {"missao_id": col.missao_id, "coluna": col.to_dict(com_cartoes=True)}, email)
        return jsonify({"coluna": col.to_dict(com_cartoes=True)})

    # Reordenar as próprias colunas do board: {missao_id, colunas_ids:[...]}
    missao_id = data.get("missao_id")
    if missao_id and isinstance(data.get("colunas_ids"), list):
        m = Missao.query.get_or_404(int(missao_id))
        colunas = {c.id: c for c in m.colunas}
        pos = 0
        for i in data["colunas_ids"]:
            c = colunas.pop(int(i), None)
            if c is not None:
                c.ordem = pos
                pos += 1
        for c in sorted(colunas.values(), key=lambda x: x.ordem or 0):
            c.ordem = pos
            pos += 1
        db.session.commit()
        _emit("MISSAO_COLUNAS_REORDENADAS", {"missao_id": m.id}, email)
        return jsonify({"ok": True})

    return jsonify({"erro": "envie coluna_id, cartao_id ou missao_id"}), 400


# ── USUÁRIOS E ETIQUETAS (p/ os seletores do modal) ──────────────────────────

@missoes_bp.route("/api/missoes/usuarios", methods=["GET"])
@require_role("admin", "gestor", "tecnico")
def listar_usuarios_leve():
    """Só nomes de usuários ativos, para o autocomplete de responsáveis.
    (O /api/users completo é gestor+; aqui técnicos precisam da lista, então
    expõe-se apenas o nome — nada de email/role/áreas.)"""
    users = User.query.filter_by(ativo=True).order_by(User.nome).all()
    return jsonify({"usuarios": [u.nome for u in users if (u.nome or "").strip()]})


@missoes_bp.route("/api/missoes/etiquetas", methods=["GET"])
@require_role("admin", "gestor", "tecnico")
def listar_etiquetas():
    """Vocabulário das etiquetas já usadas, com frequência.

    Campo livre fragmenta o dado — "urgência", "urgencia" e "Urgente" viravam
    três etiquetas diferentes e nenhum filtro fechava. Sugerir o que já existe é
    o remédio barato antes de normalizar em tabela."""
    linhas = (db.session.query(MissaoCartao.etiquetas)
              .join(Missao, MissaoCartao.missao_id == Missao.id)
              .filter(Missao.arquivado == False,
                      MissaoCartao.etiquetas != "").all())
    freq = {}
    for (csv,) in linhas:
        for tag in (csv or "").split(","):
            tag = tag.strip()
            if tag:
                freq[tag] = freq.get(tag, 0) + 1
    itens = sorted(freq.items(), key=lambda kv: (-kv[1], kv[0].lower()))
    return jsonify({"etiquetas": [{"nome": n, "n": q} for n, q in itens]})


# ── MEUS CARTÕES (visão cross-missão por responsável) ────────────────────────

@missoes_bp.route("/api/missoes/meus-cartoes", methods=["GET"])
@require_role("admin", "gestor", "tecnico")
def meus_cartoes():
    """Cartões (não concluídos, de missões ativas) atribuídos ao usuário logado."""
    user = User.query.filter_by(email=get_jwt_identity()).first()
    if not user:
        return jsonify({"cartoes": []})
    cartoes = (MissaoCartao.query
               .join(Missao, MissaoCartao.missao_id == Missao.id)
               .filter(Missao.arquivado == False,
                       MissaoCartao.concluido == False,
                       _cond_responsavel(user))
               .order_by(MissaoCartao.prazo == "", MissaoCartao.prazo,
                         MissaoCartao.missao_id, MissaoCartao.ordem)
               .all())
    refs = _mapa_refs(cartoes)
    extras = _mapa_extras(cartoes)
    vazio = {"label": "", "status": "", "status_global": "", "ativo": True}
    out = []
    for c in cartoes:
        d = c.to_dict(ref_info=refs.get((c.ref_tipo, c.ref_id), vazio),
                      extras=extras.get(c.id))
        d["missao_nome"] = c.missao.nome if c.missao else ""
        d["coluna_nome"] = c.coluna.nome if c.coluna else ""
        out.append(d)
    return jsonify({"cartoes": out, "total": len(out),
                    "atrasados": sum(1 for c in cartoes if c.atrasado)})


# ── BUSCA DE REFERÊNCIAS (p/ o seletor de vínculo do modal) ──────────────────

@missoes_bp.route("/api/missoes/refs", methods=["GET"])
@require_role("admin", "gestor", "tecnico")
def buscar_refs():
    """Busca leve por tipo p/ o dropdown de vínculo. ?tipo=equipamento&q=..."""
    tipo = (request.args.get("tipo") or "").strip()
    q = (request.args.get("q") or "").strip()
    if tipo not in REF_TIPOS_CARTAO:
        return jsonify({"erro": f"tipo inválido. Use: {', '.join(REF_TIPOS_CARTAO)}"}), 400
    out = []
    if tipo == "equipamento":
        query = Equipamento.query.filter(Equipamento.ativo == True)
        if q:
            query = query.filter(Equipamento.nome.ilike(f"%{q}%"))
        out = [{"id": e.id, "label": e.nome} for e in
               query.order_by(Equipamento.nome).limit(20).all()]
    elif tipo == "projeto":
        query = Projeto.query.filter(Projeto.ativo == True)
        if q:
            query = query.filter(Projeto.nome.ilike(f"%{q}%"))
        out = [{"id": p.id, "label": p.nome} for p in
               query.order_by(Projeto.nome).limit(20).all()]
    else:  # documento
        query = Documento.query.filter(Documento.ativo == True)
        if q:
            query = query.filter(Documento.documento.ilike(f"%{q}%"))
        out = [{"id": d.id, "label": d.documento} for d in
               query.order_by(Documento.documento).limit(20).all()]
    return jsonify({"itens": out})


# ── CARTÕES VINCULADOS A UMA ENTIDADE (p/ a ficha do documento no dashboard) ──

@missoes_bp.route("/api/missoes/cartoes-vinculados", methods=["GET"])
@require_role("admin", "gestor", "tecnico")
def cartoes_vinculados():
    """Cartões (de missões ativas) vinculados às entidades pedidas, em lote:
    ?tipo=documento&ids=1,2,3 — 1 chamada cobre os documentos do modal inteiro.
    Nota: o dashboard aceita role `leitura`, que aqui recebe 403 — o front deve
    degradar escondendo a seção."""
    tipo = (request.args.get("tipo") or "").strip()
    if tipo not in REF_TIPOS_CARTAO:
        return jsonify({"erro": f"tipo inválido. Use: {', '.join(REF_TIPOS_CARTAO)}"}), 400
    try:
        ids = [int(i) for i in (request.args.get("ids") or "").split(",") if i.strip()]
    except ValueError:
        return jsonify({"erro": "ids deve ser lista de números separados por vírgula"}), 400
    if not ids:
        return jsonify({"cartoes": []})
    cartoes = (MissaoCartao.query
               .join(Missao, MissaoCartao.missao_id == Missao.id)
               .filter(Missao.arquivado == False,
                       MissaoCartao.ref_tipo == tipo,
                       MissaoCartao.ref_id.in_(ids))
               .order_by(MissaoCartao.missao_id, MissaoCartao.ordem)
               .all())
    out = [{"id": c.id, "titulo": (c.titulo or "").strip(),
            "concluido": bool(c.concluido), "prioridade": c.prioridade or "media",
            "atrasado": c.atrasado, "prazo": c.prazo or "",
            "responsaveis": c.responsaveis or "",
            "ref_id": c.ref_id, "missao_id": c.missao_id,
            "missao_nome": c.missao.nome if c.missao else "",
            "coluna_nome": c.coluna.nome if c.coluna else "",
            "coluna_categoria": (c.coluna.categoria or "") if c.coluna else ""}
           for c in cartoes]
    return jsonify({"cartoes": out})


@missoes_bp.route("/api/missoes/cartao-rapido", methods=["POST"])
@require_role("admin", "gestor", "tecnico")
def cartao_rapido():
    """Cria um cartão já vinculado a uma entidade, sem sair da ficha dela.

    O vínculo só nascia de dentro do board: da ficha do documento dava para ver
    os cartões, mas não para abrir um. Cai na primeira coluna 'todo' da missão
    (ou na primeira coluna, se a missão não tiver categorias)."""
    data = request.get_json(silent=True) or {}
    titulo = (data.get("titulo") or "").strip()
    if not titulo:
        return jsonify({"erro": "titulo é obrigatório"}), 400
    erro = _erro_tamanho(titulo, MAX_TITULO, "titulo")
    if erro:
        return jsonify({"erro": erro}), 400
    m = Missao.query.get_or_404(int(data.get("missao_id") or 0))
    if m.arquivado:
        return jsonify({"erro": "missão arquivada"}), 400
    destino = (MissaoColuna.query.filter_by(missao_id=m.id, categoria="todo")
               .order_by(MissaoColuna.ordem).first()) or \
              (MissaoColuna.query.filter_by(missao_id=m.id)
               .order_by(MissaoColuna.ordem).first())
    if destino is None:
        return jsonify({"erro": "a missão não tem nenhuma coluna"}), 400
    email = get_jwt_identity()
    agora = datetime.now()
    ult = (MissaoCartao.query.filter_by(coluna_id=destino.id)
           .order_by(MissaoCartao.ordem.desc()).first())
    cartao = MissaoCartao(missao_id=m.id, coluna_id=destino.id, titulo=titulo,
                          criado_por=email, atualizado_por=email,
                          criado_em=agora, entrou_coluna_em=agora,
                          ordem=(ult.ordem + 1) if ult else 0)
    erro = _aplicar_campos_cartao(cartao, data)
    if erro:
        return jsonify({"erro": erro}), 400
    db.session.add(cartao)
    db.session.flush()
    _sync_responsaveis(cartao)
    _hist(cartao, "criado", por=email, em=agora, coluna_destino_id=destino.id,
          campo="titulo", novo=titulo)
    db.session.commit()
    log_action(email, "CREATE", entidade=f"Cartão: {titulo}", ip=get_client_ip())
    _emit("MISSAO_CARTAO_CREATED",
          {"missao_id": m.id, "cartao": cartao.to_dict()}, email, ip=get_client_ip())
    return jsonify({"cartao": cartao.to_dict(), "missao_nome": m.nome,
                    "coluna_nome": destino.nome}), 201


# ── MÉTRICAS ─────────────────────────────────────────────────────────────────

def _percentil(valores, p):
    """Percentil por interpolação linear (p em 0..1). p85 do cycle time é a
    promessa realista de prazo; a média sozinha esconde a cauda."""
    if not valores:
        return None
    vs = sorted(valores)
    k = (len(vs) - 1) * p
    baixo = int(k)
    alto = min(baixo + 1, len(vs) - 1)
    return round(vs[baixo] + (vs[alto] - vs[baixo]) * (k - baixo), 1)


def _tempo_por_coluna(missao_id, colunas, desde=None):
    """Tempo médio de permanência por coluna, em dias.

    Intervalos fechados vêm do histórico (evento → evento seguinte do mesmo
    cartão); o trecho aberto do cartão que ainda está parado entra pelo
    `entrou_coluna_em`. É a leitura que aponta o gargalo do fluxo."""
    q = (MissaoCartaoHistorico.query
         .filter(MissaoCartaoHistorico.missao_id == missao_id,
                 MissaoCartaoHistorico.evento.in_(("criado", "movido"))))
    if desde:
        q = q.filter(MissaoCartaoHistorico.em >= desde)
    eventos = q.order_by(MissaoCartaoHistorico.cartao_id,
                         MissaoCartaoHistorico.em).all()
    acc = {c.id: {"dias": 0.0, "n": 0} for c in colunas}
    por_cartao = {}
    for ev in eventos:
        por_cartao.setdefault(ev.cartao_id, []).append(ev)
    for evs in por_cartao.values():
        for atual, seguinte in zip(evs, evs[1:]):
            alvo = acc.get(atual.coluna_destino_id)
            if alvo is None or not atual.em or not seguinte.em:
                continue
            alvo["dias"] += (seguinte.em - atual.em).total_seconds() / 86400.0
            alvo["n"] += 1
    # trecho ainda em curso (cartão aberto parado na coluna atual)
    agora = datetime.now()
    for c in (MissaoCartao.query.filter_by(missao_id=missao_id, concluido=False)
              .all()):
        alvo = acc.get(c.coluna_id)
        base = c.entrou_coluna_em or c.criado_em
        if alvo is None or not base:
            continue
        alvo["dias"] += (agora - base).total_seconds() / 86400.0
        alvo["n"] += 1
    return {cid: {"media": round(v["dias"] / v["n"], 1) if v["n"] else 0.0,
                  "amostras": v["n"]}
            for cid, v in acc.items()}


def _metricas_missao(m, dias=30):
    cartoes = MissaoCartao.query.filter_by(missao_id=m.id).all()
    colunas = list(m.colunas)
    corte = datetime.now() - timedelta(days=dias)
    abertos = [c for c in cartoes if not c.concluido]
    concluidos = [c for c in cartoes if c.concluido]
    por_col = {c.id: [] for c in colunas}
    for c in cartoes:
        por_col.setdefault(c.coluna_id, []).append(c)

    tempos = _tempo_por_coluna(m.id, colunas)
    lista_colunas = []
    wip = 0
    for col in colunas:
        itens = por_col.get(col.id, [])
        n_abertos = len([c for c in itens if not c.concluido])
        if (col.categoria or "") == "doing":
            wip += n_abertos
        lista_colunas.append({
            "id": col.id, "nome": col.nome, "categoria": col.categoria or "",
            "limite_wip": col.limite_wip or 0,
            "total": len(itens), "abertos": n_abertos,
            "excedido": bool(col.limite_wip and n_abertos > col.limite_wip),
            "dias_medios": tempos.get(col.id, {}).get("media", 0.0),
            "amostras": tempos.get(col.id, {}).get("amostras", 0),
        })

    carga = {}
    for c in abertos:
        nomes = [n.strip() for n in (c.responsaveis or "").split(",") if n.strip()]
        for n in nomes or ["(sem responsável)"]:
            reg = carga.setdefault(n, {"nome": n, "abertos": 0, "atrasados": 0,
                                       "peso": 0.0})
            reg["abertos"] += 1
            reg["peso"] += (c.peso if c.peso is not None else 1.0)
            if c.atrasado:
                reg["atrasados"] += 1
    for reg in carga.values():
        reg["peso"] = round(reg["peso"], 1)

    janela = [c for c in concluidos if c.concluido_em and c.concluido_em >= corte]
    semanas = {}
    for c in janela:
        chave = c.concluido_em.date().isocalendar()
        rotulo = f"{chave[0]}-S{chave[1]:02d}"
        semanas[rotulo] = semanas.get(rotulo, 0) + 1
    ciclos = [c.dias_ciclo for c in janela if c.dias_ciclo is not None]

    peso_total = sum((c.peso if c.peso is not None else 1.0) for c in cartoes)
    peso_feito = sum((c.peso if c.peso is not None else 1.0) for c in concluidos)

    aging = sorted(abertos, key=lambda c: c.dias_parado, reverse=True)[:8]
    return {
        "missao_id": m.id, "missao": m.nome, "janela_dias": dias,
        "totais": {
            "total": len(cartoes), "abertos": len(abertos),
            "concluidos": len(concluidos),
            "atrasados": sum(1 for c in abertos if c.atrasado),
            "sem_responsavel": sum(1 for c in abertos if not (c.responsaveis or "").strip()),
            "sem_prazo": sum(1 for c in abertos if not (c.prazo or "").strip()),
            "wip": wip,
        },
        "avanco": {
            "por_cartao": round(100 * len(concluidos) / len(cartoes)) if cartoes else 0,
            # Ponderado: um cartão de 3 meses não vale o mesmo que um de 10 min.
            "ponderado": round(100 * peso_feito / peso_total) if peso_total else 0,
            "peso_total": round(peso_total, 1), "peso_concluido": round(peso_feito, 1),
        },
        "por_coluna": lista_colunas,
        "por_prioridade": {p: sum(1 for c in abertos if (c.prioridade or "media") == p)
                           for p in PRIORIDADES_CARTAO},
        "por_responsavel": sorted(carga.values(),
                                  key=lambda r: (-r["abertos"], r["nome"])),
        "throughput": {
            "concluidos": len(janela),
            "por_semana": [{"semana": k, "n": semanas[k]} for k in sorted(semanas)],
        },
        "cycle_time": {
            "amostra": len(ciclos),
            "media": round(sum(ciclos) / len(ciclos), 1) if ciclos else None,
            "p50": _percentil(ciclos, 0.50), "p85": _percentil(ciclos, 0.85),
        },
        "aging": [{"cartao_id": c.id, "titulo": c.titulo, "dias": c.dias_parado,
                   "coluna": c.coluna.nome if c.coluna else "",
                   "responsaveis": c.responsaveis or ""} for c in aging],
    }


@missoes_bp.route("/api/missoes/<int:mid>/metricas", methods=["GET"])
@require_role("admin", "gestor", "tecnico")
def metricas(mid):
    m = Missao.query.get_or_404(mid)
    dias = max(1, min(request.args.get("dias", default=30, type=int) or 30, 365))
    return jsonify(_metricas_missao(m, dias))


# ── ALERTAS (o que precisa de atenção hoje) ──────────────────────────────────

@missoes_bp.route("/api/missoes/alertas", methods=["GET"])
@require_role("admin", "gestor", "tecnico")
def alertas():
    """Fatos acionáveis derivados do que já está no banco — mesmo formato dos
    alertas de projetos ({tipo, severidade, titulo, detalhe}).

    ?dias_parado=N (padrão 7) · ?meus=1 restringe aos cartões do usuário logado.
    """
    dias_parado = max(1, request.args.get("dias_parado", default=7, type=int) or 7)
    so_meus = request.args.get("meus", "").strip() == "1"
    user = User.query.filter_by(email=get_jwt_identity()).first()

    q = (MissaoCartao.query.join(Missao, MissaoCartao.missao_id == Missao.id)
         .filter(Missao.arquivado == False, MissaoCartao.concluido == False))
    if so_meus and user:
        q = q.filter(_cond_responsavel(user))
    cartoes = q.all()
    refs = _mapa_refs(cartoes)

    itens = []

    def add(cartao, tipo, sev, titulo, detalhe):
        itens.append({
            "tipo": tipo, "severidade": sev, "titulo": titulo, "detalhe": detalhe,
            "missao_id": cartao.missao_id,
            "missao": cartao.missao.nome if cartao.missao else "",
            "cartao_id": cartao.id, "cartao": (cartao.titulo or "").strip(),
            "coluna": cartao.coluna.nome if cartao.coluna else "",
        })

    hoje = date.today()
    for c in cartoes:
        if c.atrasado:
            d = _parse_data(c.prazo)
            add(c, "cartao_vencido", "critico", "Prazo vencido",
                f"previsto para {d.strftime('%d/%m/%Y')} · {(hoje - d).days} dia(s) em atraso")
        if not (c.responsaveis or "").strip():
            add(c, "cartao_sem_responsavel", "atencao", "Sem responsável",
                "ninguém atribuído — o cartão não tem a quem cobrar")
        if c.dias_parado >= dias_parado:
            add(c, "cartao_parado", "atencao", "Sem movimentação",
                f"parado há {c.dias_parado} dia(s) em \"{c.coluna.nome if c.coluna else ''}\"")
        # A coluna é o estado: cartão na coluna de concluído com o campo em falso
        # some do throughput e continua contando como WIP.
        if c.coluna is not None and (c.coluna.categoria or "") == "done":
            add(c, "cartao_done_nao_concluido", "atencao", "Na coluna de concluído, mas aberto",
                f"está em \"{c.coluna.nome}\" sem estar marcado como concluído")
        if c.ref_tipo and c.ref_id:
            meta = refs.get((c.ref_tipo, c.ref_id))
            if meta is None:
                add(c, "vinculo_inexistente", "atencao", "Vínculo quebrado",
                    f"o {c.ref_tipo} #{c.ref_id} não existe mais")
            elif not meta.get("ativo", True):
                add(c, "vinculo_inativo", "atencao", "Vínculo desativado",
                    f"{c.ref_tipo} \"{meta.get('label') or c.ref_id}\" foi desativado")

    # WIP por coluna: sinal de fluxo, não de cartão — entra uma vez por coluna.
    if not so_meus:
        colunas = (MissaoColuna.query.join(Missao, MissaoColuna.missao_id == Missao.id)
                   .filter(Missao.arquivado == False, MissaoColuna.limite_wip > 0).all())
        abertos_por_col = {}
        for c in cartoes:
            abertos_por_col[c.coluna_id] = abertos_por_col.get(c.coluna_id, 0) + 1
        for col in colunas:
            n = abertos_por_col.get(col.id, 0)
            if n > (col.limite_wip or 0):
                itens.append({
                    "tipo": "wip_excedido", "severidade": "atencao",
                    "titulo": "Limite de WIP estourado",
                    "detalhe": f"{n} cartões abertos em \"{col.nome}\" (limite {col.limite_wip})",
                    "missao_id": col.missao_id,
                    "missao": col.missao.nome if col.missao else "",
                    "cartao_id": None, "cartao": "", "coluna": col.nome,
                })

    ordem = {"critico": 0, "atencao": 1, "info": 2}
    itens.sort(key=lambda a: (ordem.get(a["severidade"], 9), a["missao"], a["cartao"]))
    return jsonify({"alertas": itens, "total": len(itens),
                    "criticos": sum(1 for i in itens if i["severidade"] == "critico")})


# ── HISTÓRICO E SNAPSHOTS ────────────────────────────────────────────────────

@missoes_bp.route("/api/missoes/<int:mid>/historico", methods=["GET"])
@require_role("admin", "gestor", "tecnico")
def historico_missao(mid):
    Missao.query.get_or_404(mid)
    limite = max(1, min(request.args.get("limite", default=200, type=int) or 200, 1000))
    q = MissaoCartaoHistorico.query.filter_by(missao_id=mid)
    cartao_id = request.args.get("cartao_id", type=int)
    if cartao_id:
        q = q.filter_by(cartao_id=cartao_id)
    linhas = q.order_by(MissaoCartaoHistorico.em.desc()).limit(limite).all()
    nomes = {c.id: c.nome for c in MissaoColuna.query.filter_by(missao_id=mid).all()}
    titulos = {c.id: (c.titulo or "").strip()
               for c in MissaoCartao.query.filter_by(missao_id=mid).all()}
    out = []
    for h in linhas:
        d = h.to_dict(nomes_coluna=nomes)
        d["cartao"] = titulos.get(h.cartao_id, "")
        out.append(d)
    return jsonify({"historico": out, "total": len(out)})


@missoes_bp.route("/api/missoes/cartoes/<int:cid>/historico", methods=["GET"])
@require_role("admin", "gestor", "tecnico")
def historico_cartao(cid):
    cartao = MissaoCartao.query.get_or_404(cid)
    nomes = {c.id: c.nome for c in
             MissaoColuna.query.filter_by(missao_id=cartao.missao_id).all()}
    linhas = (MissaoCartaoHistorico.query.filter_by(cartao_id=cid)
              .order_by(MissaoCartaoHistorico.em.desc()).limit(200).all())
    return jsonify({"historico": [h.to_dict(nomes_coluna=nomes) for h in linhas]})


def _snapshot_missao(m, dia):
    """Grava/atualiza a foto do dia. Idempotente (UNIQUE missao_id+data)."""
    cartoes = MissaoCartao.query.filter_by(missao_id=m.id).all()
    if not cartoes:
        return None
    abertos = [c for c in cartoes if not c.concluido]
    doing = {c.id for c in m.colunas if (c.categoria or "") == "doing"}
    snap = MissaoSnapshot.query.filter_by(missao_id=m.id, data=dia).first()
    if snap is None:
        snap = MissaoSnapshot(missao_id=m.id, data=dia)
        db.session.add(snap)
    snap.total = len(cartoes)
    snap.abertos = len(abertos)
    snap.concluidos = len(cartoes) - len(abertos)
    snap.atrasados = sum(1 for c in abertos if c.atrasado)
    snap.wip = sum(1 for c in abertos if c.coluna_id in doing)
    snap.sem_responsavel = sum(1 for c in abertos if not (c.responsaveis or "").strip())
    snap.peso_total = sum((c.peso if c.peso is not None else 1.0) for c in cartoes)
    snap.peso_concluido = sum((c.peso if c.peso is not None else 1.0)
                              for c in cartoes if c.concluido)
    snap.criado_em = datetime.now()
    return snap


def snapshot_do_dia(dia=None):
    """Foto de todas as missões ativas. Chamado na subida do servidor e pelo
    endpoint (cron) — recalcular com as datas de hoje reescreveria o passado."""
    dia = dia or date.today().isoformat()
    n = 0
    for m in Missao.query.filter_by(arquivado=False).all():
        if _snapshot_missao(m, dia) is not None:
            n += 1
    if n:
        db.session.commit()
    return n


@missoes_bp.route("/api/missoes/snapshot", methods=["POST"])
@require_role("admin", "gestor")
def gravar_snapshot():
    n = snapshot_do_dia()
    return jsonify({"ok": True, "missoes": n, "data": date.today().isoformat()})


@missoes_bp.route("/api/missoes/<int:mid>/snapshots", methods=["GET"])
@require_role("admin", "gestor", "tecnico")
def listar_snapshots(mid):
    Missao.query.get_or_404(mid)
    dias = max(1, min(request.args.get("dias", default=90, type=int) or 90, 730))
    desde = (date.today() - timedelta(days=dias)).isoformat()
    linhas = (MissaoSnapshot.query
              .filter(MissaoSnapshot.missao_id == mid, MissaoSnapshot.data >= desde)
              .order_by(MissaoSnapshot.data).all())
    return jsonify({"snapshots": [s.to_dict() for s in linhas]})


# ── MODELOS DE MISSÃO ────────────────────────────────────────────────────────

def _aplicar_modelo(m, modelo, email):
    """Materializa as colunas (e cartões) do template na missão recém-criada."""
    agora = datetime.now()
    for i, col in enumerate(modelo.colunas()):
        categoria = (col.get("categoria") or "").strip()
        coluna = MissaoColuna(
            missao_id=m.id, nome=(col.get("nome") or f"Coluna {i + 1}")[:MAX_NOME_COLUNA],
            categoria=categoria if categoria in CATEGORIAS_COLUNA else "",
            limite_wip=int(col.get("limite_wip") or 0), ordem=i)
        db.session.add(coluna)
        db.session.flush()
        for j, card in enumerate(col.get("cartoes") or []):
            titulo = (card.get("titulo") or "").strip()
            if not titulo:
                continue
            pri = (card.get("prioridade") or "media").strip()
            cartao = MissaoCartao(
                missao_id=m.id, coluna_id=coluna.id, titulo=titulo[:MAX_TITULO],
                descricao=(card.get("descricao") or "")[:4000],
                etiquetas=(card.get("etiquetas") or "")[:MAX_ETIQUETAS],
                prioridade=pri if pri in PRIORIDADES_CARTAO else "media",
                peso=float(card.get("peso") or 1.0),
                criado_por=email, atualizado_por=email,
                criado_em=agora, entrou_coluna_em=agora, ordem=j)
            db.session.add(cartao)
            db.session.flush()
            _hist(cartao, "criado", origem="modelo", por=email, em=agora,
                  coluna_destino_id=coluna.id, campo="modelo", novo=modelo.nome)


@missoes_bp.route("/api/missoes/modelos", methods=["GET"])
@require_role("admin", "gestor", "tecnico")
def listar_modelos():
    com_estrutura = request.args.get("estrutura", "").strip() == "1"
    modelos = MissaoModelo.query.order_by(MissaoModelo.nome).all()
    return jsonify({"modelos": [m.to_dict(com_estrutura=com_estrutura) for m in modelos]})


@missoes_bp.route("/api/missoes/modelos", methods=["POST"])
@require_role("admin", "gestor", "tecnico")
def criar_modelo():
    """Salva uma missão existente como template ({missao_id, nome, com_cartoes}).
    Processos daqui se repetem (validação de equipamento novo, submissão ANVISA)
    e toda missão nascia com as mesmas 3 colunas vazias."""
    data = request.get_json(silent=True) or {}
    nome = (data.get("nome") or "").strip()
    if not nome:
        return jsonify({"erro": "nome é obrigatório"}), 400
    erro = _erro_tamanho(nome, MAX_NOME_MISSAO, "nome")
    if erro:
        return jsonify({"erro": erro}), 400
    m = Missao.query.get_or_404(int(data.get("missao_id") or 0))
    com_cartoes = bool(data.get("com_cartoes", True))
    estrutura = []
    for col in m.colunas:
        item = {"nome": col.nome, "categoria": col.categoria or "",
                "limite_wip": col.limite_wip or 0, "cartoes": []}
        if com_cartoes:
            item["cartoes"] = [
                {"titulo": c.titulo, "descricao": c.descricao or "",
                 "prioridade": c.prioridade or "media", "etiquetas": c.etiquetas or "",
                 "peso": c.peso if c.peso is not None else 1.0}
                for c in col.cartoes if not c.concluido]
        estrutura.append(item)
    email = get_jwt_identity()
    modelo = MissaoModelo(nome=nome,
                          descricao=(data.get("descricao") or "").strip(),
                          accent=m.accent or "",
                          estrutura=json.dumps(estrutura, ensure_ascii=False),
                          criado_por=email)
    db.session.add(modelo)
    db.session.commit()
    log_action(email, "CREATE", entidade=f"Modelo de missão: {nome}", ip=get_client_ip())
    _emit("MISSAO_MODELO_CREATED", {"modelo": modelo.to_dict()}, email)
    return jsonify({"modelo": modelo.to_dict()}), 201


@missoes_bp.route("/api/missoes/modelos/<int:mid>", methods=["DELETE"])
@require_role("admin", "gestor")
def excluir_modelo(mid):
    modelo = MissaoModelo.query.get_or_404(mid)
    nome = modelo.nome
    db.session.delete(modelo)
    db.session.commit()
    log_action(get_jwt_identity(), "DELETE", entidade=f"Modelo de missão: {nome}",
               ip=get_client_ip())
    return jsonify({"ok": True})


# ── EXPORT EXCEL ─────────────────────────────────────────────────────────────

@missoes_bp.route("/api/missoes/<int:mid>/export", methods=["GET"])
@require_role("admin", "gestor", "tecnico")
def exportar_missao(mid):
    """Cartões + métricas + trilha, para dinâmica/BI fora do sistema."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    m = Missao.query.get_or_404(mid)
    met = _metricas_missao(m, dias=max(1, min(
        request.args.get("dias", default=30, type=int) or 30, 365)))
    cab = Font(bold=True, color="FFFFFF")
    azul = PatternFill("solid", fgColor="1F4E5F")

    def escrever_cabecalho(ws, colunas):
        for j, h in enumerate(colunas, 1):
            c = ws.cell(row=1, column=j, value=h)
            c.font = cab
            c.fill = azul
        ws.freeze_panes = "A2"

    wb = Workbook()
    ws = wb.active
    ws.title = "Cartões"
    cols = ["Coluna", "Categoria", "Cartão", "Status", "Prioridade", "Peso",
            "Responsáveis", "Etiquetas", "Início", "Prazo", "Atrasado",
            "Dias parado", "Ciclo (dias)", "Vínculo", "Recorrência",
            "Criado em", "Concluído em", "Criado por"]
    escrever_cabecalho(ws, cols)
    linha = 2
    for col in m.colunas:
        for c in col.cartoes:
            valores = [
                col.nome, col.categoria or "", (c.titulo or "").strip(),
                "Concluído" if c.concluido else "Aberto", c.prioridade or "media",
                c.peso if c.peso is not None else 1.0, c.responsaveis or "",
                c.etiquetas or "", c.data_inicio or "", c.prazo or "",
                "Sim" if c.atrasado else "", c.dias_parado,
                c.dias_ciclo if c.dias_ciclo is not None else "",
                f"{c.ref_tipo} #{c.ref_id}" if c.ref_tipo and c.ref_id else "",
                c.recorrencia or "",
                c.criado_em.strftime("%d/%m/%Y %H:%M") if c.criado_em else "",
                c.concluido_em.strftime("%d/%m/%Y %H:%M") if c.concluido_em else "",
                c.criado_por or "",
            ]
            for j, v in enumerate(valores, 1):
                ws.cell(row=linha, column=j, value=v)
            linha += 1
    ws.column_dimensions["C"].width = 44
    for letra in ("A", "G", "H"):
        ws.column_dimensions[letra].width = 22

    wm = wb.create_sheet("Métricas")
    wm["A1"] = f"Missão: {m.nome}"
    wm["A1"].font = Font(bold=True, size=13)
    wm["A2"] = f"Janela de análise: {met['janela_dias']} dias · gerado em " \
               f"{datetime.now().strftime('%d/%m/%Y %H:%M')}"
    linha = 4
    blocos = [
        ("Totais", [(k.replace("_", " ").capitalize(), v)
                    for k, v in met["totais"].items()]),
        ("Avanço", [("Por cartão (%)", met["avanco"]["por_cartao"]),
                    ("Ponderado por peso (%)", met["avanco"]["ponderado"]),
                    ("Peso total", met["avanco"]["peso_total"]),
                    ("Peso concluído", met["avanco"]["peso_concluido"])]),
        ("Cycle time (dias)", [("Amostra", met["cycle_time"]["amostra"]),
                               ("Média", met["cycle_time"]["media"]),
                               ("P50", met["cycle_time"]["p50"]),
                               ("P85", met["cycle_time"]["p85"])]),
        ("Throughput", [("Concluídos na janela", met["throughput"]["concluidos"])] +
                       [(s["semana"], s["n"]) for s in met["throughput"]["por_semana"]]),
        ("Abertos por prioridade", list(met["por_prioridade"].items())),
    ]
    for titulo, pares in blocos:
        wm.cell(row=linha, column=1, value=titulo).font = Font(bold=True)
        linha += 1
        for rotulo, valor in pares:
            wm.cell(row=linha, column=1, value=str(rotulo))
            wm.cell(row=linha, column=2, value=valor)
            linha += 1
        linha += 1
    wm.cell(row=linha, column=1, value="Por coluna").font = Font(bold=True)
    linha += 1
    for h, j in zip(["Coluna", "Total", "Abertos", "Limite WIP", "Excedido",
                     "Dias médios"], range(1, 7)):
        wm.cell(row=linha, column=j, value=h).font = Font(bold=True)
    linha += 1
    for col in met["por_coluna"]:
        for j, v in enumerate([col["nome"], col["total"], col["abertos"],
                               col["limite_wip"], "Sim" if col["excedido"] else "",
                               col["dias_medios"]], 1):
            wm.cell(row=linha, column=j, value=v)
        linha += 1
    linha += 1
    wm.cell(row=linha, column=1, value="Carga por responsável").font = Font(bold=True)
    linha += 1
    for h, j in zip(["Responsável", "Abertos", "Atrasados", "Peso"], range(1, 5)):
        wm.cell(row=linha, column=j, value=h).font = Font(bold=True)
    linha += 1
    for r in met["por_responsavel"]:
        for j, v in enumerate([r["nome"], r["abertos"], r["atrasados"], r["peso"]], 1):
            wm.cell(row=linha, column=j, value=v)
        linha += 1
    wm.column_dimensions["A"].width = 30
    wm.column_dimensions["B"].width = 16

    wh = wb.create_sheet("Histórico")
    escrever_cabecalho(wh, ["Quando", "Cartão", "Evento", "De", "Para", "Campo",
                            "Antes", "Depois", "Origem", "Por"])
    nomes = {c.id: c.nome for c in m.colunas}
    titulos = {c.id: (c.titulo or "").strip() for c in m.cartoes}
    linha = 2
    for h in (MissaoCartaoHistorico.query.filter_by(missao_id=m.id)
              .order_by(MissaoCartaoHistorico.em.desc()).limit(5000).all()):
        valores = [h.em.strftime("%d/%m/%Y %H:%M") if h.em else "",
                   titulos.get(h.cartao_id, ""), h.evento or "",
                   nomes.get(h.coluna_origem_id, ""), nomes.get(h.coluna_destino_id, ""),
                   h.campo or "", h.valor_antigo or "", h.valor_novo or "",
                   h.origem or "", h.por or ""]
        for j, v in enumerate(valores, 1):
            wh.cell(row=linha, column=j, value=v)
        linha += 1
    wh.column_dimensions["B"].width = 40
    for j in range(1, 11):
        if j != 2:
            wh.column_dimensions[get_column_letter(j)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # Content-Disposition sem acento: "Missão" vira "Missao", não "Miss_o".
    sem_acento = (unicodedata.normalize("NFKD", m.nome or "missao")
                  .encode("ascii", "ignore").decode())
    seguro = re.sub(r"[^A-Za-z0-9_-]+", "_", sem_acento).strip("_") or "missao"
    nome_arq = f"Missao_{seguro}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=nome_arq,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

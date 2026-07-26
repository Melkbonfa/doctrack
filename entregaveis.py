"""
entregaveis.py — Módulo de Projetos / Entregáveis (PMO)

Rotas:
  GET    /api/projetos                — lista com avanço calculado + filtros
  POST   /api/projetos                — criar (admin/gestor)
  GET    /api/projetos/<id>           — detalhe agrupado por categoria
  PUT    /api/projetos/<id>           — editar metadados (admin/gestor)
  DELETE /api/projetos/<id>           — arquivar (admin/gestor)
  POST   /api/projetos/<id>/restaurar — desarquivar (admin/gestor)
  GET    /api/projetos/alertas        — atrasos, estouros e projetos parados
  GET    /api/projetos/<id>/baselines — histórico de linhas de base
  POST   /api/projetos/<id>/baselines — congela nova linha de base (admin/gestor)
  GET    /api/projetos/<id>/mensal    — série mensal + métricas EVM
  PUT    /api/projetos/<id>/mensal    — lança custo do mês (admin/gestor)
  DELETE /api/projetos/<id>/mensal/<competencia>
  PUT    /api/entregaveis/<id>        — status/percentual/responsáveis
                                        (admin/gestor; técnico só nos seus)
  DELETE /api/entregaveis/<id>        — excluir entregável (admin/gestor)
  POST   /api/projetos/<id>/entregaveis — adicionar entregável (admin/gestor)
  GET    /api/entregaveis/responsaveis — usuários atribuíveis (picker)
  GET    /api/modelos                 — modelos de entregáveis por tipo (OEM/Revenda)
  POST   /api/modelos                 — adicionar item de modelo (admin/gestor)
  PUT    /api/modelos/<id>            — editar item de modelo (admin/gestor)
  DELETE /api/modelos/<id>            — excluir item de modelo (admin/gestor)
  GET    /api/entregaveis/resumo      — KPIs e visão por responsável
  GET    /api/entregaveis/export      — Excel (entregáveis + aba PMO)

Perfis: o módulo é de gestão (admin/gestor). O técnico entra em modo restrito —
enxerga apenas os projetos onde tem entregáveis atribuídos, sem nenhum valor em
R$, e só pode mexer no status/percentual dos entregáveis que são dele.
"""
import io
import re
from datetime import datetime

from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import get_jwt_identity
from sqlalchemy.orm import selectinload

from models import (db, Projeto, Entregavel, EntregavelHistorico, ProjetoMensal,
                    ProjetoBaseline, ModeloEntregavel, User,
                    CATEGORIAS_ENTREGAVEL, STATUS_ENTREGAVEL, MOSCOW,
                    TIPOS_PROJETO, STATUS_PROJETO, STATUS_PROJETO_ABERTO,
                    _parse_iso)
from auth import require_role, get_client_ip

entregaveis_bp = Blueprint("entregaveis", __name__)

# Campos de cronograma do projeto (datas ISO em texto livre)
DATAS_PROJETO = ("data_inicio_prev", "data_inicio_real", "data_fim_prev", "data_fim_real")
# Campos que compõem a linha de base: mexer neles versiona a baseline.
CAMPOS_BASELINE = ("data_inicio_prev", "data_fim_prev", "orcamento")

_RE_COMPET = re.compile(r"^\d{4}-(0[1-9]|1[0-2])$")   # 'YYYY-MM'
_RE_ISO    = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _parse_orcamento(v):
    """Aceita número ou string ('1.234,56' / '1234.56'); retorna float >= 0 ou None se inválido."""
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v) if v >= 0 else None
    s = str(v).strip().replace("R$", "").replace(" ", "")
    # 1.234,56 -> 1234.56 ; 1234.56 -> 1234.56
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        f = float(s)
    except ValueError:
        return None
    return f if f >= 0 else None


def _parse_peso(v):
    """Peso (esforço relativo) > 0. None quando inválido; 1.0 quando ausente."""
    if v is None or v == "":
        return 1.0
    try:
        f = float(str(v).replace(",", "."))
    except (TypeError, ValueError):
        return None
    return f if f > 0 else None


def _parse_data(v):
    """'' ou ISO yyyy-mm-dd. Devolve (valor, erro)."""
    s = (v or "").strip()
    if not s:
        return "", None
    if not _RE_ISO.match(s) or _parse_iso(s) is None:
        return None, "data inválida (use AAAA-MM-DD)"
    return s, None


# ── Perfil do requisitante ───────────────────────────────────────────────────

def _usuario_atual():
    return User.query.filter_by(email=get_jwt_identity(), ativo=True).first()


def _role():
    u = _usuario_atual()
    return u.role if u else ""


def pode_ver_financeiro(role=None):
    """Só gestão vê dinheiro. Técnico usa o módulo para tocar as tarefas dele."""
    return (role or _role()) in ("admin", "gestor")


def pode_editar_projeto(role=None):
    return (role or _role()) in ("admin", "gestor")


def _responsavel_por(entregavel, user):
    """True se `user` está atribuído ao entregável (FK) ou aparece no texto legado."""
    if user is None:
        return False
    if any(u.id == user.id for u in entregavel.responsaveis_users):
        return True
    # Projetos antigos ainda só têm o texto livre: compara pelo primeiro nome.
    texto = (entregavel.responsaveis or "").lower()
    if not texto:
        return False
    primeiro = (user.nome or "").strip().split(" ")[0].lower()
    return bool(primeiro) and primeiro in texto


# ── Queries com carga antecipada (evita N+1) ─────────────────────────────────
# to_dict() toca entregaveis + mensais de cada projeto; sem isto, listar 50
# projetos disparava ~100 SELECTs extras.

def _query_projetos():
    return Projeto.query.options(
        selectinload(Projeto.entregaveis).selectinload(Entregavel.responsaveis_users),
        selectinload(Projeto.mensais),
    )


def _get_projeto(pid, com_pmo=False):
    q = _query_projetos()
    if com_pmo:
        q = q.options(selectinload(Projeto.baselines),
                      selectinload(Projeto.snapshots),
                      selectinload(Projeto.entregaveis).selectinload(Entregavel.historico))
    p = q.filter_by(id=pid).first()
    if p is None:
        from flask import abort
        abort(404)
    return p


# preenchido por servidor.py para emitir tempo real sem import circular
_rt = {"socketio": None, "publish_event": None, "AuditLog": None, "EventType": None}


def init_realtime(socketio, publish_event, AuditLog, EventType):
    _rt.update(socketio=socketio, publish_event=publish_event,
               AuditLog=AuditLog, EventType=EventType)


def _emit(event_type, payload, email, *, campo="", antigo=None, novo=None,
          entidade="", ip=None):
    """Publica o evento E grava a auditoria — uma linha só.

    Antes cada mutação chamava log_action() *e* publish_event(), gerando duas
    entradas em audit_logs para o mesmo fato. publish_event já persiste tudo o
    que log_action persistia (entidade/campo/antigo/novo/ip), então aqui é o
    único ponto de escrita.
    """
    if not (_rt["socketio"] and _rt["publish_event"]):
        return
    corpo = dict(payload)
    if entidade:
        corpo["entidade"] = entidade
    if antigo is not None:
        corpo["old_value"] = antigo
    if novo is not None:
        corpo["new_value"] = novo
    user = User.query.filter_by(email=email).first() if email else None
    try:
        _rt["publish_event"](event_type, corpo, user_id=(user.id if user else None),
                             user_email=email, db=db, AuditLog=_rt["AuditLog"],
                             socketio=_rt["socketio"], campo=campo,
                             ip=(get_client_ip() if ip is None else ip))
    except Exception:
        pass  # auditoria/tempo real é best-effort; a gravação de domínio já foi feita


# ── PROJETOS ─────────────────────────────────────────────────────────────────

@entregaveis_bp.route("/api/projetos", methods=["GET"])
@require_role("admin", "gestor", "tecnico")
def listar_projetos():
    user = _usuario_atual()
    role = user.role if user else ""
    financeiro = pode_ver_financeiro(role)

    # arquivados=1 lista os projetos arquivados (ativo=False); padrão = ativos.
    arquivados = request.args.get("arquivados", "").strip() == "1"
    q = _query_projetos().filter_by(ativo=not arquivados)

    ano = request.args.get("ano", type=int)
    if ano:
        q = q.filter_by(ano=ano)
    moscow = request.args.get("moscow", "").strip()
    if moscow:
        q = q.filter(Projeto.moscow.ilike(moscow))
    status = request.args.get("status", "").strip()
    if status == "abertos":
        q = q.filter(Projeto.status.in_(STATUS_PROJETO_ABERTO))
    elif status:
        if status not in STATUS_PROJETO:
            return jsonify({"erro": f"status inválido. Use: {', '.join(STATUS_PROJETO)}"}), 400
        q = q.filter_by(status=status)
    tipo = request.args.get("tipo", "").strip()
    if tipo:
        q = q.filter(Projeto.tipo.ilike(tipo))
    busca = request.args.get("busca", "").strip()
    if busca:
        alvo = f"%{busca}%"
        q = q.filter(db.or_(Projeto.nome.ilike(alvo), Projeto.sku.ilike(alvo)))

    projetos = q.all()
    com_ent = request.args.get("com_entregaveis", "").strip() == "1"
    resp = request.args.get("responsavel", "").strip().lower()

    out = []
    for p in projetos:
        # Técnico só enxerga projeto onde tem entregável atribuído.
        if role == "tecnico" and not any(_responsavel_por(e, user) for e in p.entregaveis):
            continue
        d = p.to_dict(com_entregaveis=com_ent, com_financeiro=financeiro)
        if resp:
            tipos = [e.to_dict() for e in p.entregaveis
                     if e.status != "na" and (
                         resp in (e.responsaveis or "").lower()
                         or any(resp in (u.nome or "").lower() or resp == (u.email or "").lower()
                                for u in e.responsaveis_users))]
            if not tipos:
                continue
            d["entregaveis_do_responsavel"] = tipos
        out.append(d)
    out.sort(key=lambda d: (d["prioridade"] or 999, d["nome"]))
    return jsonify({"projetos": out, "financeiro": financeiro, "role": role})


def _aplicar_responsaveis(e, ids):
    """Substitui os responsáveis (FK) de um entregável por `ids`. Mantém o campo
    texto sincronizado para as telas/exports que ainda o exibem."""
    if ids is None:
        return
    users = User.query.filter(User.id.in_(ids or []), User.ativo.is_(True)).all() if ids else []
    e.responsaveis_users = users
    if users:
        e.responsaveis = "/".join(u.nome.split(" ")[0] for u in users)


def _entregaveis_do_payload(itens, tipo):
    """Lista de entregáveis para criação: o que veio no modal tem prioridade;
    sem lista e com tipo definido, copia o modelo daquele tipo."""
    if isinstance(itens, list):
        return itens
    if not tipo:
        return []
    return [{"tipo": m.tipo, "categoria": m.categoria, "peso": m.peso,
             "responsaveis": m.responsavel_padrao}
            for m in (ModeloEntregavel.query
                      .filter_by(tipo_projeto=tipo)
                      .order_by(ModeloEntregavel.ordem, ModeloEntregavel.id).all())]


@entregaveis_bp.route("/api/projetos", methods=["POST"])
@require_role("admin", "gestor")
def criar_projeto():
    data = request.get_json(silent=True) or {}
    nome = (data.get("nome") or "").strip()
    if not nome:
        return jsonify({"erro": "nome é obrigatório"}), 400
    moscow = (data.get("moscow") or "").strip()
    if moscow and moscow not in MOSCOW:
        return jsonify({"erro": f"moscow inválido. Use: {', '.join(MOSCOW)}"}), 400
    tipo = (data.get("tipo") or "").strip()
    if tipo and tipo not in TIPOS_PROJETO:
        return jsonify({"erro": f"tipo inválido. Use: {', '.join(TIPOS_PROJETO)}"}), 400
    status = (data.get("status") or "execucao").strip()
    if status not in STATUS_PROJETO:
        return jsonify({"erro": f"status inválido. Use: {', '.join(STATUS_PROJETO)}"}), 400
    orc = _parse_orcamento(data.get("orcamento"))
    if orc is None:
        return jsonify({"erro": "orçamento inválido"}), 400

    datas = {}
    for c in DATAS_PROJETO:
        v, err = _parse_data(data.get(c))
        if err:
            return jsonify({"erro": f"{c}: {err}"}), 400
        datas[c] = v
    if datas["data_inicio_prev"] and datas["data_fim_prev"] \
            and datas["data_fim_prev"] < datas["data_inicio_prev"]:
        return jsonify({"erro": "término previsto não pode ser antes do início previsto"}), 400

    p = Projeto(
        nome=nome,
        descricao=(data.get("descricao") or "").strip(),
        tipo=tipo,
        sku=(data.get("sku") or "").strip(),
        moscow=moscow,
        status=status,
        prioridade=int(data.get("prioridade") or 0),
        consumivel=bool(data.get("consumivel")),
        lancamento=(data.get("lancamento") or "").strip(),
        ano=int(data.get("ano") or datetime.now().year),
        orcamento=orc,
        **datas,
    )
    db.session.add(p)
    db.session.flush()   # garante p.id para anexar os entregáveis

    email = get_jwt_identity()
    for it in _entregaveis_do_payload(data.get("entregaveis"), tipo):
        nome_ent = (it.get("tipo") or "").strip()
        if not nome_ent:
            continue
        peso = _parse_peso(it.get("peso"))
        if peso is None:
            return jsonify({"erro": f"peso inválido em '{nome_ent}' (use número > 0)"}), 400
        ini_prev, err1 = _parse_data(it.get("data_inicio_prev"))
        fim_prev, err2 = _parse_data(it.get("data_fim_prev"))
        if err1 or err2:
            return jsonify({"erro": f"datas previstas inválidas em '{nome_ent}'"}), 400
        e = Entregavel(
            projeto_id=p.id, tipo=nome_ent,
            categoria=(it.get("categoria") or "Produto").strip(),
            status=(it.get("status") or "pendente"),
            peso=peso,
            data_inicio_prev=ini_prev, data_fim_prev=fim_prev,
            responsaveis=(it.get("responsaveis") or "").strip(),
            atualizado_por=email)
        p.entregaveis.append(e)   # coleção + sessão (o avanço lê a coleção)
        db.session.flush()
        _aplicar_responsaveis(e, it.get("responsaveis_ids"))

    # Linha de base v1 + primeira foto: o projeto já nasce com histórico.
    p.registrar_baseline(email, motivo="Linha de base inicial")
    p.registrar_snapshot()
    db.session.commit()

    _emit("PROJETO_CREATED", {"projeto": p.to_dict(), "projeto_id": p.id}, email,
          entidade=f"Projeto:{p.nome}")
    return jsonify({"projeto": p.to_dict()}), 201


@entregaveis_bp.route("/api/projetos/<int:pid>", methods=["GET"])
@require_role("admin", "gestor", "tecnico")
def detalhe_projeto(pid):
    p = _get_projeto(pid, com_pmo=True)
    user = _usuario_atual()
    role = user.role if user else ""
    if role == "tecnico" and not any(_responsavel_por(e, user) for e in p.entregaveis):
        return jsonify({"erro": "Sem acesso a este projeto"}), 403

    grupos = {c: [] for c in CATEGORIAS_ENTREGAVEL}
    extras = {}
    for e in p.entregaveis:
        (grupos if e.categoria in grupos else extras).setdefault(e.categoria, [])
        (grupos.get(e.categoria) if e.categoria in grupos
         else extras[e.categoria]).append(e.to_dict())
    categorias = [{"categoria": c, "entregaveis": grupos[c]}
                  for c in CATEGORIAS_ENTREGAVEL if grupos[c]]
    categorias += [{"categoria": c, "entregaveis": v} for c, v in extras.items()]
    d = p.to_dict(com_pmo=True, com_financeiro=pode_ver_financeiro(role))
    d["categorias"] = categorias
    return jsonify(d)


@entregaveis_bp.route("/api/projetos/<int:pid>", methods=["PUT"])
@require_role("admin", "gestor")
def editar_projeto(pid):
    p = _get_projeto(pid)
    data = request.get_json(silent=True) or {}
    email = get_jwt_identity()

    # 1) VALIDA TUDO ANTES DE ESCREVER QUALQUER COISA.
    #    Antes, log_action() commitava dentro do laço: um campo inválido no meio
    #    devolvia 400 com os campos anteriores já persistidos (update parcial).
    novos, mudancas = {}, []
    for campo in ("nome", "descricao", "tipo", "sku", "moscow", "status",
                  "lancamento", *DATAS_PROJETO):
        if campo not in data:
            continue
        novo = (data.get(campo) or "").strip()
        if campo == "nome" and not novo:
            return jsonify({"erro": "nome não pode ficar vazio"}), 400
        if campo == "moscow" and novo and novo not in MOSCOW:
            return jsonify({"erro": "moscow inválido"}), 400
        if campo == "tipo" and novo and novo not in TIPOS_PROJETO:
            return jsonify({"erro": f"tipo inválido. Use: {', '.join(TIPOS_PROJETO)}"}), 400
        if campo == "status" and novo and novo not in STATUS_PROJETO:
            return jsonify({"erro": f"status inválido. Use: {', '.join(STATUS_PROJETO)}"}), 400
        if campo in DATAS_PROJETO:
            novo, err = _parse_data(novo)
            if err:
                return jsonify({"erro": f"{campo}: {err}"}), 400
        novos[campo] = novo

    ini_prev = novos.get("data_inicio_prev", p.data_inicio_prev or "")
    fim_prev = novos.get("data_fim_prev", p.data_fim_prev or "")
    if ini_prev and fim_prev and fim_prev < ini_prev:
        return jsonify({"erro": "término previsto não pode ser antes do início previsto"}), 400

    if "orcamento" in data:
        orc = _parse_orcamento(data.get("orcamento"))
        if orc is None:
            return jsonify({"erro": "orçamento inválido"}), 400
        novos["orcamento"] = orc

    # 2) APLICA
    for campo, novo in novos.items():
        antigo = getattr(p, campo)
        antigo = antigo if campo == "orcamento" else (antigo or "")
        if novo != antigo:
            setattr(p, campo, novo)
            mudancas.append((campo, antigo, novo))
    if "prioridade" in data:
        p.prioridade = int(data.get("prioridade") or 0)
    if "consumivel" in data:
        p.consumivel = bool(data.get("consumivel"))

    # 3) Mexeu na linha de base? Versiona em vez de sobrescrever silenciosamente.
    mudou_baseline = [c for c, _, _ in mudancas if c in CAMPOS_BASELINE]
    baseline = None
    if mudou_baseline:
        motivo = (data.get("motivo_replanejamento") or "").strip() \
            or f"Replanejamento ({', '.join(mudou_baseline)})"
        baseline = p.registrar_baseline(email, motivo=motivo)

    p.registrar_snapshot()
    db.session.commit()

    for campo, antigo, novo in mudancas:
        _emit("PROJETO_UPDATED", {"projeto_id": p.id}, email, campo=campo,
              antigo=str(antigo), novo=str(novo), entidade=f"Projeto:{p.nome}")
    if baseline is not None:
        _emit("BASELINE_CREATED",
              {"projeto_id": p.id, "baseline": baseline.to_dict()}, email,
              entidade=f"Projeto:{p.nome}", campo="linha_de_base")
    _emit("PROJETO_UPDATED", {"projeto": p.to_dict(), "projeto_id": p.id}, email)
    return jsonify({"projeto": p.to_dict()})


@entregaveis_bp.route("/api/projetos/<int:pid>", methods=["DELETE"])
@require_role("admin", "gestor")
def arquivar_projeto(pid):
    """Arquiva. `status` no corpo registra POR QUE saiu (concluído × cancelado)."""
    p = _get_projeto(pid)
    data = request.get_json(silent=True) or {}
    status = (data.get("status") or "").strip()
    if status and status not in STATUS_PROJETO:
        return jsonify({"erro": f"status inválido. Use: {', '.join(STATUS_PROJETO)}"}), 400
    p.ativo = False
    if status:
        p.status = status
    elif p.status in STATUS_PROJETO_ABERTO:
        # Sem informação explícita, deduz pelo avanço em vez de deixar ambíguo.
        p.status = "concluido" if p.avanco >= 100 else "cancelado"
    if p.status == "concluido" and not (p.data_fim_real or "").strip():
        p.data_fim_real = datetime.now().date().isoformat()
    p.registrar_snapshot()
    db.session.commit()
    email = get_jwt_identity()
    _emit("PROJETO_ARQUIVADO", {"projeto": p.to_dict(), "projeto_id": p.id}, email,
          entidade=f"Projeto:{p.nome}", campo="ativo", antigo="True", novo="False")
    return jsonify({"ok": True, "projeto": p.to_dict()})


@entregaveis_bp.route("/api/projetos/<int:pid>/restaurar", methods=["POST"])
@require_role("admin", "gestor")
def restaurar_projeto(pid):
    p = _get_projeto(pid)
    p.ativo = True
    if p.status in ("concluido", "cancelado"):
        p.status = "execucao"
    p.registrar_snapshot()
    db.session.commit()
    email = get_jwt_identity()
    _emit("PROJETO_UPDATED", {"projeto": p.to_dict(), "projeto_id": p.id}, email,
          entidade=f"Projeto:{p.nome}", campo="ativo", antigo="False", novo="True")
    return jsonify({"projeto": p.to_dict()})


# ── LINHA DE BASE ────────────────────────────────────────────────────────────

@entregaveis_bp.route("/api/projetos/<int:pid>/baselines", methods=["GET"])
@require_role("admin", "gestor")
def listar_baselines(pid):
    p = _get_projeto(pid)
    return jsonify({"baselines": [b.to_dict() for b in p.baselines]})


@entregaveis_bp.route("/api/projetos/<int:pid>/baselines", methods=["POST"])
@require_role("admin", "gestor")
def criar_baseline(pid):
    p = _get_projeto(pid)
    data = request.get_json(silent=True) or {}
    motivo = (data.get("motivo") or "").strip()
    if not motivo:
        return jsonify({"erro": "informe o motivo do replanejamento"}), 400
    email = get_jwt_identity()
    b = p.registrar_baseline(email, motivo=motivo)
    db.session.commit()
    _emit("BASELINE_CREATED", {"projeto_id": p.id, "baseline": b.to_dict()}, email,
          entidade=f"Projeto:{p.nome}", campo="linha_de_base", novo=motivo)
    return jsonify({"baseline": b.to_dict()}), 201


# ── ALERTAS (o que precisa de atenção hoje) ──────────────────────────────────

@entregaveis_bp.route("/api/projetos/alertas", methods=["GET"])
@require_role("admin", "gestor", "tecnico")
def alertas():
    """Fatos acionáveis derivados do que já está no banco.

    Tipos: entregável atrasado, projeto com término previsto vencido, estouro
    de orçamento projetado e projeto sem atualização há 30+ dias.
    """
    user = _usuario_atual()
    role = user.role if user else ""
    financeiro = pode_ver_financeiro(role)
    hoje = datetime.now().date()
    dias_parado = request.args.get("dias_parado", default=30, type=int)

    itens = []
    projetos = _query_projetos().filter(
        Projeto.ativo.is_(True),
        Projeto.status.in_(STATUS_PROJETO_ABERTO)).all()

    for p in projetos:
        meus = [e for e in p.entregaveis if _responsavel_por(e, user)]
        if role == "tecnico" and not meus:
            continue
        escopo = meus if role == "tecnico" else p.entregaveis

        for e in escopo:
            if not e.atrasado:
                continue
            fim = _parse_iso(e.data_fim_prev)
            itens.append({
                "tipo": "entregavel_atrasado", "severidade": "critico",
                "projeto_id": p.id, "projeto": p.nome, "entregavel_id": e.id,
                "titulo": f"{e.tipo} atrasado",
                "detalhe": f"previsto para {fim.strftime('%d/%m/%Y')} · {(hoje - fim).days} dia(s) em atraso",
            })

        fim_p = _parse_iso(p.data_fim_prev)
        if fim_p and fim_p < hoje and p.avanco < 100:
            itens.append({
                "tipo": "projeto_vencido", "severidade": "critico",
                "projeto_id": p.id, "projeto": p.nome,
                "titulo": "Prazo do projeto vencido",
                "detalhe": f"término previsto {fim_p.strftime('%d/%m/%Y')} · avanço {p.avanco}%",
            })

        m = p.pmo_metrics()
        if financeiro and m.get("eac") and m.get("bac") and m["eac"] > m["bac"] * 1.02:
            over = m["eac"] - m["bac"]
            itens.append({
                "tipo": "estouro_orcamento", "severidade": "atencao",
                "projeto_id": p.id, "projeto": p.nome,
                "titulo": "Estouro de orçamento projetado",
                "detalhe": f"EAC R$ {over:,.0f} acima do orçado (CPI {m['cpi']:.2f})"
                           .replace(",", "."),
            })

        ults = [e.atualizado_em for e in p.entregaveis if e.atualizado_em]
        if ults:
            parado = (datetime.now() - max(ults)).days
            if parado >= dias_parado and p.avanco < 100:
                itens.append({
                    "tipo": "projeto_parado", "severidade": "atencao",
                    "projeto_id": p.id, "projeto": p.nome,
                    "titulo": "Sem movimentação",
                    "detalhe": f"nenhum entregável atualizado há {parado} dias",
                })

    ordem = {"critico": 0, "atencao": 1}
    itens.sort(key=lambda a: (ordem.get(a["severidade"], 9), a["projeto"]))
    return jsonify({"alertas": itens, "total": len(itens)})


# ── ACOMPANHAMENTO MENSAL (PMO / EVM) ────────────────────────────────────────

def _validar_pct(v, campo):
    """int 0-100 ou erro (str)."""
    try:
        n = int(v)
    except (TypeError, ValueError):
        return None, f"{campo} deve ser número"
    if not (0 <= n <= 100):
        return None, f"{campo} deve estar entre 0 e 100"
    return n, None


@entregaveis_bp.route("/api/projetos/<int:pid>/mensal", methods=["GET"])
@require_role("admin", "gestor")
def listar_mensal(pid):
    p = _get_projeto(pid)
    return jsonify({
        "projeto_id": p.id,
        "orcamento": p.orcamento or 0.0,
        "serie": p.serie_mensal(),
        "pmo": p.pmo_metrics(),
    })


@entregaveis_bp.route("/api/projetos/<int:pid>/mensal", methods=["PUT"])
@require_role("admin", "gestor")
def upsert_mensal(pid):
    """Cria ou atualiza o CUSTO de uma competência (YYYY-MM).

    O realizado (avanço) é automático, vindo da conclusão das tarefas — não se informa aqui.
    O previsto é calculado pelas datas. Este lançamento serve só ao custo (para o CPI).
    """
    p = _get_projeto(pid)
    data = request.get_json(silent=True) or {}
    comp = (data.get("competencia") or "").strip()
    if not _RE_COMPET.match(comp):
        return jsonify({"erro": "competência inválida (use AAAA-MM)"}), 400

    # Aceita 'custo_mes' (incremental, novo) ou 'custo_acumulado' (compat. antiga).
    bruto = data.get("custo_mes", data.get("custo_acumulado"))
    custo = _parse_orcamento(bruto)
    if custo is None:
        return jsonify({"erro": "custo do mês inválido"}), 400

    email = get_jwt_identity()
    reg = ProjetoMensal.query.filter_by(projeto_id=p.id, competencia=comp).first()
    acao = "MENSAL_UPDATED"
    if reg is None:
        reg = ProjetoMensal(projeto_id=p.id, competencia=comp)
        # Precisa entrar na COLEÇÃO, não só na sessão: `p.mensais` já veio
        # carregado pelo selectinload e recompute_acumulados() itera sobre ela.
        p.mensais.append(reg)
        acao = "MENSAL_CREATED"

    import calendar
    ano_c, mes_c = int(comp[:4]), int(comp[5:7])
    ultimo_dia = datetime(ano_c, mes_c, calendar.monthrange(ano_c, mes_c)[1]).date()
    reg.pct_previsto  = p.previsto_em(comp) or 0        # informativo
    reg.pct_realizado = p.realizado_em(ultimo_dia)      # informativo
    reg.custo_mes = custo
    reg.atualizado_por = email
    reg.atualizado_em = datetime.now()
    db.session.flush()           # garante que reg participe do recálculo
    p.recompute_acumulados()     # atualiza o custo acumulado (AC) de toda a série
    p.registrar_snapshot()
    db.session.commit()

    _emit(acao, {"projeto_id": p.id, "mensal": reg.to_dict(), "pmo": p.pmo_metrics()},
          email, entidade=f"{p.nome} · {comp}", campo="custo_mensal",
          novo=f"R$ {custo:.2f}")
    return jsonify({"mensal": reg.to_dict(), "pmo": p.pmo_metrics()})


@entregaveis_bp.route("/api/projetos/<int:pid>/mensal/<competencia>", methods=["DELETE"])
@require_role("admin", "gestor")
def remover_mensal(pid, competencia):
    p = _get_projeto(pid)
    reg = ProjetoMensal.query.filter_by(projeto_id=p.id, competencia=competencia).first()
    if reg is None:
        return jsonify({"erro": "lançamento não encontrado"}), 404
    p.mensais.remove(reg)        # delete-orphan cuida do DELETE; a coleção fica coerente
    db.session.flush()
    p.recompute_acumulados()     # reajusta o acumulado dos meses restantes
    p.registrar_snapshot()
    db.session.commit()
    email = get_jwt_identity()
    _emit("MENSAL_DELETED", {"projeto_id": p.id, "competencia": competencia,
                             "pmo": p.pmo_metrics()}, email,
          entidade=f"{p.nome} · {competencia}", campo="acompanhamento_mensal")
    return jsonify({"ok": True, "pmo": p.pmo_metrics()})


# ── ENTREGÁVEIS ──────────────────────────────────────────────────────────────

@entregaveis_bp.route("/api/entregaveis/responsaveis", methods=["GET"])
@require_role("admin", "gestor", "tecnico")
def listar_responsaveis():
    """Usuários atribuíveis a um entregável (picker do front)."""
    users = (User.query.filter(User.ativo.is_(True),
                               User.role.in_(("admin", "gestor", "tecnico")))
             .order_by(User.nome).all())
    return jsonify({"usuarios": [{"id": u.id, "nome": u.nome, "email": u.email,
                                  "role": u.role} for u in users]})


@entregaveis_bp.route("/api/entregaveis/<int:eid>", methods=["PUT"])
@require_role("admin", "gestor", "tecnico")
def atualizar_entregavel(eid):
    e = (Entregavel.query.options(selectinload(Entregavel.responsaveis_users))
         .filter_by(id=eid).first())
    if e is None:
        return jsonify({"erro": "entregável não encontrado"}), 404

    data = request.get_json(silent=True) or {}
    email = get_jwt_identity()
    user = _usuario_atual()
    role = user.role if user else ""
    gestor = pode_editar_projeto(role)

    # Técnico só mexe no que é dele, e só no andamento — nunca em quem é
    # responsável, no peso ou no cronograma planejado.
    if not gestor:
        if not _responsavel_por(e, user):
            return jsonify({"erro": "Você não é responsável por este entregável"}), 403
        proibidos = {"responsaveis", "responsaveis_ids", "peso",
                     "data_inicio_prev", "data_fim_prev"}
        if proibidos & set(data):
            return jsonify({"erro": "Perfil técnico só pode alterar status e percentual"}), 403

    status_antigo, pct_antigo = e.status, e.percentual
    mudancas = []
    hoje_iso = datetime.now().strftime("%Y-%m-%d")

    # datas explícitas primeiro (aceita ISO yyyy-mm-dd ou ""), para que os
    # autopreenchimentos abaixo respeitem o que o usuário informou.
    for campo in ("data_inicio", "data_conclusao", "data_inicio_prev", "data_fim_prev"):
        if campo in data:
            val, err = _parse_data(data.get(campo))
            if err:
                return jsonify({"erro": f"{campo}: {err}"}), 400
            setattr(e, campo, val)

    if "peso" in data:
        peso = _parse_peso(data.get("peso"))
        if peso is None:
            return jsonify({"erro": "peso deve ser um número maior que zero"}), 400
        if peso != e.peso:
            mudancas.append(("peso", e.peso, peso))
            e.peso = peso

    if "status" in data:
        novo = (data.get("status") or "").strip()
        if novo not in STATUS_ENTREGAVEL:
            return jsonify({"erro": f"status inválido. Use: {', '.join(STATUS_ENTREGAVEL)}"}), 400
        if novo != e.status:
            mudancas.append(("status", e.status, novo))
            e.status = novo
        if novo == "concluido":
            e.percentual = 100
            if not (e.data_conclusao or "").strip():       # conclusão = hoje (se não veio explícita)
                e.data_conclusao = hoje_iso
            if not (e.data_inicio or "").strip():          # início = conclusão (se não veio)
                e.data_inicio = e.data_conclusao
        elif novo == "na":
            e.percentual = None
            e.data_conclusao = ""                          # não se aplica: some do cálculo
        elif novo == "pendente":
            e.percentual = 0
            # A data de conclusão é PRESERVADA de propósito: apagá-la reescrevia
            # a curva-S histórica. `pct_em` já ignora o que foi reaberto.
        elif novo == "em_progresso" and not (e.data_inicio or "").strip():
            e.data_inicio = hoje_iso

    if "percentual" in data and e.status == "em_progresso":
        try:
            pct = int(data.get("percentual"))
        except (TypeError, ValueError):
            return jsonify({"erro": "percentual deve ser número"}), 400
        if not (0 <= pct <= 100):
            return jsonify({"erro": "percentual deve estar entre 0 e 100"}), 400
        if pct != e.percentual:
            mudancas.append(("percentual", e.percentual, pct))
            e.percentual = pct

    if "responsaveis" in data:
        novo = (data.get("responsaveis") or "").strip()
        if novo != (e.responsaveis or ""):
            mudancas.append(("responsaveis", e.responsaveis, novo))
            e.responsaveis = novo
    if "responsaveis_ids" in data:
        ids = data.get("responsaveis_ids") or []
        if not isinstance(ids, list):
            return jsonify({"erro": "responsaveis_ids deve ser uma lista"}), 400
        antes = sorted(u.id for u in e.responsaveis_users)
        _aplicar_responsaveis(e, ids)
        depois = sorted(u.id for u in e.responsaveis_users)
        if antes != depois:
            mudancas.append(("responsaveis_ids", str(antes), str(depois)))

    # Registra a transição: é o que permite reconstruir o passado com fidelidade.
    if e.status != status_antigo or e.percentual != pct_antigo:
        db.session.add(EntregavelHistorico(
            entregavel_id=e.id, status_antigo=status_antigo or "",
            status_novo=e.status or "", percentual=e.percentual,
            em=datetime.now(), por=email))

    e.atualizado_por = email
    e.atualizado_em = datetime.now()
    if e.projeto is not None:
        e.projeto.registrar_snapshot()
    db.session.commit()

    for campo, antigo, novo in mudancas:
        _emit("ENTREGAVEL_UPDATED", {"projeto_id": e.projeto_id, "entregavel_id": e.id},
              email, campo=campo, antigo=str(antigo), novo=str(novo),
              entidade=f"{e.projeto.nome} · {e.tipo}")
    _emit("ENTREGAVEL_UPDATED",
          {"entregavel": e.to_dict(), "projeto_id": e.projeto_id,
           "avanco_projeto": e.projeto.avanco}, email)
    return jsonify({"entregavel": e.to_dict(), "avanco_projeto": e.projeto.avanco})


@entregaveis_bp.route("/api/entregaveis/<int:eid>/historico", methods=["GET"])
@require_role("admin", "gestor", "tecnico")
def historico_entregavel(eid):
    e = Entregavel.query.get_or_404(eid)
    return jsonify({"historico": [h.to_dict() for h in e.historico]})


@entregaveis_bp.route("/api/projetos/<int:pid>/entregaveis", methods=["POST"])
@require_role("admin", "gestor")
def adicionar_entregavel(pid):
    p = _get_projeto(pid)
    data = request.get_json(silent=True) or {}
    tipo = (data.get("tipo") or "").strip()
    if not tipo:
        return jsonify({"erro": "tipo é obrigatório"}), 400
    peso = _parse_peso(data.get("peso"))
    if peso is None:
        return jsonify({"erro": "peso deve ser um número maior que zero"}), 400
    ini_prev, err1 = _parse_data(data.get("data_inicio_prev"))
    fim_prev, err2 = _parse_data(data.get("data_fim_prev"))
    if err1 or err2:
        return jsonify({"erro": err1 or err2}), 400

    categoria = (data.get("categoria") or "Produto").strip()
    email = get_jwt_identity()
    e = Entregavel(projeto_id=p.id, tipo=tipo, categoria=categoria,
                   status=(data.get("status") or "pendente"),
                   peso=peso, data_inicio_prev=ini_prev, data_fim_prev=fim_prev,
                   responsaveis=(data.get("responsaveis") or "").strip(),
                   atualizado_por=email)
    p.entregaveis.append(e)   # entra na coleção: o avanço/snapshot já contam com ele
    db.session.flush()
    _aplicar_responsaveis(e, data.get("responsaveis_ids"))
    p.registrar_snapshot()
    db.session.commit()

    _emit("ENTREGAVEL_CREATED",
          {"entregavel": e.to_dict(), "projeto_id": p.id, "avanco_projeto": p.avanco},
          email, entidade=f"{p.nome} · {tipo}")
    return jsonify({"entregavel": e.to_dict(), "avanco_projeto": p.avanco}), 201


@entregaveis_bp.route("/api/entregaveis/<int:eid>", methods=["DELETE"])
@require_role("admin", "gestor")
def excluir_entregavel(eid):
    e = Entregavel.query.get_or_404(eid)
    pid, nome, tipo, projeto = e.projeto_id, e.projeto.nome, e.tipo, e.projeto
    projeto.entregaveis.remove(e)   # delete-orphan apaga; o avanço já reflete a saída
    db.session.flush()
    projeto.registrar_snapshot()
    db.session.commit()
    email = get_jwt_identity()
    _emit("ENTREGAVEL_DELETED",
          {"entregavel_id": eid, "projeto_id": pid, "avanco_projeto": projeto.avanco},
          email, entidade=f"{nome} · {tipo}")
    return jsonify({"ok": True, "avanco_projeto": projeto.avanco})


# ── MODELOS DE ENTREGÁVEIS (templates por tipo de projeto) ───────────────────

@entregaveis_bp.route("/api/modelos", methods=["GET"])
@require_role("admin", "gestor")
def listar_modelos():
    """Itens de modelo agrupados por tipo de projeto (OEM/Revenda)."""
    tipo = (request.args.get("tipo") or "").strip()
    q = ModeloEntregavel.query
    if tipo:
        if tipo not in TIPOS_PROJETO:
            return jsonify({"erro": "tipo inválido"}), 400
        q = q.filter_by(tipo_projeto=tipo)
    itens = q.order_by(ModeloEntregavel.tipo_projeto,
                       ModeloEntregavel.ordem, ModeloEntregavel.id).all()
    out = {t: [] for t in TIPOS_PROJETO}
    for m in itens:
        out.setdefault(m.tipo_projeto, []).append(m.to_dict())
    return jsonify({"tipos": TIPOS_PROJETO, "modelos": out})


@entregaveis_bp.route("/api/modelos", methods=["POST"])
@require_role("admin", "gestor")
def adicionar_modelo():
    data = request.get_json(silent=True) or {}
    tipo_projeto = (data.get("tipo_projeto") or "").strip()
    if tipo_projeto not in TIPOS_PROJETO:
        return jsonify({"erro": f"tipo_projeto inválido. Use: {', '.join(TIPOS_PROJETO)}"}), 400
    tipo = (data.get("tipo") or "").strip()
    if not tipo:
        return jsonify({"erro": "tipo (nome do entregável) é obrigatório"}), 400
    peso = _parse_peso(data.get("peso"))
    if peso is None:
        return jsonify({"erro": "peso deve ser um número maior que zero"}), 400
    ult = (ModeloEntregavel.query.filter_by(tipo_projeto=tipo_projeto)
           .order_by(ModeloEntregavel.ordem.desc()).first())
    m = ModeloEntregavel(
        tipo_projeto=tipo_projeto,
        categoria=(data.get("categoria") or "Produto").strip(),
        tipo=tipo,
        peso=peso,
        responsavel_padrao=(data.get("responsavel_padrao") or "").strip(),
        ordem=((ult.ordem + 1) if ult else 0))
    db.session.add(m)
    db.session.commit()
    _emit("PROJETO_UPDATED", {"modelo": m.to_dict()}, get_jwt_identity(),
          entidade=f"Modelo {tipo_projeto} · {tipo}", campo="modelo_criado")
    return jsonify({"modelo": m.to_dict()}), 201


@entregaveis_bp.route("/api/modelos/<int:mid>", methods=["PUT"])
@require_role("admin", "gestor")
def editar_modelo(mid):
    m = ModeloEntregavel.query.get_or_404(mid)
    data = request.get_json(silent=True) or {}
    if "tipo" in data:
        novo = (data.get("tipo") or "").strip()
        if not novo:
            return jsonify({"erro": "tipo não pode ficar vazio"}), 400
        m.tipo = novo
    if "categoria" in data:
        m.categoria = (data.get("categoria") or "Produto").strip()
    if "responsavel_padrao" in data:
        m.responsavel_padrao = (data.get("responsavel_padrao") or "").strip()
    if "peso" in data:
        peso = _parse_peso(data.get("peso"))
        if peso is None:
            return jsonify({"erro": "peso deve ser um número maior que zero"}), 400
        m.peso = peso
    if "ordem" in data:
        try:
            m.ordem = int(data.get("ordem"))
        except (TypeError, ValueError):
            return jsonify({"erro": "ordem deve ser número"}), 400
    db.session.commit()
    _emit("PROJETO_UPDATED", {"modelo": m.to_dict()}, get_jwt_identity(),
          entidade=f"Modelo {m.tipo_projeto} · {m.tipo}", campo="modelo_editado")
    return jsonify({"modelo": m.to_dict()})


@entregaveis_bp.route("/api/modelos/<int:mid>", methods=["DELETE"])
@require_role("admin", "gestor")
def excluir_modelo(mid):
    m = ModeloEntregavel.query.get_or_404(mid)
    info = f"Modelo {m.tipo_projeto} · {m.tipo}"
    db.session.delete(m)
    db.session.commit()
    _emit("PROJETO_UPDATED", {}, get_jwt_identity(),
          entidade=info, campo="modelo_excluido")
    return jsonify({"ok": True})


# ── RESUMO ───────────────────────────────────────────────────────────────────

@entregaveis_bp.route("/api/entregaveis/resumo", methods=["GET"])
@require_role("admin", "gestor", "tecnico")
def resumo():
    user = _usuario_atual()
    role = user.role if user else ""
    projetos = _query_projetos().filter_by(ativo=True).all()
    if role == "tecnico":
        projetos = [p for p in projetos
                    if any(_responsavel_por(e, user) for e in p.entregaveis)]

    pend = conc = prog = 0
    por_resp = {}
    for p in projetos:
        for e in p.entregaveis:
            if e.status == "pendente":
                pend += 1
            elif e.status == "concluido":
                conc += 1
            elif e.status == "em_progresso":
                prog += 1
            if e.status not in ("pendente", "em_progresso"):
                continue
            item = {"projeto": p.nome, "tipo": e.tipo, "status": e.status,
                    "percentual": e.percentual, "id": e.id,
                    "atrasado": e.atrasado}
            # Uma linha por pessoa (FK). Só cai no texto livre quando o
            # entregável ainda não foi migrado — antes "Guilherme/Melk" virava
            # uma "pessoa" separada de "Melk" e a carga não fechava com ninguém.
            if e.responsaveis_users:
                for u in e.responsaveis_users:
                    por_resp.setdefault(u.nome, []).append(item)
            elif e.responsaveis:
                por_resp.setdefault(e.responsaveis, []).append(item)

    # Só projetos com escopo definido entram na média — projeto sem entregável
    # devolve avanço 0 e puxava o indicador para baixo sem significar nada.
    avancos = [p.avanco for p in projetos
               if any(e.status != "na" for e in p.entregaveis)]
    return jsonify({
        "projetos": len(projetos),
        "projetos_com_escopo": len(avancos),
        "avanco_medio": round(sum(avancos) / len(avancos)) if avancos else 0,
        "pendentes": pend, "em_progresso": prog, "concluidos": conc,
        "atrasados": sum(p.atrasados for p in projetos),
        "por_responsavel": por_resp,
    })


# ── EXPORT EXCEL ─────────────────────────────────────────────────────────────

@entregaveis_bp.route("/api/entregaveis/export", methods=["GET"])
@require_role("admin", "gestor")
def exportar_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    projetos = (_query_projetos().filter_by(ativo=True)
                .order_by(Projeto.prioridade, Projeto.nome).all())
    # união ordenada de tipos (categoria, tipo) preservando ordem de aparição
    tipos = []
    for p in projetos:
        for e in p.entregaveis:
            chave = (e.categoria, e.tipo)
            if chave not in tipos:
                tipos.append(chave)

    wb = Workbook()
    ws = wb.active
    ano = datetime.now().year
    ws.title = f"Entregáveis {ano}"
    CORES = {"concluido": "C6EFCE", "em_progresso": "FFEB9C",
             "pendente": "FFC7CE", "na": "D9D9D9"}
    cab = Font(bold=True, color="FFFFFF")
    azul = PatternFill("solid", fgColor="1F4E5F")

    headers = ["Projeto", "Status", "MoSCoW", "SKU", "Lançamento", "Avanço %"] + \
              [f"{t}\n({c})" for c, t in tipos]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = cab
        cell.fill = azul
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for i, p in enumerate(projetos, 2):
        ws.cell(row=i, column=1, value=p.nome).font = Font(bold=True)
        ws.cell(row=i, column=2, value=p.status or "")
        ws.cell(row=i, column=3, value=p.moscow)
        ws.cell(row=i, column=4, value=p.sku)
        ws.cell(row=i, column=5, value=p.lancamento)
        ws.cell(row=i, column=6, value=p.avanco)
        mapa = {(e.categoria, e.tipo): e for e in p.entregaveis}
        for j, chave in enumerate(tipos, 7):
            e = mapa.get(chave)
            if e is None:
                continue
            if e.status == "concluido":
                v = "OK"
            elif e.status == "em_progresso":
                v = f"{e.percentual or 0}%"
            elif e.status == "pendente":
                v = "Pendente"
            else:
                v = "NA"
            cell = ws.cell(row=i, column=j, value=v)
            cell.fill = PatternFill("solid", fgColor=CORES[e.status])
            cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 24
    for j in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 13
    ws.freeze_panes = "B2"

    # ── Aba PMO: o que o export não trazia (cronograma, R$, índices) ──────────
    wp = wb.create_sheet("PMO")
    cols_pmo = ["Projeto", "Status", "Tipo", "Início prev.", "Término prev.",
                "Início real", "Término real", "Previsão (velocidade)",
                "Avanço %", "Previsto %", "SPI", "CPI",
                "Orçado (BAC)", "Gasto (AC)", "Projetado (EAC)", "Desvio",
                "Entregáveis", "Atrasados"]
    for j, h in enumerate(cols_pmo, 1):
        c = wp.cell(row=1, column=j, value=h)
        c.font = cab
        c.fill = azul
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for i, p in enumerate(projetos, 2):
        m = p.pmo_metrics()
        bac, eac = m.get("bac") or 0, m.get("eac")
        vals = [p.nome, p.status or "", p.tipo or "",
                p.data_inicio_prev or "", p.data_fim_prev or "",
                p.data_inicio_real or "", p.data_fim_real or "",
                p.previsao_termino() or "",
                p.avanco, m.get("pct_previsto"), m.get("spi"), m.get("cpi"),
                bac, m.get("ac"), eac,
                (eac - bac) if (eac is not None and bac) else None,
                sum(1 for e in p.entregaveis if e.status != "na"), p.atrasados]
        for j, v in enumerate(vals, 1):
            wp.cell(row=i, column=j, value=v)
    wp.column_dimensions["A"].width = 28
    for j in range(2, len(cols_pmo) + 1):
        wp.column_dimensions[get_column_letter(j)].width = 14
    wp.freeze_panes = "B2"

    # ── Aba Entregáveis: linha a linha, para dinâmica/BI ──────────────────────
    we = wb.create_sheet("Detalhe")
    cols_det = ["Projeto", "Categoria", "Entregável", "Status", "%", "Peso",
                "Responsáveis", "Início prev.", "Término prev.",
                "Início real", "Conclusão", "Atrasado", "Atualizado em"]
    for j, h in enumerate(cols_det, 1):
        c = we.cell(row=1, column=j, value=h)
        c.font = cab
        c.fill = azul
    linha = 2
    for p in projetos:
        for e in p.entregaveis:
            nomes = ", ".join(u.nome for u in e.responsaveis_users) or (e.responsaveis or "")
            for j, v in enumerate([
                    p.nome, e.categoria or "", e.tipo, e.status or "",
                    e.percentual, e.peso if e.peso is not None else 1.0, nomes,
                    e.data_inicio_prev or "", e.data_fim_prev or "",
                    e.data_inicio or "", e.data_conclusao or "",
                    "Sim" if e.atrasado else "",
                    e.atualizado_em.strftime("%d/%m/%Y %H:%M") if e.atualizado_em else ""], 1):
                we.cell(row=linha, column=j, value=v)
            linha += 1
    we.column_dimensions["A"].width = 26
    we.column_dimensions["C"].width = 34
    we.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome = f"Projetos_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=nome,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

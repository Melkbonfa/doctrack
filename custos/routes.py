"""
custos/routes.py — Blueprint do módulo Custos (formação de custo de produto).

Montado sob /custos no servidor mestre. Reutiliza autenticação, audit log e
event_bus do mestre.

## Acesso

Custo é dinheiro: todas as rotas exigem **admin ou gestor** — o mesmo corte de
`pode_ver_financeiro()` em entregaveis.py e da exportação de Projetos. Técnico e
leitura não enxergam valor, e aqui não existe versão "despida" da API como em
Projetos: lá o dinheiro é um adorno de uma tela que técnico precisa usar; aqui o
valor *é* o conteúdo.

O gate é `@require_role`, não `@require_area("pde")`, para ficar igual aos
módulos vizinhos da área (Documentos, Projetos, Missões, Equipamentos). Só o PDR
usa `require_area`. Trocar isso aqui trancaria todo gestor cujo `users.areas`
ainda está vazio — que é o estado da maioria — sem ganho real de segurança, já
que o perfil já limita a gestão.

## Rotas

  GET    /custos/                                  página
  GET    /custos/api/meta                          vocabulários
  GET    /custos/api/composicoes                   lista (filtros: q, status, tipo)
  POST   /custos/api/composicoes                   cria (com lançamentos padrão)
  GET    /custos/api/composicoes/<id>              detalhe + lançamentos + cálculo
  PUT    /custos/api/composicoes/<id>              atualiza (versiona a baseline)
  DELETE /custos/api/composicoes/<id>              arquiva
  POST   /custos/api/composicoes/<id>/versoes      congela versão manualmente
  POST   /custos/api/composicoes/<id>/lancamentos  cria lançamento
  PUT    /custos/api/lancamentos/<id>              atualiza lançamento
  DELETE /custos/api/lancamentos/<id>              exclui lançamento
  GET    /custos/api/portfolio                     comparativo entre composições
  GET    /custos/api/saude                         diagnóstico (índice + checks)
  GET    /custos/api/cotacoes                      série de câmbio
  POST   /custos/api/cotacoes                      registro manual
  POST   /custos/api/cotacoes/sincronizar          força a busca no BCB
  GET    /custos/api/export/composicoes.csv        CSV (`;` + BOM)
  GET    /custos/api/export/custos.xlsx            XLSX (3 abas)
"""
import csv
import io
import json
from datetime import date, datetime

from flask import Blueprint, jsonify, render_template, request, send_file
from flask_jwt_extended import get_jwt_identity

from models import db, User, Projeto
from auth import require_role, log_action, get_client_ip
from event_bus import EventType
# Parser de dinheiro em formato BR ("R$ 1.234,56"). Reusado de propósito: era a
# quarta cópia possível dessa função no repositório.
from entregaveis import _parse_orcamento as parse_valor_br

from . import cambio
from .core import calcular, diagnostico, recalcular_linha
from .models import (
    Composicao, Lancamento, Cotacao, Versao,
    CATEGORIAS_CUSTO, NATUREZAS, NATUREZAS_LABELS, TIPOS_CALCULO,
    PERFIS_HORA, PERFIS_HORA_LABELS, PROCEDENCIAS, PROCEDENCIAS_LABELS,
    CONFIANCAS, STATUS_COMPOSICAO, TIPOS_COMPOSICAO, MOEDAS, MOEDAS_ESTRANGEIRAS,
    INCOTERMS, TIPOS_COTACAO, CAMPOS_BASELINE, LIMITE_DESVIO_CAMBIO,
    lancamentos_padrao,
)

custos_bp = Blueprint(
    "custos", __name__,
    url_prefix="/custos",
    template_folder="templates",
)

# CSS/JS ficam em static/ do mestre (1º nível) de propósito: _static_version()
# só varre esse nível, então o cache-busting funciona sem token próprio. O PDR
# precisou de um ASSET_V só porque seus assets ficam dentro do pacote.

_rt = {"socketio": None, "publish_event": None, "AuditLog": None, "EventType": None}


def init_realtime(socketio, publish_event, AuditLog, EventType):
    _rt.update(socketio=socketio, publish_event=publish_event,
               AuditLog=AuditLog, EventType=EventType)


def current_user():
    return User.query.filter_by(email=get_jwt_identity()).first()


def _emit(event_type, payload):
    if _rt["socketio"] and _rt["publish_event"]:
        u = current_user()
        try:
            _rt["publish_event"](event_type, payload,
                                 user_id=u.id if u else None,
                                 user_email=u.email if u else None,
                                 db=db, AuditLog=_rt["AuditLog"],
                                 socketio=_rt["socketio"])
        except Exception:
            pass  # tempo real é best-effort; a gravação já foi feita


# Custo é dinheiro — gestão para cima, sempre. Um só nome para o gate, para que
# afrouxá-lo por engano numa rota isolada seja visível na revisão.
def so_gestao(fn):
    return require_role("admin", "gestor")(fn)


# ── HELPERS ───────────────────────────────────────────────────────────────────

def _hoje_iso():
    return date.today().isoformat()


def _proximo_codigo():
    ano = date.today().year
    prefixo = f"CC-{ano}-"
    ultimo = (Composicao.query.filter(Composicao.codigo.like(f"{prefixo}%"))
              .order_by(Composicao.codigo.desc()).first())
    seq = 1
    if ultimo and ultimo.codigo:
        try:
            seq = int(ultimo.codigo.rsplit("-", 1)[1]) + 1
        except (ValueError, IndexError):
            seq = 1
    return f"{prefixo}{seq:03d}"


def _num(v, default=None):
    """Número livre (taxas, alíquotas, horas). None quando inválido."""
    if v is None or v == "":
        return default
    try:
        return float(str(v).replace(",", "."))
    except (TypeError, ValueError):
        return None


def _valida_escolha(valor, opcoes, campo):
    if valor and valor not in opcoes:
        return f"{campo} inválido (use um de: {', '.join(opcoes)})"
    return None


def _congelar(comp, motivo, autor, inicial=False):
    """Grava uma Versao com o cálculo do momento. A v1 é o baseline do estimado.

    Fora da abertura, o número parte do maior entre as versões já gravadas e o
    `versao` da própria composição. Sem essa âncora, uma composição criada fora
    desta API (seed, importação, migração) — que nasce com `versao=1` e nenhuma
    linha em `custo_versoes` — teria a sua primeira alteração gravada como "v1",
    como se a mudança fosse o baseline.
    """
    numero = 1 if inicial else max([v.numero for v in comp.versoes] + [comp.versao or 0]) + 1
    snap = comp.to_dict(com_lancamentos=True)
    snap["calculo"] = calcular(comp)
    v = Versao(composicao_id=comp.id, numero=numero, motivo=motivo or "",
               criado_por=autor or "", snapshot_json=json.dumps(snap, ensure_ascii=False))
    db.session.add(v)
    comp.versao = numero
    return v


def _referencia():
    """(valor USD, data ISO) — a referência principal, usada nos alertas."""
    valor, dia = cambio.referencia("USD")
    return valor, (dia.isoformat() if dia else None)


def _referencias():
    """Última cotação de cada moeda estrangeira, para o painel de câmbio."""
    out = {}
    for m in MOEDAS_ESTRANGEIRAS:
        valor, dia = cambio.referencia(m)
        out[m] = {"valor": valor, "data": dia.isoformat() if dia else None}
    return out


def _composicoes_filtradas():
    """Query compartilhada entre listagem e exports — o filtro tem que ser o mesmo.

    Devolve (query, erro). `erro` já é a resposta 400 pronta, como em
    `_filtrar_projetos` de entregaveis.py.
    """
    q = Composicao.query.filter_by(ativo=True)
    status = (request.args.get("status") or "").strip()
    if status:
        if status not in STATUS_COMPOSICAO:
            return None, (jsonify({"erro": f"status inválido: {status}"}), 400)
        q = q.filter_by(status=status)
    tipo = (request.args.get("tipo") or "").strip()
    if tipo:
        if tipo not in TIPOS_COMPOSICAO:
            return None, (jsonify({"erro": f"tipo inválido: {tipo}"}), 400)
        q = q.filter_by(tipo=tipo)
    termo = (request.args.get("q") or "").strip()
    if termo:
        like = f"%{termo}%"
        q = q.filter(db.or_(Composicao.produto.ilike(like),
                            Composicao.sku.ilike(like),
                            Composicao.codigo.ilike(like),
                            Composicao.fornecedor.ilike(like)))
    return q.order_by(Composicao.codigo.desc()), None


# ── PÁGINA ────────────────────────────────────────────────────────────────────
@custos_bp.route("/")
@custos_bp.route("")
def index():
    # A página valida o token no front (igual ao PDR e ao hub); o acesso aos
    # dados é barrado em cada rota /custos/api/*.
    from servidor import _static_version
    return render_template("custos/dashboard.html", asset_v=_static_version())


# ── META ──────────────────────────────────────────────────────────────────────
@custos_bp.route("/api/meta", methods=["GET"])
@so_gestao
def api_meta():
    return jsonify({
        "categorias": CATEGORIAS_CUSTO,
        "naturezas": NATUREZAS,
        "naturezas_labels": NATUREZAS_LABELS,
        "tipos_calculo": TIPOS_CALCULO,
        "perfis_hora": PERFIS_HORA,
        "perfis_hora_labels": PERFIS_HORA_LABELS,
        "procedencias": PROCEDENCIAS,
        "procedencias_labels": PROCEDENCIAS_LABELS,
        "confiancas": CONFIANCAS,
        "status": STATUS_COMPOSICAO,
        "tipos": TIPOS_COMPOSICAO,
        "moedas": MOEDAS,
        "moedas_estrangeiras": MOEDAS_ESTRANGEIRAS,
        "incoterms": INCOTERMS,
        "limite_desvio_cambio": LIMITE_DESVIO_CAMBIO,
        "projetos": [{"id": p.id, "nome": p.nome}
                     for p in Projeto.query.filter_by(ativo=True)
                     .order_by(Projeto.nome).all()],
    })


# ── COMPOSIÇÕES ───────────────────────────────────────────────────────────────
@custos_bp.route("/api/composicoes", methods=["GET"])
@so_gestao
def listar_composicoes():
    q, erro = _composicoes_filtradas()
    if erro:
        return erro
    itens = []
    for c in q.all():
        d = c.to_dict()
        d["calculo"] = calcular(c)
        itens.append(d)
    ref, ref_data = _referencia()
    return jsonify({"composicoes": itens, "total": len(itens),
                    "referencia": ref, "referencia_data": ref_data,
                    "referencias": _referencias()})


@custos_bp.route("/api/composicoes/<int:cid>", methods=["GET"])
@so_gestao
def obter_composicao(cid):
    c = Composicao.query.get_or_404(cid)
    d = c.to_dict(com_lancamentos=True, com_calculo=True)
    d["versoes"] = [v.to_dict() for v in c.versoes]
    ref, ref_data = _referencia()
    d["referencia"] = ref
    d["referencia_data"] = ref_data
    return jsonify(d)


@custos_bp.route("/api/composicoes", methods=["POST"])
@so_gestao
def criar_composicao():
    dados = request.get_json(silent=True) or {}
    produto = (dados.get("produto") or "").strip()
    if not produto:
        return jsonify({"erro": "produto é obrigatório"}), 400

    for campo, opcoes in (("tipo", TIPOS_COMPOSICAO), ("status", STATUS_COMPOSICAO),
                          ("moeda_base", MOEDAS), ("incoterm", INCOTERMS)):
        e = _valida_escolha((dados.get(campo) or "").strip(), opcoes, campo)
        if e:
            return jsonify({"erro": e}), 400

    taxa = _num(dados.get("taxa_planejamento"), 1.0)
    if taxa is None or taxa <= 0:
        return jsonify({"erro": "taxa_planejamento deve ser maior que zero"}), 400
    fob = parse_valor_br(dados.get("valor_fob"))
    if fob is None:
        return jsonify({"erro": "valor_fob inválido"}), 400
    preco = dados.get("preco_venda")
    preco = None if preco in (None, "") else parse_valor_br(preco)
    if preco is None and dados.get("preco_venda") not in (None, ""):
        return jsonify({"erro": "preco_venda inválido"}), 400

    u = current_user()
    c = Composicao(
        codigo=_proximo_codigo(),
        produto=produto,
        sku=(dados.get("sku") or "").strip(),
        projeto_id=dados.get("projeto_id") or None,
        equipamento_id=dados.get("equipamento_id") or None,
        fornecedor=(dados.get("fornecedor") or "").strip(),
        tipo=(dados.get("tipo") or "OEM").strip(),
        incoterm=(dados.get("incoterm") or "FOB").strip(),
        moeda_base=(dados.get("moeda_base") or "USD").strip(),
        status=(dados.get("status") or "rascunho").strip(),
        valor_fob=fob,
        qtd_invoice=int(_num(dados.get("qtd_invoice"), 1) or 1),
        volume_projetado=max(1, int(_num(dados.get("volume_projetado"), 1) or 1)),
        preco_venda=preco,
        custo_hora_engenharia=parse_valor_br(dados.get("custo_hora_engenharia")) or 0,
        custo_hora_producao=parse_valor_br(dados.get("custo_hora_producao")) or 0,
        reserva_cambial_pct=_num(dados.get("reserva_cambial_pct"), 10) or 0,
        taxa_planejamento=taxa,
        taxa_planejamento_data=(dados.get("taxa_planejamento_data") or _hoje_iso()),
        taxa_planejamento_autor=(u.nome if u else ""),
        taxa_planejamento_justificativa=(dados.get("taxa_planejamento_justificativa") or "").strip(),
        observacoes=(dados.get("observacoes") or "").strip(),
        criado_por=(u.email if u else ""),
    )
    db.session.add(c)
    db.session.flush()

    # Sem lançamentos, uma composição nova é uma folha em branco e alguém acaba
    # digitando a estrutura de importação de novo, diferente da anterior.
    if dados.get("com_padrao", True):
        for i, base in enumerate(lancamentos_padrao(c.tipo)):
            db.session.add(Lancamento(composicao_id=c.id, ordem=i, **base))
    db.session.flush()
    for l in c.lancamentos:
        recalcular_linha(l, c)

    _congelar(c, "Abertura da composição", u.email if u else "", inicial=True)
    db.session.commit()

    log_action(u.email if u else "", "CREATE", entidade=f"Composicao #{c.id}",
               campo="produto", novo=c.produto, ip=get_client_ip())
    _emit(EventType.CUSTO_COMPOSICAO_CREATED, {"id": c.id, "produto": c.produto})
    return jsonify(c.to_dict(com_lancamentos=True, com_calculo=True)), 201


@custos_bp.route("/api/composicoes/<int:cid>", methods=["PUT"])
@so_gestao
def atualizar_composicao(cid):
    c = Composicao.query.get_or_404(cid)
    dados = request.get_json(silent=True) or {}
    u = current_user()

    for campo, opcoes in (("tipo", TIPOS_COMPOSICAO), ("status", STATUS_COMPOSICAO),
                          ("moeda_base", MOEDAS), ("incoterm", INCOTERMS)):
        if campo in dados:
            e = _valida_escolha((dados.get(campo) or "").strip(), opcoes, campo)
            if e:
                return jsonify({"erro": e}), 400

    # Foto dos campos de baseline antes de mexer: se algum mudar, versiona.
    antes = {k: getattr(c, k) for k in CAMPOS_BASELINE}

    texto = ("produto", "sku", "fornecedor", "tipo", "incoterm", "moeda_base",
             "status", "di_numero", "di_data", "observacoes",
             "taxa_planejamento_justificativa")
    for k in texto:
        if k in dados:
            setattr(c, k, (dados.get(k) or "").strip())

    if "projeto_id" in dados:
        c.projeto_id = dados["projeto_id"] or None
    if "equipamento_id" in dados:
        c.equipamento_id = dados["equipamento_id"] or None

    for k in ("valor_fob", "custo_hora_engenharia", "custo_hora_producao"):
        if k in dados:
            v = parse_valor_br(dados[k])
            if v is None:
                return jsonify({"erro": f"{k} inválido"}), 400
            setattr(c, k, v)

    if "preco_venda" in dados:
        pv = dados["preco_venda"]
        if pv in (None, ""):
            c.preco_venda = None
        else:
            v = parse_valor_br(pv)
            if v is None:
                return jsonify({"erro": "preco_venda inválido"}), 400
            c.preco_venda = v

    for k, minimo in (("taxa_planejamento", 0.000001), ("reserva_cambial_pct", 0),
                      ("taxa_realizada", 0.000001)):
        if k in dados:
            if k == "taxa_realizada" and dados[k] in (None, ""):
                c.taxa_realizada = None
                continue
            v = _num(dados[k])
            if v is None or v < minimo:
                return jsonify({"erro": f"{k} inválido"}), 400
            setattr(c, k, v)

    for k in ("qtd_invoice", "volume_projetado"):
        if k in dados:
            v = _num(dados[k])
            if v is None or v < 1:
                return jsonify({"erro": f"{k} deve ser >= 1"}), 400
            setattr(c, k, int(v))

    if "taxa_planejamento" in dados:
        c.taxa_planejamento_data = dados.get("taxa_planejamento_data") or _hoje_iso()
        c.taxa_planejamento_autor = u.nome if u else ""

    mudou = [k for k in CAMPOS_BASELINE
             if str(antes[k]) != str(getattr(c, k))]
    db.session.flush()
    for l in c.lancamentos:
        recalcular_linha(l, c)

    if mudou:
        _congelar(c, "Alteração de " + ", ".join(mudou), u.email if u else "")
    db.session.commit()

    log_action(u.email if u else "", "UPDATE", entidade=f"Composicao #{c.id}",
               campo=",".join(mudou) or "dados", ip=get_client_ip())
    _emit(EventType.CUSTO_COMPOSICAO_UPDATED, {"id": c.id, "produto": c.produto})
    return jsonify(c.to_dict(com_lancamentos=True, com_calculo=True))


@custos_bp.route("/api/composicoes/<int:cid>", methods=["DELETE"])
@so_gestao
def arquivar_composicao(cid):
    c = Composicao.query.get_or_404(cid)
    c.ativo = False
    c.status = "arquivada"
    db.session.commit()
    u = current_user()
    log_action(u.email if u else "", "DELETE", entidade=f"Composicao #{c.id}",
               campo="ativo", antigo="True", novo="False", ip=get_client_ip())
    _emit(EventType.CUSTO_COMPOSICAO_DELETED, {"id": c.id})
    return jsonify({"ok": True, "id": c.id})


@custos_bp.route("/api/composicoes/<int:cid>/versoes", methods=["POST"])
@so_gestao
def criar_versao(cid):
    c = Composicao.query.get_or_404(cid)
    dados = request.get_json(silent=True) or {}
    u = current_user()
    v = _congelar(c, (dados.get("motivo") or "Congelamento manual").strip(),
                  u.email if u else "")
    db.session.commit()
    _emit(EventType.CUSTO_VERSAO_CREATED, {"composicao_id": c.id, "numero": v.numero})
    return jsonify(v.to_dict()), 201


# ── LANÇAMENTOS ───────────────────────────────────────────────────────────────

def _aplicar_lancamento(l, dados, comp):
    """Whitelist de campos editáveis. Devolve mensagem de erro ou None."""
    for campo, opcoes in (("natureza", NATUREZAS), ("categoria", CATEGORIAS_CUSTO),
                          ("tipo_calculo", TIPOS_CALCULO), ("moeda", MOEDAS),
                          ("procedencia", PROCEDENCIAS), ("confianca", CONFIANCAS),
                          ("perfil_hora", PERFIS_HORA)):
        if campo in dados:
            valor = (dados.get(campo) or "").strip()
            e = _valida_escolha(valor, opcoes, campo)
            if e:
                return e
            setattr(l, campo, valor)

    for k in ("subcategoria", "descricao", "observacao", "realizado_data", "realizado_doc"):
        if k in dados:
            setattr(l, k, (dados.get(k) or "").strip())

    if "aplicavel" in dados:
        l.aplicavel = bool(dados["aplicavel"])
    if "ordem" in dados:
        l.ordem = int(_num(dados["ordem"], 0) or 0)

    for k in ("valor_moeda", "horas", "aliquota"):
        if k in dados:
            v = parse_valor_br(dados[k]) if k == "valor_moeda" else _num(dados[k], 0)
            if v is None or v < 0:
                return f"{k} inválido"
            setattr(l, k, v)

    if "realizado_valor_brl" in dados:
        rv = dados["realizado_valor_brl"]
        if rv in (None, ""):
            l.realizado_valor_brl = None
        else:
            v = parse_valor_br(rv)
            if v is None:
                return "realizado_valor_brl inválido"
            l.realizado_valor_brl = v

    if l.tipo_calculo == "horas" and not l.perfil_hora:
        l.perfil_hora = "eng"
    recalcular_linha(l, comp)
    return None


@custos_bp.route("/api/composicoes/<int:cid>/lancamentos", methods=["POST"])
@so_gestao
def criar_lancamento(cid):
    c = Composicao.query.get_or_404(cid)
    dados = request.get_json(silent=True) or {}
    if not (dados.get("subcategoria") or "").strip():
        return jsonify({"erro": "subcategoria é obrigatória"}), 400

    ordem = max([l.ordem or 0 for l in c.lancamentos], default=-1) + 1
    l = Lancamento(composicao_id=c.id, ordem=ordem,
                   natureza="cogs", categoria=CATEGORIAS_CUSTO[1],
                   tipo_calculo="montante", moeda="BRL",
                   procedencia="estimativa", confianca="media")
    erro = _aplicar_lancamento(l, dados, c)
    if erro:
        return jsonify({"erro": erro}), 400
    db.session.add(l)
    db.session.commit()

    u = current_user()
    log_action(u.email if u else "", "CREATE", entidade=f"Lancamento #{l.id}",
               campo="subcategoria", novo=l.subcategoria, ip=get_client_ip())
    _emit(EventType.CUSTO_LANCAMENTO_CREATED, {"composicao_id": c.id, "id": l.id})
    return jsonify({"lancamento": l.to_dict(), "calculo": calcular(c)}), 201


@custos_bp.route("/api/lancamentos/<int:lid>", methods=["PUT"])
@so_gestao
def atualizar_lancamento(lid):
    l = Lancamento.query.get_or_404(lid)
    c = l.composicao
    dados = request.get_json(silent=True) or {}
    antigo = float(l.valor_brl or 0)
    erro = _aplicar_lancamento(l, dados, c)
    if erro:
        return jsonify({"erro": erro}), 400
    db.session.commit()

    u = current_user()
    log_action(u.email if u else "", "UPDATE", entidade=f"Lancamento #{l.id}",
               campo="valor_brl", antigo=antigo, novo=float(l.valor_brl or 0),
               ip=get_client_ip())
    _emit(EventType.CUSTO_LANCAMENTO_UPDATED, {"composicao_id": c.id, "id": l.id})
    return jsonify({"lancamento": l.to_dict(), "calculo": calcular(c)})


@custos_bp.route("/api/lancamentos/<int:lid>", methods=["DELETE"])
@so_gestao
def excluir_lancamento(lid):
    l = Lancamento.query.get_or_404(lid)
    c = l.composicao
    nome = l.subcategoria
    db.session.delete(l)
    db.session.commit()

    u = current_user()
    log_action(u.email if u else "", "DELETE", entidade=f"Lancamento #{lid}",
               campo="subcategoria", antigo=nome, ip=get_client_ip())
    _emit(EventType.CUSTO_LANCAMENTO_DELETED, {"composicao_id": c.id, "id": lid})
    return jsonify({"ok": True, "calculo": calcular(c)})


# ── PORTFÓLIO E SAÚDE ─────────────────────────────────────────────────────────
@custos_bp.route("/api/portfolio", methods=["GET"])
@so_gestao
def portfolio():
    """Comparativo entre composições e agregado por categoria."""
    comps = Composicao.query.filter_by(ativo=True).all()
    itens, por_categoria = [], {}
    for c in comps:
        calc = calcular(c)
        itens.append({
            "id": c.id, "codigo": c.codigo, "produto": c.produto, "sku": c.sku,
            "tipo": c.tipo, "status": c.status, "fornecedor": c.fornecedor,
            "qtd_invoice": c.qtd_invoice, "valor_fob": float(c.valor_fob or 0),
            "custo_unitario": calc["custo_unitario"],
            "nre_unitario": calc["nre_unitario"],
            # Efetivo, não realizado: a barra do comparativo precisa de um valor
            # mesmo antes de a DI chegar, senão composições novas somem do gráfico.
            "cogs_efetivo": calc["cogs_efetivo"],
            "cogs_realizado": calc["cogs_realizado"],
            "preco_venda": calc["preco_venda"],
            "margem_pct": calc["margem_pct"],
            "desvio": calc["desvio"],
        })
        for l in c.lancamentos:
            if not l.ativo or not l.aplicavel:
                continue
            linha = next((x for x in calc["linhas"] if x["id"] == l.id), None)
            if not linha:
                continue
            v = linha["realizado"] if linha["realizado"] is not None else linha["orcado"]
            if v:
                por_categoria[l.categoria] = round(por_categoria.get(l.categoria, 0) + v, 2)

    itens.sort(key=lambda x: x["custo_unitario"] or 0, reverse=True)
    total = sum(por_categoria.values())
    categorias = [{"categoria": k, "valor": v,
                   "pct": round(v / total, 4) if total else 0}
                  for k, v in sorted(por_categoria.items(), key=lambda kv: -kv[1])]
    return jsonify({"itens": itens, "categorias": categorias, "total": total})


@custos_bp.route("/api/saude", methods=["GET"])
@so_gestao
def saude():
    comps = Composicao.query.filter_by(ativo=True).all()
    ref, ref_data = _referencia()
    return jsonify(diagnostico(comps, referencia=ref, referencia_data=ref_data,
                               hoje_iso=_hoje_iso()))


# ── COTAÇÕES ──────────────────────────────────────────────────────────────────
@custos_bp.route("/api/cotacoes", methods=["GET"])
@so_gestao
def listar_cotacoes():
    q = Cotacao.query
    moeda = (request.args.get("moeda") or "").strip().upper()
    if moeda:
        if moeda not in MOEDAS_ESTRANGEIRAS:
            return jsonify({"erro": f"moeda inválida: {moeda}"}), 400
        q = q.filter_by(moeda=moeda)
    limite = min(int(_num(request.args.get("limite"), 120) or 120), 1000)
    itens = q.order_by(Cotacao.data.desc()).limit(limite).all()
    ref, ref_data = _referencia()
    return jsonify({
        "cotacoes": [c.to_dict() for c in reversed(itens)],
        "referencia": ref, "referencia_data": ref_data,
        "habilitado": cambio.habilitado(),
    })


@custos_bp.route("/api/cotacoes", methods=["POST"])
@so_gestao
def registrar_cotacao():
    """Registro manual — o caminho que mantém o módulo utilizável sem internet."""
    dados = request.get_json(silent=True) or {}
    moeda = (dados.get("moeda") or "").strip().upper()
    if moeda not in MOEDAS_ESTRANGEIRAS:
        return jsonify({"erro": "moeda deve ser USD ou EUR"}), 400
    valor = _num(dados.get("valor"))
    if valor is None or valor <= 0:
        return jsonify({"erro": "valor deve ser maior que zero"}), 400
    tipo = (dados.get("tipo") or "ptax_venda").strip()
    if tipo not in TIPOS_COTACAO:
        return jsonify({"erro": f"tipo inválido: {tipo}"}), 400
    try:
        dia = (datetime.strptime(dados["data"], "%Y-%m-%d").date()
               if dados.get("data") else date.today())
    except (ValueError, TypeError):
        return jsonify({"erro": "data inválida (use AAAA-MM-DD)"}), 400

    c = Cotacao.query.filter_by(moeda=moeda, data=dia, tipo=tipo).first()
    if c:
        c.valor, c.fonte, c.obtido_em = valor, "manual", datetime.now()
    else:
        c = Cotacao(moeda=moeda, data=dia, tipo=tipo, valor=valor, fonte="manual")
        db.session.add(c)
    db.session.commit()
    u = current_user()
    log_action(u.email if u else "", "CREATE", entidade="Cotacao",
               campo=f"{moeda} {dia}", novo=valor, ip=get_client_ip())
    return jsonify(c.to_dict()), 201


@custos_bp.route("/api/cotacoes/sincronizar", methods=["POST"])
@so_gestao
def sincronizar_cotacoes():
    resultado = cambio.sincronizar()
    ref, ref_data = _referencia()
    return jsonify({"resultado": resultado, "referencia": ref,
                    "referencia_data": ref_data})


# ── EXPORTS ───────────────────────────────────────────────────────────────────
# Convenção da casa (PR #30): CSV com `;` e BOM utf-8, nome datado vindo do
# servidor, e o export respeita os mesmos filtros da listagem.

_CAB_COMPOSICOES = [
    "Codigo", "Produto", "SKU", "Projeto", "Tipo", "Status", "Fornecedor",
    "Incoterm", "Moeda", "Taxa planejamento", "Taxa realizada", "Qtd invoice",
    "Volume projetado", "NRE (R$)", "COGS orcado (R$)", "COGS realizado (R$)",
    "Desvio (R$)", "Custo unitario (R$)", "Preco venda (R$)", "Margem (%)",
]


def _ou_vazio(v):
    """None vira célula vazia — em planilha, zero e 'não medido' não são a mesma coisa."""
    return "" if v is None else v


def _linha_composicao(c, calc):
    return [
        c.codigo, c.produto, c.sku, c.projeto.nome if c.projeto else "",
        c.tipo, c.status, c.fornecedor, c.incoterm, c.moeda_base,
        float(c.taxa_planejamento or 0),
        float(c.taxa_realizada) if c.taxa_realizada is not None else "",
        c.qtd_invoice, c.volume_projetado,
        calc["nre_realizado"], calc["cogs_orcado"], _ou_vazio(calc["cogs_realizado"]),
        _ou_vazio(calc["desvio"]), calc["custo_unitario"],
        _ou_vazio(calc["preco_venda"]),
        round(calc["margem_pct"] * 100, 1) if calc["margem_pct"] is not None else "",
    ]


@custos_bp.route("/api/export/composicoes.csv", methods=["GET"])
@so_gestao
def exportar_csv():
    q, erro = _composicoes_filtradas()
    if erro:
        return erro
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow(_CAB_COMPOSICOES)
    for c in q.all():
        w.writerow(_linha_composicao(c, calcular(c)))
    out = io.BytesIO(buf.getvalue().encode("utf-8-sig"))
    return send_file(out, mimetype="text/csv", as_attachment=True,
                     download_name=f"custos_composicoes_{datetime.now():%Y%m%d}.csv")


@custos_bp.route("/api/export/custos.xlsx", methods=["GET"])
@so_gestao
def exportar_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    q, erro = _composicoes_filtradas()
    if erro:
        return erro
    comps = q.all()

    wb = Workbook()
    cab_fonte = Font(bold=True, color="FFFFFF")
    cab_fundo = PatternFill("solid", fgColor="1F4E5F")

    def nova_aba(titulo, cabecalho, larguras, primeira=False):
        ws = wb.active if primeira else wb.create_sheet()
        ws.title = titulo
        ws.append(cabecalho)
        for cel in ws[1]:
            cel.font, cel.fill = cab_fonte, cab_fundo
        for letra, larg in larguras.items():
            ws.column_dimensions[letra].width = larg
        ws.freeze_panes = "A2"
        return ws

    ws = nova_aba("Composicoes", _CAB_COMPOSICOES,
                  {"A": 15, "B": 32, "C": 13, "D": 24, "G": 16}, primeira=True)
    for c in comps:
        ws.append(_linha_composicao(c, calcular(c)))

    wl = nova_aba("Lancamentos", [
        "Codigo", "Produto", "Natureza", "Categoria", "Subcategoria", "Descricao",
        "Aplicavel", "Forma", "Moeda", "Valor moeda", "Aliquota (%)",
        "Procedencia", "Confianca", "Orcado (R$)", "Realizado (R$)", "Desvio (R$)",
        "Taxa aplicada", "Observacao",
    ], {"A": 15, "B": 30, "D": 20, "E": 32, "F": 34, "R": 32})
    for c in comps:
        calc = calcular(c)
        por_id = {x["id"]: x for x in calc["linhas"]}
        for l in c.lancamentos:
            if not l.ativo:
                continue
            x = por_id.get(l.id, {})
            wl.append([
                c.codigo, c.produto, l.natureza.upper(), l.categoria, l.subcategoria,
                l.descricao, "Sim" if l.aplicavel else "Nao", l.tipo_calculo,
                l.moeda, float(l.valor_moeda or 0), float(l.aliquota or 0),
                PROCEDENCIAS_LABELS.get(l.procedencia, l.procedencia), l.confianca,
                x.get("orcado"), x.get("realizado"), x.get("desvio"),
                float(l.taxa_aplicada) if l.taxa_aplicada is not None else "",
                l.observacao,
            ])

    wc = nova_aba("Cotacoes", ["Data", "Moeda", "Tipo", "Valor", "Fonte", "Obtido em"],
                  {"A": 13, "F": 20})
    for c in (Cotacao.query.order_by(Cotacao.data.desc()).limit(400).all()):
        wc.append([c.data.strftime("%d/%m/%Y") if c.data else "", c.moeda, c.tipo,
                   float(c.valor), c.fonte,
                   c.obtido_em.strftime("%d/%m/%Y %H:%M") if c.obtido_em else ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True,
        download_name=f"custos_{datetime.now():%Y%m%d}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
